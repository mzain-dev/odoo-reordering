import base64
import io
from datetime import date, datetime, timedelta
from unittest.mock import patch

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

from odoo import fields
from odoo.exceptions import AccessError, UserError, ValidationError
from odoo.tests import tagged
from odoo.tests.common import TransactionCase


@tagged('post_install', '-at_install')
class TestReorderCalculations(TransactionCase):
    """Pure-function unit tests for the static calculation helpers.
    No DB fixtures needed beyond a config record for ABC thresholds.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.env.company.id,
            'abc_a_threshold': 5.0,
            'abc_b_threshold': 1.0,
        })

    def test_round_to_moq(self):
        m = self.Suggestion._round_to_moq
        self.assertEqual(m(0, 10), 0.0)
        self.assertEqual(m(5, 10), 10.0)
        self.assertEqual(m(10, 10), 10.0)
        self.assertEqual(m(11, 10), 11.0)
        self.assertEqual(m(5, 0), 5.0, 'MOQ <= 0 should pass qty through unmodified')
        self.assertEqual(m(-5, 10), 0.0, 'negative qty must never round to a negative number')

    def test_calc_months_of_stock(self):
        m = self.Suggestion._calc_months_of_stock
        # Rule 1: Net available <= 0 returns 0.0
        self.assertEqual(m(-5.0, 10.0), 0.0)
        self.assertEqual(m(0.0, 10.0), 0.0)
        self.assertEqual(m(-1.0, 0.0), 0.0)
        # Rule 2: Net available > 0 with positive demand returns net available / average monthly demand
        self.assertEqual(m(10.0, 2.0), 5.0)
        self.assertEqual(m(5.0, 2.0), 2.5)
        # Rule 3: Net available > 0 with zero demand returns 999.0
        self.assertEqual(m(10.0, 0.0), 999.0)
        self.assertEqual(m(10.0, -1.0), 999.0)

    def test_classify_abc(self):
        m = self.Suggestion._classify_abc
        self.assertEqual(m(10, self.config), 'A')
        self.assertEqual(m(5, self.config), 'A')
        self.assertEqual(m(3, self.config), 'B')
        self.assertEqual(m(1, self.config), 'B')
        self.assertEqual(m(0.5, self.config), 'C')

    def test_determine_urgency(self):
        m = self.Suggestion._determine_urgency
        # negative on-hand always wins, regardless of every other input
        self.assertEqual(
            m(qty_on_hand=-1, months_of_stock=999, avg_monthly=0,
              lead_months=1, is_dead_stock=False, suggested_qty=0),
            'critical',
        )
        self.assertEqual(
            m(qty_on_hand=5, months_of_stock=999, avg_monthly=0,
              lead_months=1, is_dead_stock=True, suggested_qty=0),
            'dead',
        )
        self.assertEqual(
            m(qty_on_hand=1, months_of_stock=0.5, avg_monthly=2,
              lead_months=1, is_dead_stock=False, suggested_qty=3),
            'urgent',
        )
        self.assertEqual(
            m(qty_on_hand=10, months_of_stock=5, avg_monthly=2,
              lead_months=1, is_dead_stock=False, suggested_qty=3),
            'normal',
        )
        self.assertEqual(
            m(qty_on_hand=10, months_of_stock=5, avg_monthly=2,
              lead_months=1, is_dead_stock=False, suggested_qty=0),
            'ok',
        )

    def test_calc_trend(self):
        m = self.Suggestion._calc_trend
        self.assertEqual(m(0, 0, 3, 3), ('new', 0.0))
        self.assertEqual(m(30, 0, 3, 3), ('up', 100.0))
        trend, pct = m(115, 100, 1, 1)
        self.assertEqual(trend, 'up')
        self.assertGreaterEqual(pct, 15)
        trend, pct = m(80, 100, 1, 1)
        self.assertEqual(trend, 'down')
        self.assertLessEqual(pct, -15)
        trend, pct = m(105, 100, 1, 1)
        self.assertEqual(trend, 'stable')

    def test_calc_seasonal_note(self):
        m = self.Suggestion._calc_seasonal_note
        self.assertIn('No data', m(10, 0))
        self.assertIn('spike', m(150, 100))
        self.assertIn('dip', m(50, 100))
        self.assertIn('Consistent', m(102, 100))

    def test_robust_demand_identical_months_no_outliers(self):
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [5.0, 5.0, 5.0, 5.0, 5.0]
        )
        self.assertEqual(forecast, 5.0)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

    def test_robust_demand_normal_variation_no_outliers(self):
        # Some natural month-to-month variance. Expecting the true average of all months.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [4.0, 5.0, 6.0, 5.0, 4.0]
        )
        self.assertEqual(forecast, 4.8)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

    def test_robust_demand_rejects_high_bulk_order_outlier(self):
        # One freak 500-unit bulk-order month against a 4-6 baseline is excluded.
        # Expecting the mean of clean months (4.8).
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [4.0, 5.0, 6.0, 5.0, 4.0, 500.0]
        )
        self.assertEqual(forecast, 4.8)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, False, False, True],
                          'only the 500-unit month must be flagged as excluded')

    def test_robust_demand_keeps_low_months_no_outliers(self):
        # Low months/zeros must never be excluded as outliers.
        # Expecting the true average of all months (42.0).
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [50.0, 49.0, 51.0, 52.0, 50.0, 0.0]
        )
        self.assertEqual(forecast, 42.0)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 6)

    def test_robust_demand_zero_mad_fallback(self):
        # [0, 0, 0, 5, 0] has median 0.0 and MAD 0.0. The lone 5-unit month holds
        # 100% of total sales and the next-highest month is 0, so the ratio test
        # alone would call it a spike. But it falls below the default minimum
        # spike size (10 units), so it is NOT excluded — a single small sale
        # must not be zeroed out and flagged as a one-time bulk order.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 5.0, 0.0]
        )
        self.assertEqual(forecast, 1.0)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

    def test_robust_demand_single_month_one_unit_not_spike(self):
        # A part that sold exactly 1 unit in one month out of the window and
        # nothing else must NOT be zeroed out as a "one-time big order" —
        # it's a normal slow/rare mover, not a bulk-order spike.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 1.0, 0.0]
        )
        self.assertEqual(forecast, 0.2)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

    def test_robust_demand_single_month_three_units_not_spike(self):
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 3.0, 0.0]
        )
        self.assertEqual(forecast, 0.6)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

    def test_robust_demand_single_month_large_qty_still_spike(self):
        # A genuine 2000-unit one-off bulk order (well above the default
        # minimum spike size) must still be quarantined as a spike.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 2000.0, 0.0]
        )
        self.assertEqual(forecast, 0.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, True, False])

    def test_robust_demand_min_spike_size_is_configurable(self):
        # Lowering min_spike_size below the sale quantity flips a small
        # single-month sale back into spike territory, proving the threshold
        # is actually wired through rather than hardcoded.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 5.0, 0.0], min_spike_size=1.0
        )
        self.assertEqual(forecast, 0.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, True, False])

    def test_robust_demand_spike_multiplier_is_configurable(self):
        # March=400, July=1550 (most-recent-first: [1550, 0, 0, 0, 400, 0]).
        # At the default 4x multiplier, 1550 is NOT >= 4*400=1600, so it's kept
        # in the average — a real conversation example (325/month forecast).
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [1550.0, 0.0, 0.0, 0.0, 400.0, 0.0]
        )
        self.assertEqual(forecast, 325.0)
        self.assertEqual(excluded, 0)

        # Lowering the multiplier to 3x: 1550 >= 3*400=1200 now passes, so July
        # gets excluded and the forecast drops to the average of the other 5
        # months (400/5 = 80/month) — proving the threshold is wired through.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [1550.0, 0.0, 0.0, 0.0, 400.0, 0.0], spike_multiplier=3.0
        )
        self.assertEqual(forecast, 80.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [True, False, False, False, False, False])

    def test_robust_demand_spike_dominance_pct_is_configurable(self):
        # Five months of 7 units plus one month of 30. The 30 clears the
        # multiplier test (30 >= 4*7=28) but NOT the default 50% dominance
        # test (30 < 0.5*65=32.5) — so by default it's kept in the average.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [7.0, 7.0, 7.0, 7.0, 7.0, 30.0]
        )
        self.assertAlmostEqual(forecast, 65.0 / 6)
        self.assertEqual(excluded, 0)

        # Lowering dominance to 40%: 30 >= 0.4*65=26 now passes, so the 30 is
        # excluded and the forecast becomes the plain average of the five 7s.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [7.0, 7.0, 7.0, 7.0, 7.0, 30.0], spike_dominance_pct=40.0
        )
        self.assertEqual(forecast, 7.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, False, False, True])

    def test_robust_demand_zero_mad_no_bulk_order(self):
        # [0, 0, 0, 5, 5] has median 0.0 and MAD 0.0.
        # Zero count is 3 (>= 2.5), but max month (5.0) is only 5.0 / 10.0 = 50% (< 75%) of total sales.
        # So it is NOT treated as a bulk order, and falls back to the raw average (2.0).
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 5.0, 5.0]
        )
        self.assertEqual(forecast, 2.0)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

    def test_robust_demand_safety_net(self):
        # Clean months [-2.0, 0.0, 2.0] average to 0.0, but total sales (100.0) is > 0.0.
        # Since the safety net fallback was removed, it must return the clean average of 0.0.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [-2.0, 0.0, 2.0, 100.0]
        )
        self.assertEqual(forecast, 0.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, True])

    def test_robust_demand_one_off_bulk_order_exclusion(self):
        # Case A: meets both conditions: at least half zero (3/5), single largest (150) is >= 75% of total sales (100%).
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 150.0, 0.0]
        )
        self.assertEqual(forecast, 0.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, True, False])

        # Case B: meets zero condition (3/5) but largest (100) is only 100/150 = 66.7% (< 75%) of total sales.
        # No exclusion, falls back to raw average (150/5 = 30).
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 0.0, 100.0, 50.0]
        )
        self.assertEqual(forecast, 30.0)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [False] * 5)

        # Case C: largest (150) is >= 75% of total, but zeros are only 2 out of 5 (40% < 50%).
        # In this case mad != 0, so it uses standard MAD logic to exclude the outlier.
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 20.0, 30.0, 150.0, 0.0]
        )
        self.assertEqual(forecast, 12.5)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, True, False])

    def test_robust_demand_empty_series(self):
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand([])
        self.assertEqual(forecast, 0.0)
        self.assertEqual(excluded, 0)
        self.assertIsNone(direction)
        self.assertEqual(flags, [])

    def test_robust_demand_fallback_spike_test(self):
        # Spike holds >= 50% total sales AND is >= 4 times larger than the next-highest month.
        # [0, 0, 2, 150, 0, 0] has total sales 152. Max 150 holds 98.7% of total sales.
        # Next-highest is 2. 150 >= 8 (4*2). Treated as spike!
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 0.0, 2.0, 150.0, 0.0, 0.0]
        )
        self.assertEqual(forecast, 0.4)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, False, False, True, False, False])

        # [0, 2000, 0, 0, 0, 0] has total sales 2000. Max 2000 holds 100% of total sales.
        # Next highest is 0. 2000 >= 0 (4*0). Treated as spike!
        forecast, excluded, direction, flags = self.Suggestion._compute_robust_monthly_demand(
            [0.0, 2000.0, 0.0, 0.0, 0.0, 0.0]
        )
        self.assertEqual(forecast, 0.0)
        self.assertEqual(excluded, 1)
        self.assertEqual(direction, 'high')
        self.assertEqual(flags, [False, True, False, False, False, False])

    def test_calc_delta_pct(self):
        m = self.Suggestion._calc_delta_pct
        self.assertEqual(m(0, 5), 100.0, 'brand-new suggestion must read as +100%, not divide by zero')
        self.assertEqual(m(0, 0), 0.0, 'no prior qty and no new qty is genuinely no change')
        self.assertEqual(m(10, 20), 100.0)
        self.assertEqual(m(20, 10), -50.0)
        self.assertEqual(m(10, 10), 0.0)
        self.assertEqual(m(4, 12), 200.0)

    def test_compute_needs_review_large_swing(self):
        m = self.Suggestion._compute_needs_review
        flagged, reason = m(avg_monthly=5, suggested_qty=10, is_dead_stock=False,
                             delta_pct=80.0, delta_threshold=50.0)
        self.assertTrue(flagged)
        self.assertIn('80%', reason)

        not_flagged, reason = m(avg_monthly=5, suggested_qty=10, is_dead_stock=False,
                                 delta_pct=30.0, delta_threshold=50.0)
        self.assertFalse(not_flagged)
        self.assertEqual(reason, '')

    def test_compute_needs_review_swing_trigger_disabled_at_zero_threshold(self):
        flagged, reason = self.Suggestion._compute_needs_review(
            avg_monthly=5, suggested_qty=10, is_dead_stock=False,
            delta_pct=500.0, delta_threshold=0,
        )
        self.assertFalse(flagged, 'threshold 0 must disable the swing trigger entirely')
        self.assertEqual(reason, '')

    def test_compute_needs_review_zero_demand_with_positive_suggestion(self):
        flagged, reason = self.Suggestion._compute_needs_review(
            avg_monthly=0, suggested_qty=8, is_dead_stock=False,
            delta_pct=0.0, delta_threshold=50.0,
        )
        self.assertTrue(flagged)
        self.assertIn('No real monthly demand', reason)

    def test_compute_needs_review_dead_stock_contradiction(self):
        flagged, reason = self.Suggestion._compute_needs_review(
            avg_monthly=2, suggested_qty=5, is_dead_stock=True,
            delta_pct=0.0, delta_threshold=50.0,
        )
        self.assertTrue(flagged)
        self.assertIn('dead stock', reason)

    def test_compute_needs_review_normal_dead_stock_not_flagged(self):
        # Dead stock with NO suggested reorder (the common, non-contradictory case).
        flagged, reason = self.Suggestion._compute_needs_review(
            avg_monthly=0, suggested_qty=0, is_dead_stock=True,
            delta_pct=0.0, delta_threshold=50.0,
        )
        self.assertFalse(flagged)
        self.assertEqual(reason, '')

    def test_compute_needs_review_multiple_triggers_combine(self):
        flagged, reason = self.Suggestion._compute_needs_review(
            avg_monthly=0, suggested_qty=10, is_dead_stock=True,
            delta_pct=200.0, delta_threshold=50.0,
        )
        self.assertTrue(flagged)
        self.assertIn('changed', reason)
        self.assertIn('No real monthly demand', reason)
        self.assertIn('dead stock', reason)

    def test_confidence_score_clean_history_is_100(self):
        score, reason = self.Suggestion._compute_confidence_score(
            monthly_series=[4.0, 5.0, 6.0, 5.0, 4.0, 5.0], excluded_months=0
        )
        self.assertEqual(score, 100.0)
        self.assertEqual(reason, 'Full clean history — no deductions.')

    def test_confidence_score_deducts_per_zero_month(self):
        score, reason = self.Suggestion._compute_confidence_score(
            monthly_series=[0.0, 0.0, 5.0, 4.0, 5.0, 4.0], excluded_months=0
        )
        self.assertEqual(score, 90.0, '2 zero months x 5-point penalty = -10')
        self.assertIn('2 of 6 months had zero sales (-10)', reason)

    def test_confidence_score_deducts_per_outlier_month(self):
        score, reason = self.Suggestion._compute_confidence_score(
            monthly_series=[4.0, 5.0, 6.0, 5.0, 4.0, 500.0], excluded_months=1
        )
        self.assertEqual(score, 85.0, '1 outlier month x 15-point penalty = -15')
        self.assertIn('1 month(s) excluded as an outlier (-15)', reason)

    def test_confidence_score_deducts_for_very_new_product(self):
        score, reason = self.Suggestion._compute_confidence_score(
            monthly_series=[0.0, 0.0, 0.0, 0.0, 0.0, 5.0], excluded_months=0
        )
        # 5 zero months (-25) + very-limited-history, only 1 non-zero month (-10) = 65
        self.assertEqual(score, 65.0)
        self.assertIn('5 of 6 months had zero sales (-25)', reason)
        self.assertIn('Very limited sales history — only 1 month(s) with any sales (-10)', reason)

    def test_confidence_score_floors_at_zero(self):
        score, _ = self.Suggestion._compute_confidence_score(
            monthly_series=[0.0] * 6, excluded_months=10
        )
        self.assertEqual(score, 0.0, 'must never go negative')

    def test_confidence_score_caps_at_100(self):
        score, _ = self.Suggestion._compute_confidence_score(
            monthly_series=[4.0, 5.0, 6.0, 5.0, 4.0, 5.0], excluded_months=0
        )
        self.assertLessEqual(score, 100.0)

    def test_confidence_score_deducts_concentration_by_default(self):
        # One month holds exactly 50% of total sales — concentration penalty
        # applies for the default ("Let the System Decide") behavior.
        score, reason = self.Suggestion._compute_confidence_score(
            monthly_series=[50.0, 10.0, 10.0, 10.0, 10.0, 10.0], excluded_months=0
        )
        self.assertEqual(score, 70.0, '1 concentrated month x 30-point penalty = -30')
        self.assertIn('Demand heavily concentrated in a single month (-30)', reason)

    def test_confidence_score_skips_concentration_for_bulk_regular(self):
        # Same concentrated series, but the buyer has confirmed the product as
        # "Customer Buys in Bulk Regularly" — the concentration is expected,
        # confirmed behavior, so no penalty should apply, and the note should
        # say so instead of silently omitting the deduction.
        score, reason = self.Suggestion._compute_confidence_score(
            monthly_series=[50.0, 10.0, 10.0, 10.0, 10.0, 10.0], excluded_months=0,
            reorder_behavior='bulk_regular',
        )
        self.assertEqual(score, 100.0, 'no deduction once the concentration is buyer-confirmed')
        self.assertNotIn('-30', reason)
        self.assertIn('no penalty applied', reason)
        self.assertIn('Customer Buys in Bulk Regularly', reason)

    def test_config_order_cycle_validation(self):
        with self.assertRaises(ValidationError):
            self.config.write({'order_cycle_months': -0.5})


@tagged('post_install', '-at_install')
class TestReorderGenerateSuggestions(TransactionCase):
    """Integration tests for generate_suggestions() — the main engine."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company

        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'analysis_period': '6',
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.0,
            'dead_stock_months': 6,
            'auto_flag_on_negative': True,
        })

        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Test Warehouse',
                'code': 'TWH',
                'company_id': cls.company.id,
            })

        cls.product = cls.env['product.product'].create({
            'name': 'Test Reorder Widget',
            'type': 'product',
            'standard_price': 10.0,
        })

    def _set_quant(self, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            self.product, self.warehouse.lot_stock_id, qty_delta
        )

    def test_generate_suggestions_flags_negative_stock_critical(self):
        self._set_quant(-5.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertEqual(suggestion.urgency, 'critical')
        self.assertTrue(suggestion.reorder_needed)
        self.assertTrue(suggestion.active)

    def test_orphan_suggestion_is_archived_not_deleted(self):
        # Run 1: negative stock creates an active critical suggestion.
        self._set_quant(-5.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])
        self.assertEqual(len(suggestion), 1)
        original_id = suggestion.id

        # Run 2: bring stock back to zero (out of scope — no sales, no negative qty).
        self._set_quant(5.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        archived = self.Suggestion.with_context(active_test=False).search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])
        self.assertEqual(len(archived), 1, 'orphan must be archived, never hard-deleted')
        self.assertEqual(archived.id, original_id, 'archiving must reuse the same record')
        self.assertFalse(archived.active)

    def test_reactivated_suggestion_reuses_archived_record(self):
        # Create, archive, then bring the product back into scope.
        self._set_quant(-5.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        original_id = self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
        ]).id

        self._set_quant(5.0)  # archive it
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )

        self._set_quant(-5.0)  # bring it back into scope (negative again)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )

        all_records = self.Suggestion.with_context(active_test=False).search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])
        self.assertEqual(
            len(all_records), 1,
            'reactivation must reuse the archived record, not violate the unique constraint '
            'by creating a duplicate'
        )
        self.assertEqual(all_records.id, original_id)
        self.assertTrue(all_records.active)
        self.assertEqual(all_records.urgency, 'critical')

    def test_generate_suggestions_min_max_policy_with_demand(self):
        self.config.write({
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.5,
            'order_cycle_months': 2.0,
        })
        self._set_quant(5.0)  # qty_on_hand = 5.0, qty_available = 5.0

        # Patch _compute_robust_monthly_demand to return average monthly demand of 10.0
        with patch.object(self.Suggestion, '_compute_robust_monthly_demand', return_value=(10.0, 0, None, [False]*6)):
            self.Suggestion.generate_suggestions(
                company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
            )
        suggestion = self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        # Min Level = 10.0 * (1.5 + 1.0) = 25.0
        # Max Level = 25.0 + 10.0 * 2.0 = 45.0
        # qty_available = 5.0 < Min Level (25.0) -> Triggered!
        # raw_qty = Max Level - qty_available = 45.0 - 5.0 = 40.0
        # MOQ = 1.0 -> suggested_qty = 40.0
        self.assertEqual(suggestion.min_stock_level, 25.0)
        self.assertEqual(suggestion.max_stock_level, 45.0)
        self.assertEqual(suggestion.suggested_reorder_qty, 40.0)

    def test_generate_suggestions_min_max_policy_not_triggered(self):
        self.config.write({
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.5,
            'order_cycle_months': 2.0,
        })
        self._set_quant(30.0)  # qty_on_hand = 30.0, qty_available = 30.0

        # Patch _compute_robust_monthly_demand to return average monthly demand of 10.0
        with patch.object(self.Suggestion, '_compute_robust_monthly_demand', return_value=(10.0, 0, None, [False]*6)):
            self.Suggestion.generate_suggestions(
                company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
            )
        suggestion = self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        # Min Level = 10.0 * (1.5 + 1.0) = 25.0
        # Max Level = 25.0 + 10.0 * 2.0 = 45.0
        # qty_available = 30.0 >= Min Level (25.0) -> Not triggered!
        # raw_qty = 0.0 -> suggested_qty = 0.0
        self.assertEqual(suggestion.min_stock_level, 25.0)
        self.assertEqual(suggestion.max_stock_level, 45.0)
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)


@tagged('post_install', '-at_install')
class TestReorderSuggestionActions(TransactionCase):
    """Snooze / unsnooze behavior on an existing suggestion."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.product = cls.env['product.product'].create({
            'name': 'Test Snooze Widget',
            'type': 'product',
            'standard_price': 5.0,
        })
        cls.suggestion = cls.env['smart.reorder.suggestion'].create({
            'company_id': cls.company.id,
            'warehouse_id': cls.warehouse.id,
            'product_id': cls.product.id,
            'qty_on_hand': -3.0,
            'suggested_reorder_qty': 10.0,
            'urgency': 'critical',
            'reorder_needed': True,
        })

    def test_snooze_30_suppresses_reorder_needed(self):
        self.suggestion.action_snooze_30()
        self.assertFalse(self.suggestion.reorder_needed)
        self.assertTrue(self.suggestion.is_snoozed)
        self.assertTrue(self.suggestion.snoozed_until)

    def test_unsnooze_recomputes_reorder_needed(self):
        self.suggestion.action_snooze_30()
        self.suggestion.action_unsnooze()
        self.assertFalse(self.suggestion.snoozed_until)
        self.assertFalse(self.suggestion.is_snoozed)
        self.assertTrue(
            self.suggestion.reorder_needed,
            'on-hand is still negative, so unsnoozing must re-flag it as needing reorder'
        )

    def test_refresh_vendor_performance(self):
        vendor = self.env['res.partner'].create({'name': 'Vendor Performance Partner'})
        self.suggestion.write({
            'vendor_id': vendor.id,
            'vendor_stated_lead_days': 10,
        })

        # Create two purchase orders in state 'purchase'
        po1 = self.env['purchase.order'].create({
            'partner_id': vendor.id,
            'company_id': self.company.id,
            'date_approve': date(2026, 6, 1),
            'effective_date': date(2026, 6, 11), # 10 days
        })
        po1.write({'state': 'purchase'})
        self.env['purchase.order.line'].create({
            'order_id': po1.id,
            'product_id': self.product.id,
            'name': self.product.name,
            'product_qty': 10,
            'price_unit': 5.0,
        })

        po2 = self.env['purchase.order'].create({
            'partner_id': vendor.id,
            'company_id': self.company.id,
            'date_approve': date(2026, 6, 1),
            'effective_date': date(2026, 6, 15), # 14 days
        })
        po2.write({'state': 'purchase'})
        self.env['purchase.order.line'].create({
            'order_id': po2.id,
            'product_id': self.product.id,
            'name': self.product.name,
            'product_qty': 10,
            'price_unit': 5.0,
        })

        # Call refresh
        self.suggestion.action_refresh_vendor_performance()
        self.suggestion.invalidate_recordset()

        # Average is (10 + 14) / 2 = 12.0 days
        self.assertEqual(self.suggestion.vendor_actual_avg_days, 12.0)
        self.assertIn("12 days", self.suggestion.vendor_performance_note)


@tagged('post_install', '-at_install')
class TestGenerateSuggestionsWizard(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.env['smart.reorder.config'].create({'company_id': cls.env.company.id})

    def test_warehouse_scope_without_selection_raises(self):
        wizard = self.env['smart.reorder.wizard'].create({
            'scope': 'warehouse',
            'company_ids': [],
            'warehouse_ids': [],
        })
        with self.assertRaises(UserError):
            wizard.action_generate()

    def test_all_scope_runs_without_error(self):
        wizard = self.env['smart.reorder.wizard'].create({'scope': 'all'})
        result = wizard.action_generate()
        self.assertEqual(result.get('res_model'), 'smart.reorder.suggestion')


@tagged('post_install', '-at_install')
class TestCronLogRetention(TransactionCase):
    """Recommendation: Run History had no retention, unlike forecast
    snapshots — it would grow one row per company per run forever. Mirrors
    ForecastSnapshot's retention purge; never touches a 'running' row."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.CronLog = self.env['smart.reorder.cron.log']
        self.config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.config:
            self.config = self.env['smart.reorder.config'].create({'company_id': self.company.id})

    def _make_log(self, days_old, status='completed'):
        return self.CronLog.create({
            'company_id': self.company.id,
            'started_at': fields.Datetime.now() - timedelta(days=days_old),
            'trigger_type': 'cron',
            'status': status,
        })

    def test_purges_logs_older_than_retention(self):
        self.config.write({'cron_log_retention_months': 6})
        old_log = self._make_log(400)      # ~13 months
        recent_log = self._make_log(30)
        self.CronLog._purge_old_logs()
        self.assertFalse(old_log.exists())
        self.assertTrue(recent_log.exists())

    def test_never_purges_running_log_even_if_old(self):
        self.config.write({'cron_log_retention_months': 1})
        stuck_log = self._make_log(400, status='running')
        self.CronLog._purge_old_logs()
        self.assertTrue(stuck_log.exists(), 'a running log must never be purged, regardless of age')

    def test_retention_zero_disables_purge(self):
        self.config.write({'cron_log_retention_months': 0})
        old_log = self._make_log(1000)
        self.CronLog._purge_old_logs()
        self.assertTrue(old_log.exists())

    def test_weekly_cron_calls_purge(self):
        self.config.write({'cron_log_retention_months': 1})
        old_log = self._make_log(400)
        self.env['smart.reorder.suggestion'].action_run_weekly_cron()
        self.assertFalse(old_log.exists(), 'the weekly cron entry point must trigger the purge')


@tagged('post_install', '-at_install')
class TestCronOverdueDetection(TransactionCase):
    """Recommendation: missed-run detection — both the passive dashboard
    field and the active daily heartbeat notification, so a stuck worker or
    an accidentally-disabled cron doesn't go silently unnoticed."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.config:
            self.config = self.env['smart.reorder.config'].create({'company_id': self.company.id})
        self.config.write({'cron_frequency': 'weekly', 'overdue_alert_sent': False})
        self.cron = self.env.ref('smart_reorder_advisor.cron_smart_reorder_weekly')
        self.cron.write({'active': True})
        self.CronLog = self.env['smart.reorder.cron.log']
        self.user = self.env['res.users'].create({
            'name': 'Overdue Alert Recipient',
            'login': 'overdue_alert_recipient_test',
            'email': 'overdue_alert_recipient_test@example.com',
            'groups_id': [(6, 0, [self.env.ref('base.group_user').id])],
        })
        self.config.write({'notify_user_ids': [(6, 0, [self.user.id])]})

    def _make_log(self, days_old, status='completed'):
        return self.CronLog.create({
            'company_id': self.company.id,
            'started_at': fields.Datetime.now() - timedelta(days=days_old),
            'finished_at': fields.Datetime.now() - timedelta(days=days_old),
            'trigger_type': 'cron',
            'status': status,
        })

    def test_dashboard_not_overdue_with_recent_run(self):
        self._make_log(1)
        dashboard = self.env['smart.reorder.observability.dashboard'].create({'company_id': self.company.id})
        self.assertFalse(dashboard.is_run_overdue)

    def test_dashboard_overdue_with_old_run(self):
        self._make_log(30)  # well past weekly (7d) * 1.5 grace
        dashboard = self.env['smart.reorder.observability.dashboard'].create({'company_id': self.company.id})
        self.assertTrue(dashboard.is_run_overdue)
        self.assertEqual(dashboard.expected_interval_days, 7)

    def test_dashboard_not_overdue_when_cron_disabled(self):
        self._make_log(30)
        self.cron.write({'active': False})
        dashboard = self.env['smart.reorder.observability.dashboard'].create({'company_id': self.company.id})
        self.assertFalse(dashboard.is_run_overdue, 'a deliberately disabled cron is not "overdue"')

    def test_heartbeat_sends_once_and_dedupes(self):
        self._make_log(30)
        count_before = self.env['mail.message'].search_count([('subject', 'ilike', 'overdue')])

        self.env['smart.reorder.config']._check_overdue_runs()
        self.config.invalidate_recordset()
        self.assertTrue(self.config.overdue_alert_sent)
        count_after_first = self.env['mail.message'].search_count([('subject', 'ilike', 'overdue')])
        self.assertGreater(count_after_first, count_before)

        # Second call must NOT send again.
        self.env['smart.reorder.config']._check_overdue_runs()
        count_after_second = self.env['mail.message'].search_count([('subject', 'ilike', 'overdue')])
        self.assertEqual(count_after_second, count_after_first, 'must not resend once already flagged')

    def test_heartbeat_clears_flag_once_resolved(self):
        self._make_log(30)
        self.env['smart.reorder.config']._check_overdue_runs()
        self.config.invalidate_recordset()
        self.assertTrue(self.config.overdue_alert_sent)

        self._make_log(0)  # a fresh successful run resolves it
        self.env['smart.reorder.config']._check_overdue_runs()
        self.config.invalidate_recordset()
        self.assertFalse(self.config.overdue_alert_sent)

    def test_heartbeat_noop_when_main_cron_disabled(self):
        self._make_log(30)
        self.cron.write({'active': False})
        self.env['smart.reorder.config']._check_overdue_runs()
        self.config.invalidate_recordset()
        self.assertFalse(self.config.overdue_alert_sent)


@tagged('post_install', '-at_install')
class TestReorderRunLock(TransactionCase):
    """Per-company run lock (T-21/T-22) — backed by smart.reorder.cron.log as the
    auditable source of truth (any 'running' record IS the lock). Cron and the
    manual wizard both funnel through generate_suggestions(), so locking it there
    covers both triggers without a separate check in each caller. config.is_running
    is just a computed view onto the log table, not an independent flag.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.CronLog = cls.env['smart.reorder.cron.log']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
        })
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.product = cls.env['product.product'].create({
            'name': 'Test Lock Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

    def _set_quant(self, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            self.product, self.warehouse.lot_stock_id, qty_delta
        )

    def _search_suggestion(self):
        return self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])

    def _open_running_logs(self):
        return self.CronLog.search([
            ('company_id', '=', self.company.id),
            ('status', '=', 'running'),
        ])

    def test_lock_released_after_successful_run(self):
        self._set_quant(-1.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        self.assertFalse(self.config.is_running)
        self.assertFalse(self._open_running_logs())
        log = self.CronLog.search([('company_id', '=', self.company.id)], limit=1)
        self.assertEqual(log.status, 'completed')
        self.assertEqual(log.trigger_type, 'cron')
        self.assertTrue(log.finished_at)
        self.assertGreaterEqual(log.duration_seconds, 0.0)

    def test_concurrent_run_is_skipped(self):
        self.CronLog.create({
            'company_id': self.company.id,
            'started_at': fields.Datetime.now(),
            'trigger_type': 'cron',
            'status': 'running',
        })
        self._set_quant(-1.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        self.assertFalse(self._search_suggestion(), 'a fresh lock must skip the company entirely')
        self.assertTrue(self.config.is_running, 'skip path must not touch an in-progress lock')
        self.assertEqual(len(self._open_running_logs()), 1, 'skip path must not create a second log row')

    def test_stuck_lock_is_overridden(self):
        stale_log = self.CronLog.create({
            'company_id': self.company.id,
            'started_at': fields.Datetime.now() - timedelta(minutes=61),
            'trigger_type': 'cron',
            'status': 'running',
        })
        self._set_quant(-1.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        self.assertTrue(self._search_suggestion(), 'a stale (>60min) lock must be overridden, not honored')
        self.assertFalse(self.config.is_running, 'lock must be released once the overriding run completes')
        self.assertEqual(stale_log.status, 'aborted', 'the stale lock itself must be marked aborted')
        completed_logs = self.CronLog.search([
            ('company_id', '=', self.company.id),
            ('status', '=', 'completed'),
        ])
        self.assertEqual(len(completed_logs), 1, 'overriding must create a fresh log, not reuse the stale one')

    def test_lock_released_even_if_company_processing_raises(self):
        self._set_quant(-1.0)
        with patch.object(
            type(self.Suggestion), '_send_notifications',
            side_effect=RuntimeError('boom'),
        ):
            with self.assertRaises(RuntimeError):
                self.Suggestion.generate_suggestions(
                    company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
                )
        self.assertFalse(self.config.is_running, 'the log must be closed out even on exception')
        log = self.CronLog.search([('company_id', '=', self.company.id)], limit=1)
        self.assertEqual(log.status, 'aborted')
        self.assertIn('boom', log.error_notes)

    def test_clear_lock_action(self):
        self.CronLog.create({
            'company_id': self.company.id,
            'started_at': fields.Datetime.now(),
            'trigger_type': 'cron',
            'status': 'running',
        })
        self.assertTrue(self.config.is_running)
        self.config.action_clear_lock()
        self.assertFalse(self.config.is_running)
        self.assertFalse(self._open_running_logs())

    def test_manual_trigger_recorded_as_manual(self):
        wizard = self.env['smart.reorder.wizard'].create({'scope': 'all'})
        wizard.action_generate()
        log = self.CronLog.search(
            [('company_id', '=', self.company.id)], limit=1, order='started_at desc'
        )
        self.assertEqual(log.trigger_type, 'manual')

    def test_email_report_failure_marks_completed_with_errors(self):
        self.config.write({'send_email_report': True})
        self._set_quant(-1.0)
        with patch.object(type(self.Suggestion), '_send_email_report', return_value=False):
            self.Suggestion.generate_suggestions(
                company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
            )
        log = self.CronLog.search([('company_id', '=', self.company.id)], limit=1)
        self.assertEqual(log.status, 'completed_with_errors')

    def test_lock_scoping_for_provisional(self):
        other_company = self.env['res.company'].create({'name': 'Other Company'})
        other_warehouse = self.env['stock.warehouse'].create({
            'name': 'Other WH',
            'code': 'OWH',
            'company_id': other_company.id,
        })
        other_config = self.env['smart.reorder.config'].create({
            'company_id': other_company.id,
        })
        other_product = self.env['product.product'].create({
            'name': 'Other Lock Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

        sug_current = self.Suggestion.create({
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'product_id': self.product.id,
            'is_provisional': True,
        })
        sug_other = self.Suggestion.create({
            'company_id': other_company.id,
            'warehouse_id': other_warehouse.id,
            'product_id': other_product.id,
            'is_provisional': True,
        })

        self.CronLog.create({
            'company_id': other_company.id,
            'started_at': fields.Datetime.now(),
            'trigger_type': 'cron',
            'status': 'running',
        })

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id, other_company.id],
            warehouse_ids=[self.warehouse.id, other_warehouse.id]
        )

        sug_current.invalidate_recordset()
        sug_other.invalidate_recordset()

        self.assertFalse(sug_current.is_provisional, "Current company's suggestion should be cleared")
        self.assertTrue(sug_other.is_provisional, "Locked other company's suggestion should retain provisional flag")


@tagged('post_install', '-at_install')
class TestPoWizardSecurityGuard(TransactionCase):
    """action_confirm_consolidation() must reject non-managers before touching
    active_ids, mirroring the guard on action_create_draft_po()."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.env['smart.reorder.config'].create({'company_id': cls.env.company.id})
        user_group = cls.env.ref('smart_reorder_advisor.group_smart_reorder_user')
        cls.basic_user = cls.env['res.users'].create({
            'name': 'Basic Reorder User',
            'login': 'basic_reorder_user_test',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id, user_group.id])],
        })

    def test_non_manager_cannot_confirm_consolidation(self):
        wizard = self.env['smart.reorder.po.wizard'].with_user(self.basic_user).create({})
        with self.assertRaises(UserError):
            wizard.action_confirm_consolidation()


@tagged('post_install', '-at_install')
class TestDraftPoVendorFallbackGuard(TransactionCase):
    """Task 8: a single-click Draft PO must never silently fall back to the
    configured Default Vendor — it must require an explicit vendor on the
    product, directing the user to the bulk wizard (which shows the fallback
    before confirming) instead."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.default_vendor = cls.env['res.partner'].create({'name': 'Fallback Default Vendor'})
        cls.config = cls.env['smart.reorder.config'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.config:
            cls.config = cls.env['smart.reorder.config'].create({'company_id': cls.company.id})
        cls.config.write({
            'allow_draft_po': True,
            'default_vendor_id': cls.default_vendor.id,
        })
        cls.manager = cls.env['res.users'].create({
            'name': 'Draft PO Fallback Manager',
            'login': 'draft_po_fallback_manager_test',
            'groups_id': [(6, 0, [
                cls.env.ref('base.group_user').id,
                cls.env.ref('smart_reorder_advisor.group_smart_reorder_manager').id,
                cls.env.ref('purchase.group_purchase_user').id,
            ])],
        })
        cls.product = cls.env['product.product'].create({
            'name': 'No Vendor Widget', 'type': 'product', 'standard_price': 1.0,
        })
        cls.suggestion = cls.env['smart.reorder.suggestion'].create({
            'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id,
            'product_id': cls.product.id, 'urgency': 'urgent',
            'suggested_reorder_qty': 5.0, 'reorder_needed': True,
        })

    def test_no_vendor_blocks_single_click_draft_po_even_with_default_configured(self):
        with self.assertRaises(UserError) as cm:
            self.suggestion.with_user(self.manager).action_create_draft_po()
        self.assertIn('Generate Consolidated POs', str(cm.exception))
        self.assertFalse(self.suggestion.po_ids, 'no PO must be created when the guard fires')

    def test_no_vendor_and_no_default_configured_gives_clear_message(self):
        self.config.write({'default_vendor_id': False})
        with self.assertRaises(UserError) as cm:
            self.suggestion.with_user(self.manager).action_create_draft_po()
        self.assertIn('Set a vendor', str(cm.exception))

    def test_po_wizard_reports_fallback_count_before_confirming(self):
        wizard = self.env['smart.reorder.po.wizard'].with_context(
            active_ids=[self.suggestion.id]
        ).with_user(self.manager).create({})
        self.assertEqual(wizard.fallback_vendor_count, 1)
        self.assertEqual(wizard.fallback_vendor_names, 'Fallback Default Vendor')


@tagged('post_install', '-at_install')
class TestMarkButtonsSecurityGuard(TransactionCase):
    """action_mark_one_time_order() / action_mark_regular_order() must reject
    non-managers before touching the product template, mirroring the guard on
    action_create_draft_po() / action_create_internal_transfer()."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.env['smart.reorder.config'].create({'company_id': cls.company.id})
        user_group = cls.env.ref('smart_reorder_advisor.group_smart_reorder_user')
        cls.basic_user = cls.env['res.users'].create({
            'name': 'Basic Reorder User',
            'login': 'basic_reorder_user_mark_test',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id, user_group.id])],
        })
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Test Warehouse',
                'code': 'TWH3',
                'company_id': cls.company.id,
            })
        cls.product = cls.env['product.product'].create({
            'name': 'Mark Button Guard Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        cls.suggestion = cls.env['smart.reorder.suggestion'].create({
            'company_id':   cls.company.id,
            'warehouse_id': cls.warehouse.id,
            'product_id':   cls.product.id,
            'needs_review': True,
        })

    def test_non_manager_cannot_mark_one_time_order(self):
        product_tmpl = self.product.product_tmpl_id
        with self.assertRaises(UserError):
            self.suggestion.with_user(self.basic_user).action_mark_one_time_order()
        # The guard must fire before the product template is touched.
        self.assertEqual(product_tmpl.reorder_behavior, 'system')

    def test_non_manager_cannot_mark_regular_order(self):
        product_tmpl = self.product.product_tmpl_id
        with self.assertRaises(UserError):
            self.suggestion.with_user(self.basic_user).action_mark_regular_order()
        self.assertEqual(product_tmpl.reorder_behavior, 'system')


@tagged('post_install', '-at_install')
class TestDashboardSingleCompanyUserAccess(TransactionCase):
    """Fix 13 / Fix 5: get_dashboard_data() must be scoped by the calling
    user's OWN allowed companies. Unlike test_dashboard_multi_company_security
    (which uses with_context(allowed_company_ids=...) on an otherwise
    unrestricted admin env), this uses a genuinely single-company res.users
    record — company_ids restricted to one company — and calls the method
    with_user(), so env.companies resolves from the user's real access
    rather than an explicit context override. Proves the scoping is real
    access control, not just a UI-layer convenience filter."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company_a = cls.env.company
        cls.company_b = cls.env['res.company'].create({'name': 'Dashboard Access Test Company B'})

        cls.warehouse_a = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company_a.id)], limit=1
        )
        if not cls.warehouse_a:
            cls.warehouse_a = cls.env['stock.warehouse'].create({
                'name': 'Dash Access WH A', 'code': 'DAWA', 'company_id': cls.company_a.id,
            })
        cls.warehouse_b = cls.env['stock.warehouse'].create({
            'name': 'Dash Access WH B', 'code': 'DAWB', 'company_id': cls.company_b.id,
        })

        product_a = cls.env['product.product'].create({'name': 'Dash Access Widget A', 'type': 'product'})
        product_b = cls.env['product.product'].create({'name': 'Dash Access Widget B', 'type': 'product'})

        cls.env['smart.reorder.suggestion'].create({
            'company_id': cls.company_a.id, 'warehouse_id': cls.warehouse_a.id,
            'product_id': product_a.id, 'urgency': 'ok', 'abc_class': 'C', 'sales_pattern': 'new',
        })
        cls.suggestion_b = cls.env['smart.reorder.suggestion'].create({
            'company_id': cls.company_b.id, 'warehouse_id': cls.warehouse_b.id,
            'product_id': product_b.id, 'urgency': 'critical', 'abc_class': 'A',
            'sales_pattern': 'new', 'reorder_needed': True, 'suggested_reorder_qty': 5.0,
            'vendor_price': 100.0,
        })

        user_group = cls.env.ref('smart_reorder_advisor.group_smart_reorder_user')
        cls.single_company_user = cls.env['res.users'].create({
            'name': 'Single Company Dashboard User',
            'login': 'single_company_dashboard_user',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id, user_group.id])],
            'company_id': cls.company_a.id,
            'company_ids': [(6, 0, [cls.company_a.id])],
        })

    def test_single_company_user_cannot_see_other_company_dashboard_figures(self):
        data = self.env['smart.reorder.suggestion'].with_user(
            self.single_company_user
        ).get_dashboard_data()

        self.assertEqual(
            data['urgency'].get('critical', 0), 0,
            "a single-company user must not see another company's critical count"
        )
        self.assertNotIn(
            self.warehouse_b.id, [w['id'] for w in data['warehouses']],
            "a single-company user must not see another company's warehouse"
        )
        self.assertEqual(data['total_reorder_value'], 0.0)
        for lst in data['top'].values():
            self.assertNotIn(
                self.suggestion_b.product_id.id,
                [r['product_id'][0] for r in lst],
                "a single-company user's top lists must not include the other company's product"
            )


@tagged('post_install', '-at_install')
class TestRobustDemandForecast(TransactionCase):
    """End-to-end: generate_suggestions() must forecast off the robust average + MAD
    monthly breakdown, not a plain total/N average, so one freak bulk-order
    month doesn't inflate the reorder qty, value, budget rank, and urgency."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company

        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'analysis_period': '6',
            'safety_buffer_months': 0.0,
            'default_lead_time_months': 0.1,
        })

        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Test Warehouse',
                'code': 'TWH2',
                'company_id': cls.company.id,
            })
        cls.partner = cls.env['res.partner'].create({'name': 'Test Robust Demand Customer'})
        cls.product = cls.env['product.product'].create({
            'name': 'Test Robust Demand Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

        # 6 calendar-month buckets, most-recent-first, anchored the same way
        # generate_suggestions() anchors them (today's month, then back by 1).
        today = fields.Date.today()
        cls.month_starts = [
            (today.replace(day=1) - relativedelta(months=i)) for i in range(6)
        ]
        month_starts = cls.month_starts
        # Most recent month (i=0) is a 500-unit outlier; the other 5 months
        # have a normal 4-6 baseline with some natural variance.
        monthly_qtys = [500.0, 4.0, 5.0, 6.0, 5.0, 4.0]
        for month_start, qty in zip(month_starts, monthly_qtys):
            order_date = datetime.combine(month_start + timedelta(days=4), datetime.min.time())
            order = cls.env['sale.order'].create({
                'partner_id': cls.partner.id,
                'company_id': cls.company.id,
                'warehouse_id': cls.warehouse.id,
                'date_order': order_date,
                'order_line': [(0, 0, {
                    'product_id': cls.product.id,
                    'product_uom_qty': qty,
                    'price_unit': 10.0,
                })],
            })
            order.action_confirm()
            order.order_line.qty_delivered = qty

    def test_forecast_uses_average_with_spike_rejection(self):
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertEqual(suggestion.total_qty_sold, 524.0, 'display total must stay the raw sum')
        self.assertEqual(
            suggestion.avg_monthly_demand, 4.8,
            'forecast must be the average of the clean months, not 524/6 ≈ 87.3'
        )
        self.assertEqual(suggestion.excluded_outlier_months, 1)
        self.assertIn('Excluded 1', suggestion.demand_forecast_note)
        self.assertIn('unusually high', suggestion.demand_forecast_note)

        # T-27: the notes audit trail must list every month by name/year, and
        # tag only the actual outlier month — so a buyer can see at a glance
        # which months fed into the number and what got filtered out.
        outlier_label = self.month_starts[0].strftime('%B %Y')
        normal_label = self.month_starts[1].strftime('%B %Y')
        self.assertIn('Per-month breakdown:', suggestion.notes)
        self.assertIn(f'{outlier_label}: 500 units [excluded as outlier]', suggestion.notes)
        self.assertIn(f'{normal_label}: 4 units', suggestion.notes)
        self.assertNotIn(f'{normal_label}: 4 units [excluded as outlier]', suggestion.notes)

        # T-28: confidence score — all 6 months have sales (no zero-month penalty),
        # 1 month excluded as an outlier (-15) → 85/100.
        self.assertEqual(suggestion.confidence, 85.0)
        self.assertIn('Confidence: 85/100', suggestion.notes)
        self.assertIn('1 month(s) excluded as an outlier (-15)', suggestion.notes)

    def test_sales_today_are_included(self):
        product_today = self.env['product.product'].create({
            'name': 'Test Today Sale Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        today_datetime = datetime.combine(fields.Date.today(), datetime.min.time()) + timedelta(hours=14)
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': today_datetime,
            'order_line': [(0, 0, {
                'product_id': product_today.id,
                'product_uom_qty': 10.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()
        order.order_line.qty_delivered = 10.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', product_today.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertEqual(suggestion.total_qty_sold, 10.0, "Today's sales must be fully included")

    def test_regression_series_1_spike_safe(self):
        # 0, 0, 2, 150, 0, 0 -> forecast 0.4/month, suggestion ~2 units, big_order_mixed
        prod = self.env['product.product'].create({
            'name': 'Test Regression Widget 1',
            'type': 'product',
            'standard_price': 1.0,
        })
        self.config.write({
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.5,
            'order_cycle_months': 1.0,
        })
        qtys = [0.0, 0.0, 2.0, 150.0, 0.0, 0.0]
        for month_start, qty in zip(self.month_starts, qtys):
            if qty > 0.0:
                order_date = datetime.combine(month_start + timedelta(days=4), datetime.min.time())
                order = self.env['sale.order'].create({
                    'partner_id': self.partner.id,
                    'company_id': self.company.id,
                    'warehouse_id': self.warehouse.id,
                    'date_order': order_date,
                    'order_line': [(0, 0, {
                        'product_id': prod.id,
                        'product_uom_qty': qty,
                        'price_unit': 10.0,
                    })],
                })
                order.action_confirm()
                order.order_line.qty_delivered = qty

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertAlmostEqual(suggestion.avg_monthly_demand, 0.4)
        self.assertEqual(suggestion.sales_pattern, 'big_order_mixed')
        self.assertEqual(suggestion.suggested_reorder_qty, 2.0)

    def test_spike_multiplier_config_flows_through_generate_suggestions(self):
        # Same March(400)/July(1550) pattern discussed with the user — proves
        # the new spike_multiplier config field actually reaches the real
        # calculation pipeline (config -> config_data -> engine), not just
        # the pure function tested in TestReorderCalculations.
        prod = self.env['product.product'].create({
            'name': 'Test Spike Multiplier Config Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self.config.write({
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.5,
            'order_cycle_months': 1.0,
        })
        qtys = [1550.0, 0.0, 0.0, 0.0, 400.0, 0.0]
        for month_start, qty in zip(self.month_starts, qtys):
            if qty > 0.0:
                order_date = datetime.combine(month_start + timedelta(days=4), datetime.min.time())
                order = self.env['sale.order'].create({
                    'partner_id': self.partner.id,
                    'company_id': self.company.id,
                    'warehouse_id': self.warehouse.id,
                    'date_order': order_date,
                    'order_line': [(0, 0, {
                        'product_id': prod.id,
                        'product_uom_qty': qty,
                        'price_unit': 10.0,
                    })],
                })
                order.action_confirm()
                order.order_line.qty_delivered = qty

        # Default multiplier (4.0x): July is kept in the average -> 325/month.
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertAlmostEqual(suggestion.avg_monthly_demand, 325.0)

        # Lower to 3.0x and re-run -> July now excluded as a spike -> 80/month.
        self.config.write({'spike_multiplier': 3.0})
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertAlmostEqual(suggestion.avg_monthly_demand, 80.0)

    def test_regression_series_2_one_time_sale(self):
        # 0, 2000, 0, 0, 0, 0 -> forecast 0, suggestion 0, flagged for review (one_time_big_order)
        prod = self.env['product.product'].create({
            'name': 'Test Regression Widget 2',
            'type': 'product',
            'standard_price': 1.0,
        })
        self.config.write({
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.5,
            'order_cycle_months': 1.0,
        })
        qtys = [0.0, 2000.0, 0.0, 0.0, 0.0, 0.0]
        for month_start, qty in zip(self.month_starts, qtys):
            if qty > 0.0:
                order_date = datetime.combine(month_start + timedelta(days=4), datetime.min.time())
                order = self.env['sale.order'].create({
                    'partner_id': self.partner.id,
                    'company_id': self.company.id,
                    'warehouse_id': self.warehouse.id,
                    'date_order': order_date,
                    'order_line': [(0, 0, {
                        'product_id': prod.id,
                        'product_uom_qty': qty,
                        'price_unit': 10.0,
                    })],
                })
                order.action_confirm()
                order.order_line.qty_delivered = qty

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertEqual(suggestion.avg_monthly_demand, 0.0)
        self.assertEqual(suggestion.sales_pattern, 'one_time_big_order')
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertTrue(suggestion.needs_review)

    def test_regression_series_3_single_small_sale_not_one_time_order(self):
        # 0, 0, 0, 1, 0, 0 -> a C-class part that sold once. Must NOT be treated
        # as a one-time bulk order: forecast stays a normal small average and the
        # pattern lands in "Sells Sometimes", not "One-Time Big Order Only".
        prod = self.env['product.product'].create({
            'name': 'Test Regression Widget 3',
            'type': 'product',
            'standard_price': 1.0,
        })
        self.config.write({
            'safety_buffer_months': 1.0,
            'default_lead_time_months': 1.5,
            'order_cycle_months': 1.0,
        })
        qtys = [0.0, 0.0, 0.0, 1.0, 0.0, 0.0]
        for month_start, qty in zip(self.month_starts, qtys):
            if qty > 0.0:
                order_date = datetime.combine(month_start + timedelta(days=4), datetime.min.time())
                order = self.env['sale.order'].create({
                    'partner_id': self.partner.id,
                    'company_id': self.company.id,
                    'warehouse_id': self.warehouse.id,
                    'date_order': order_date,
                    'order_line': [(0, 0, {
                        'product_id': prod.id,
                        'product_uom_qty': qty,
                        'price_unit': 10.0,
                    })],
                })
                order.action_confirm()
                order.order_line.qty_delivered = qty

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertAlmostEqual(suggestion.avg_monthly_demand, 1.0 / 6.0)
        self.assertEqual(suggestion.sales_pattern, 'sometimes')
        self.assertEqual(suggestion.excluded_outlier_months, 0)

    def test_reorder_behavior_against_order(self):
        prod = self.env['product.product'].create({
            'name': 'Against Order Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        prod.product_tmpl_id.reorder_behavior = 'against_order'

        # Generate some sales history
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_uom_qty': 10.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertIn("ORDER ONLY AGAINST CUSTOMER ORDER", suggestion.notes)

    def test_reorder_behavior_bulk_regular(self):
        prod = self.env['product.product'].create({
            'name': 'Regular Bulk Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        prod.product_tmpl_id.reorder_behavior = 'bulk_regular'

        qtys = [0.0, 0.0, 2.0, 150.0, 0.0, 0.0]
        for month_start, qty in zip(self.month_starts, qtys):
            if qty > 0.0:
                order = self.env['sale.order'].create({
                    'partner_id': self.partner.id,
                    'company_id': self.company.id,
                    'warehouse_id': self.warehouse.id,
                    'date_order': datetime.combine(month_start + timedelta(days=4), datetime.min.time()),
                    'order_line': [(0, 0, {
                        'product_id': prod.id,
                        'product_uom_qty': qty,
                        'price_unit': 10.0,
                    })],
                })
                order.action_confirm()
                order.order_line.qty_delivered = qty

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        # Average is sum(qtys)/len(qtys) = 152 / 6 = 25.333...
        self.assertAlmostEqual(suggestion.avg_monthly_demand, 25.33, places=2)
        self.assertEqual(suggestion.excluded_outlier_months, 0)

    def test_one_click_mark_buttons(self):
        prod = self.env['product.product'].create({
            'name': 'One Click Test Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        # Standard generation (which will flag as outlier review)
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_uom_qty': 2000.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()
        order.order_line.qty_delivered = 2000.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertTrue(suggestion.needs_review)

        cron_log_count_before = self.env['smart.reorder.cron.log'].search_count([])

        # Click one time order
        result = suggestion.action_mark_one_time_order()
        self.assertEqual(prod.product_tmpl_id.reorder_behavior, 'against_order')
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertFalse(suggestion.needs_review)
        self.assertEqual(result.get('tag'), 'display_notification',
                          'button must return a lightweight notification, not trigger a full run')

        # Click regular orders
        result = suggestion.action_mark_regular_order()
        self.assertEqual(prod.product_tmpl_id.reorder_behavior, 'bulk_regular')
        # Bypasses outlier check, so average should be 2000/6 = 333.33
        self.assertAlmostEqual(suggestion.avg_monthly_demand, 333.33, places=2)
        self.assertFalse(suggestion.needs_review)
        self.assertEqual(result.get('tag'), 'display_notification')

        # Neither button may trigger a real analysis run — Run History must show
        # no phantom scheduled/manual entries from what the user experiences as
        # a one-click product-level edit.
        self.assertEqual(
            self.env['smart.reorder.cron.log'].search_count([]),
            cron_log_count_before,
            'marking a product one-time/regular must not create a cron log entry '
            '(i.e. must not trigger a full generate_suggestions() run)'
        )

    def test_needs_review_bulk_concentration(self):
        prod = self.env['product.product'].create({
            'name': 'Concentration Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        # Single large sale in month 0
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_uom_qty': 100.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()
        order.order_line.qty_delivered = 100.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertTrue(suggestion.needs_review)
        self.assertIn("Single month carries half or more of total demand", suggestion.needs_review_reason)
        # A single 100-unit month against 5 zero months is also a one-time-order
        # spike (Fix 2): excluded as an outlier and forecast to 0. Confidence:
        # 5 zero months (-25) + 1 outlier month excluded (-15) +
        # very-limited-history, only 1 month with sales (-10) +
        # concentration (-30) = 100 - 80 = 20.
        self.assertEqual(suggestion.confidence, 20.0)

    def test_bulk_regular_skips_concentration_review(self):
        """A product marked 'bulk_regular' should NOT get the concentration review
        flag or the confidence concentration deduction (Fix 1 / Fix 10), even
        when a single month carries ≥50% of total demand."""
        prod = self.env['product.product'].create({
            'name': 'Bulk Regular Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        # Mark the product as bulk-regular
        prod.product_tmpl_id.reorder_behavior = 'bulk_regular'

        # Single large sale in month 0 — would normally trigger concentration
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_uom_qty': 100.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()
        order.order_line.qty_delivered = 100.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        # Concentration reason must NOT appear because buyer confirmed bulk-regular
        if suggestion.needs_review_reason:
            self.assertNotIn(
                "Single month carries half or more of total demand",
                suggestion.needs_review_reason,
            )
        # bulk_regular also bypasses spike-rejection entirely (no excluded
        # months), so only the zero-months (-25) and very-limited-history
        # (-10) deductions apply — no outlier and no concentration penalty:
        # 100 - 25 - 10 = 65.
        self.assertEqual(suggestion.confidence, 65.0)
        self.assertNotIn('-30', suggestion.notes)
        self.assertIn('no penalty applied', suggestion.notes)
        self.assertEqual(suggestion.excluded_outlier_months, 0)

    def test_dashboard_multi_company_security(self):
        # Baseline: dashboard scoped to self.company ONLY, before the other
        # company's data exists. Compared against the same call again after —
        # a relative before/after check, since this class accumulates
        # suggestions for self.company/self.warehouse across many test methods
        # and a hardcoded absolute expectation would be brittle.
        scoped = self.Suggestion.with_context(allowed_company_ids=[self.company.id])
        before = scoped.get_dashboard_data()

        # Create a second company
        other_company = self.env['res.company'].create({'name': 'Other test company'})
        other_warehouse = self.env['stock.warehouse'].create({
            'name': 'Other Warehouse',
            'code': 'OWH',
            'company_id': other_company.id,
        })
        prod = self.env['product.product'].create({
            'name': 'Other Company Product',
            'type': 'product',
        })
        # Create a suggestion for the other company, with a non-zero purchase
        # value so a budget-sum leak would actually be visible.
        other_suggestion = self.Suggestion.create({
            'company_id': other_company.id,
            'warehouse_id': other_warehouse.id,
            'product_id': prod.id,
            'avg_monthly_demand': 10.0,
            'suggested_reorder_qty': 10.0,
            'vendor_price': 50.0,
            'reorder_needed': True,
            'within_budget': True,
            'urgency': 'critical',
            'abc_class': 'A',
            'sales_pattern': 'regular',
        })
        self.assertEqual(other_suggestion.estimated_purchase_value, 500.0)

        # And an evaluated back-test snapshot for the other company, with a
        # deliberately extreme MAPE so a leak would be obvious in the average.
        self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': other_company.id,
            'warehouse_id': other_warehouse.id,
            'product_id': prod.id,
            'snapshot_date': date.today() - timedelta(days=60),
            'forecast_demand': 10.0,
            'lead_time_days': 30,
            'abc_class': 'A',
            'evaluated': True,
            'absolute_error_pct': 999.0,
        })

        # Run get_dashboard_data as self.env (with self.company in environment)
        # The allowed companies in self.env should only be self.company (not other_company)
        after = scoped.get_dashboard_data()

        # Verify other suggestion and other warehouse are excluded from the result
        self.assertEqual(after['urgency'].get('critical', 0), 0)
        self.assertNotIn(other_warehouse.id, [w['id'] for w in after['warehouses']])

        # Budget sums and back-test accuracy for self.company must be
        # unaffected by the other company's records existing.
        self.assertEqual(after['total_reorder_value'], before['total_reorder_value'])
        self.assertEqual(after['within_budget_value'], before['within_budget_value'])
        self.assertEqual(after['backtest']['overall_mape'], before['backtest']['overall_mape'])
        self.assertTrue(
            all(v < 900.0 for v in after['backtest']['mape_by_abc'].values()),
            'the other company\'s 999.0 MAPE snapshot must not leak into this company\'s breakdown'
        )

    def test_draft_po_feedback_loop(self):
        prod = self.env['product.product'].create({
            'name': 'PO Loop Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

        # Make a reorder suggested suggestion first. Uniform demand across
        # every month — a single-month spike would be excluded as an outlier
        # and classified "one-time big order" (Fix 2), which suppresses the
        # suggested qty for a reason unrelated to what this test is checking.
        for month_start in self.month_starts:
            order = self.env['sale.order'].create({
                'partner_id': self.partner.id,
                'company_id': self.company.id,
                'warehouse_id': self.warehouse.id,
                'date_order': datetime.combine(month_start + timedelta(days=4), datetime.min.time()),
                'order_line': [(0, 0, {
                    'product_id': prod.id,
                    'product_uom_qty': 15.0,
                    'price_unit': 10.0,
                })],
            })
            order.action_confirm()
            order.order_line.qty_delivered = 15.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(suggestion), 1)
        self.assertTrue(suggestion.reorder_needed)

        # Create draft PO and link it
        po = self.env['purchase.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
        })
        suggestion.write({'po_ids': [(4, po.id)]})

        # Re-run suggestion generation
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        
        self.assertFalse(suggestion.reorder_needed)
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertIn("Draft PO already created", suggestion.notes)

    def test_draft_po_ref_names_the_po_and_releases_once_confirmed(self):
        prod = self.env['product.product'].create({
            'name': 'Draft PO Ref Widget', 'type': 'product', 'standard_price': 1.0,
        })
        # Uniform demand across every month — a single-month spike would be
        # excluded as an outlier (Fix 2) and suppress the suggestion for a
        # reason unrelated to what this test is checking.
        for month_start in self.month_starts:
            order = self.env['sale.order'].create({
                'partner_id': self.partner.id,
                'company_id': self.company.id,
                'warehouse_id': self.warehouse.id,
                'date_order': datetime.combine(month_start + timedelta(days=4), datetime.min.time()),
                'order_line': [(0, 0, {
                    'product_id': prod.id, 'product_uom_qty': 15.0, 'price_unit': 10.0,
                })],
            })
            order.action_confirm()
            order.order_line.qty_delivered = 15.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertFalse(suggestion.draft_po_ref, "no draft PO linked yet")

        po = self.env['purchase.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_qty': 50.0,
                'price_unit': 10.0,
                'name': prod.name,
            })],
        })
        suggestion.write({'po_ids': [(4, po.id)]})
        self.assertEqual(suggestion.draft_po_ref, po.name)

        # Re-run: the guard must kick in — quantity/flag suppressed, and the
        # note prominently names this exact draft PO (reference + quantity).
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertFalse(suggestion.reorder_needed)
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertIn(po.name, suggestion.notes)
        self.assertIn('50', suggestion.notes)
        self.assertTrue(suggestion.needs_review)
        self.assertIn(po.name, suggestion.needs_review_reason)

        # Once the PO is no longer draft, the guard must release — "still in
        # draft state" is the trigger, not merely having a linked po_ids
        # record. (The PO's 50 units now also count as confirmed incoming
        # supply via Q3, so whether reorder_needed ends up True or False is
        # ordinary stock math from here — what this asserts is that the
        # guard's own suppression/note no longer applies.)
        po.write({'state': 'purchase'})
        self.assertFalse(suggestion.draft_po_ref, "draft_po_ref must clear once the PO is confirmed")
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertNotIn(
            "Draft PO already created", suggestion.notes or "",
            "the guard note must not persist once the linked PO is no longer draft"
        )

    def test_stale_draft_po_alert(self):
        prod = self.env['product.product'].create({
            'name': 'Stale Draft PO Widget', 'type': 'product', 'standard_price': 1.0,
        })
        for month_start in self.month_starts:
            order = self.env['sale.order'].create({
                'partner_id': self.partner.id,
                'company_id': self.company.id,
                'warehouse_id': self.warehouse.id,
                'date_order': datetime.combine(month_start + timedelta(days=4), datetime.min.time()),
                'order_line': [(0, 0, {
                    'product_id': prod.id, 'product_uom_qty': 15.0, 'price_unit': 10.0,
                })],
            })
            order.action_confirm()
            order.order_line.qty_delivered = 15.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)

        po = self.env['purchase.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'order_line': [(0, 0, {
                'product_id': prod.id, 'product_qty': 50.0, 'price_unit': 10.0, 'name': prod.name,
            })],
        })
        suggestion.write({'po_ids': [(4, po.id)]})

        config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        config.write({'stale_draft_po_days': 7})

        # Fresh PO: must not be flagged stale yet.
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertFalse(suggestion.is_draft_po_stale)
        self.assertEqual(suggestion.draft_po_stale_days, 0)

        # Backdate the PO's create_date past the threshold.
        self.env.cr.execute(
            "UPDATE purchase_order SET create_date = %s WHERE id = %s",
            (fields.Datetime.now() - timedelta(days=10), po.id),
        )
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertTrue(suggestion.is_draft_po_stale)
        self.assertGreaterEqual(suggestion.draft_po_stale_days, 10)
        self.assertIn('Sitting unconfirmed', suggestion.notes)

        # Confirming the PO must clear the flag.
        po.write({'state': 'purchase'})
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertFalse(suggestion.is_draft_po_stale)
        self.assertEqual(suggestion.draft_po_stale_days, 0)

    def test_stale_draft_po_threshold_zero_disables_alert(self):
        prod = self.env['product.product'].create({
            'name': 'Stale Draft PO Disabled Widget', 'type': 'product', 'standard_price': 1.0,
        })
        for month_start in self.month_starts:
            order = self.env['sale.order'].create({
                'partner_id': self.partner.id,
                'company_id': self.company.id,
                'warehouse_id': self.warehouse.id,
                'date_order': datetime.combine(month_start + timedelta(days=4), datetime.min.time()),
                'order_line': [(0, 0, {
                    'product_id': prod.id, 'product_uom_qty': 15.0, 'price_unit': 10.0,
                })],
            })
            order.action_confirm()
            order.order_line.qty_delivered = 15.0

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        po = self.env['purchase.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'order_line': [(0, 0, {
                'product_id': prod.id, 'product_qty': 50.0, 'price_unit': 10.0, 'name': prod.name,
            })],
        })
        suggestion.write({'po_ids': [(4, po.id)]})
        self.env.cr.execute(
            "UPDATE purchase_order SET create_date = %s WHERE id = %s",
            (fields.Datetime.now() - timedelta(days=100), po.id),
        )
        config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        config.write({'stale_draft_po_days': 0})

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertFalse(suggestion.is_draft_po_stale, 'threshold 0 must disable the alert entirely')

    def test_incoming_internal_transfers(self):
        prod = self.env['product.product'].create({
            'name': 'Transfer Loop Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

        # Add some historical sales so there is demand
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_uom_qty': 60.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()

        # Create a donor warehouse
        donor_wh = self.env['stock.warehouse'].create({
            'name': 'Donor Wh',
            'code': 'DNW',
            'company_id': self.company.id,
        })

        # Create a confirmed incoming internal transfer move
        picking_type = self.env['stock.picking.type'].search([
            ('company_id', '=', self.company.id),
            ('code', '=', 'internal')
        ], limit=1)
        
        picking = self.env['stock.picking'].create({
            'picking_type_id': picking_type.id,
            'location_id': donor_wh.lot_stock_id.id,
            'location_dest_id': self.warehouse.lot_stock_id.id,
            'company_id': self.company.id,
            'move_ids': [(0, 0, {
                'name': prod.name,
                'product_id': prod.id,
                'product_uom_qty': 40.0,
                'product_uom': prod.uom_id.id,
                'location_id': donor_wh.lot_stock_id.id,
                'location_dest_id': self.warehouse.lot_stock_id.id,
            })]
        })
        picking.action_confirm()

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        
        # Verify the 40.0 units are counted in incoming stock
        self.assertEqual(suggestion.qty_incoming, 40.0)

    def test_wizard_returned_scope_filtering(self):
        wizard = self.env['smart.reorder.wizard'].create({
            'scope': 'warehouse',
            'warehouse_ids': [(6, 0, [self.warehouse.id])],
            'include_zero_demand': False,
        })
        action = wizard.action_generate()
        self.assertEqual(action['res_model'], 'smart.reorder.suggestion')
        self.assertEqual(action['domain'], [('warehouse_id', 'in', [self.warehouse.id])])
        self.assertEqual(action['context'].get('search_default_gb_warehouse'), 1)

    def test_snapshot_eval_date_and_notifications(self):
        # We need manager group to click button
        self.env.user.groups_id = [(4, self.env.ref('smart_reorder_advisor.group_smart_reorder_manager').id)]
        
        prod = self.env['product.product'].create({
            'name': 'Back-testing Widget',
            'type': 'product',
        })
        
        # 1. Create a young snapshot
        snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'product_id': prod.id,
            'snapshot_date': date.today(),
            'lead_time_days': 30,
            'forecast_demand': 10.0,
            'evaluated': False,
        })
        
        # Attempting to evaluate should raise UserError
        with self.assertRaises(UserError) as e:
            snap.action_evaluate()
        self.assertIn("not yet scoreable", str(e.exception))
        
        # 2. Make it ready by setting snapshot date back
        snap.snapshot_date = date.today() - timedelta(days=45)
        
        # Try evaluating again
        res = snap.action_evaluate()
        self.assertTrue(snap.evaluated)
        self.assertEqual(res['type'], 'ir.actions.client')
        self.assertEqual(res['tag'], 'display_notification')
        self.assertEqual(res['params']['type'], 'success')

    def test_snapshot_evaluate_scores_mature_and_skips_immature_in_mixed_selection(self):
        # A mixed selection of one mature + one immature snapshot must score
        # the mature one instead of refusing the whole batch, and the
        # notification must report both counts.
        prod_mature = self.env['product.product'].create({
            'name': 'Mixed Eval Mature Widget', 'type': 'product',
        })
        prod_immature = self.env['product.product'].create({
            'name': 'Mixed Eval Immature Widget', 'type': 'product',
        })
        snap_mature = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'product_id': prod_mature.id,
            'snapshot_date': date.today() - timedelta(days=45),
            'lead_time_days': 30,
            'forecast_demand': 10.0,
            'evaluated': False,
        })
        immature_ready_date = date.today() + timedelta(days=30)
        snap_immature = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'product_id': prod_immature.id,
            'snapshot_date': date.today(),
            'lead_time_days': 30,
            'forecast_demand': 5.0,
            'evaluated': False,
        })

        res = (snap_mature + snap_immature).action_evaluate()

        snap_mature.invalidate_recordset()
        snap_immature.invalidate_recordset()
        self.assertTrue(snap_mature.evaluated, 'the mature snapshot must be scored')
        self.assertFalse(snap_immature.evaluated, 'the immature snapshot must be left alone, not errored on')

        self.assertEqual(res['type'], 'ir.actions.client')
        self.assertEqual(res['tag'], 'display_notification')
        self.assertEqual(res['params']['type'], 'success')
        message = res['params']['message']
        self.assertIn('1', message)
        self.assertIn('skipped', message)
        self.assertIn(str(immature_ready_date), message)

    def test_snapshot_skip_pending_duplicates(self):
        prod = self.env['product.product'].create({
            'name': 'Duplicate Test Widget',
            'type': 'product',
        })
        # Generate some sales history
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod.id,
                'product_uom_qty': 10.0,
                'price_unit': 10.0,
            })],
        })
        order.action_confirm()

        # Generate suggestions first time - should create snapshot
        self.config.snapshot_scope = 'all'
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        
        snapshots1 = self.env['smart.reorder.forecast.snapshot'].search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertEqual(len(snapshots1), 1)
        self.assertFalse(snapshots1.evaluated)

        # Generate suggestions second time - should skip duplicate
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        
        snapshots2 = self.env['smart.reorder.forecast.snapshot'].search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        # Still exactly 1 snapshot!
        self.assertEqual(len(snapshots2), 1)

    def test_snapshot_scope_settings(self):
        # 1. Product A-Class (A+B class widget)
        prod_ab = self.env['product.product'].create({
            'name': 'AB Widget',
            'type': 'product',
        })
        # Sells a lot, so Class A. A single 60-unit month would otherwise be
        # excluded as a one-time spike (Fix 2), zeroing the forecast and
        # misclassifying it as C — bulk_regular bypasses that (plain average).
        prod_ab.product_tmpl_id.reorder_behavior = 'bulk_regular'
        order1 = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod_ab.id,
                'product_uom_qty': 60.0,
                'price_unit': 10.0,
            })],
        })
        order1.action_confirm()
        order1.order_line.qty_delivered = 60.0

        # 2. Product C-Class (Slow mover)
        prod_c = self.env['product.product'].create({
            'name': 'C Widget',
            'type': 'product',
        })
        # Sells very little, so Class C
        order2 = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(self.month_starts[0] + timedelta(days=4), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': prod_c.id,
                'product_uom_qty': 0.1,
                'price_unit': 10.0,
            })],
        })
        order2.action_confirm()
        order2.order_line.qty_delivered = 0.1

        # Set config snapshot scope to 'ab_only'
        self.config.snapshot_scope = 'ab_only'

        # Generate suggestions
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )

        # Class AB should have snapshot
        snap_ab = self.env['smart.reorder.forecast.snapshot'].search([
            ('product_id', '=', prod_ab.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(snap_ab)

        # Class C should NOT have snapshot
        snap_c = self.env['smart.reorder.forecast.snapshot'].search([
            ('product_id', '=', prod_c.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertFalse(snap_c)

    def test_cron_include_zero_demand(self):
        prod = self.env['product.product'].create({
            'name': 'Zero Demand Widget',
            'type': 'product',
        })
        # Archive any existing suggestions for this product to avoid interference
        self.Suggestion.search([('product_id', '=', prod.id)]).unlink()

        # 1. With cron_include_zero_demand = False
        self.config.cron_include_zero_demand = False
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id],
            warehouse_ids=[self.warehouse.id],
            trigger_type='cron'
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertFalse(suggestion)

        # 2. With cron_include_zero_demand = True
        self.config.cron_include_zero_demand = True
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id],
            warehouse_ids=[self.warehouse.id],
            trigger_type='cron'
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)


@tagged('post_install', '-at_install')
class TestStaleDataCarryForward(TransactionCase):
    """T-23: a single warehouse's batch failure (query/DB error) must never wipe
    or silently skip its suggestions. Existing records are left untouched and
    flagged stale instead, the company run still completes (not aborted), and
    the cron log's error_count makes the failure visible without reading logs."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.CronLog = cls.env['smart.reorder.cron.log']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
        })
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.product = cls.env['product.product'].create({
            'name': 'Test Stale Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

    def _set_quant(self, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            self.product, self.warehouse.lot_stock_id, qty_delta
        )

    def _search_suggestion(self):
        return self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])

    def test_failed_batch_leaves_existing_records_untouched_and_flags_stale(self):
        # Baseline: a clean run creates a critical suggestion.
        self._set_quant(-5.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        baseline = self._search_suggestion()
        self.assertEqual(len(baseline), 1)
        self.assertFalse(baseline.is_stale)
        baseline_id = baseline.id

        # Second run: the per-warehouse batch blows up partway through (deep
        # inside the product loop, well past the point where it would otherwise
        # have archived/recreated suggestions).
        with patch.object(
            type(self.Suggestion), '_round_to_moq',
            side_effect=RuntimeError('simulated DB blip'),
        ):
            self.Suggestion.generate_suggestions(
                company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
            )  # must not raise — the failure is caught and handled internally

        # The company run must be reported as completed-with-errors, not aborted,
        # and the error must be counted on the log record.
        log = self.CronLog.search([('company_id', '=', self.company.id)], limit=1)
        self.assertEqual(log.status, 'completed_with_errors')
        self.assertEqual(log.error_count, 1)

        # The existing suggestion must survive untouched, just flagged stale.
        still_there = self._search_suggestion()
        self.assertEqual(len(still_there), 1)
        self.assertEqual(still_there.id, baseline_id, 'must not delete/archive and recreate')
        self.assertTrue(still_there.is_stale)
        self.assertIn('RuntimeError', still_there.stale_reason)
        self.assertEqual(still_there.urgency, 'critical', 'old data must be preserved, not wiped')

    def test_is_stale_clears_on_next_successful_run(self):
        self._set_quant(-5.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self._search_suggestion()
        suggestion.write({'is_stale': True, 'stale_reason': 'Pretend a previous run failed.'})

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        self.assertFalse(suggestion.is_stale, 'a successful refresh must clear the stale flag')
        self.assertFalse(suggestion.stale_reason)


@tagged('post_install', '-at_install')
class TestDashboardData(TransactionCase):
    """T-24: get_dashboard_data() must compute counts/sums DB-side (read_group /
    search_count) and return the same shape/values it always did — verified here
    against hand-built records rather than loading everything in Python."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Test Warehouse', 'code': 'DWH1', 'company_id': cls.company.id,
            })
        cls.warehouse2 = cls.env['stock.warehouse'].create({
            'name': 'Second Test Warehouse', 'code': 'DWH2', 'company_id': cls.company.id,
        })

        def make_product(name):
            return cls.env['product.product'].create({
                'name': name, 'type': 'product', 'standard_price': 1.0,
            })

        cls.p_critical = make_product('Dash Critical Widget')
        cls.p_urgent   = make_product('Dash Urgent Widget')
        cls.p_dead     = make_product('Dash Dead Widget')
        cls.p_fallingA = make_product('Dash Falling A Widget')
        cls.p_fallingB = make_product('Dash Falling B Widget')
        cls.p_other_wh = make_product('Dash Other Warehouse Widget')

        common = {'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id}
        # The dashboard's value tiles sum estimated_purchase_value (vendor-price
        # basis) — a stored computed field (suggested_reorder_qty * vendor_price),
        # so the fixtures drive it through those two inputs (20*5=100, 20*10=200)
        # instead of writing a raw amount the compute would not reproduce.
        cls.Suggestion.create([
            dict(common, product_id=cls.p_critical.id, urgency='critical', abc_class='A',
                 demand_trend='down', trend_pct=-10.0, is_dead_stock=False,
                 reorder_needed=True, within_budget=True, vendor_price=5.0,
                 qty_on_hand=-5.0, suggested_reorder_qty=20.0),
            dict(common, product_id=cls.p_urgent.id, urgency='urgent', abc_class='B',
                 demand_trend='up', trend_pct=30.0, is_dead_stock=False,
                 reorder_needed=True, within_budget=False, vendor_price=10.0,
                 months_of_stock=0.5, avg_monthly_demand=10.0, suggested_reorder_qty=20.0),
            dict(common, product_id=cls.p_dead.id, urgency='dead', abc_class='C',
                 demand_trend='stable', is_dead_stock=True, reorder_needed=False,
                 within_budget=True, reorder_value=0.0,
                 months_since_last_sale=8, qty_on_hand=3.0),
            dict(common, product_id=cls.p_fallingA.id, urgency='ok', abc_class='A',
                 demand_trend='down', trend_pct=-15.0, is_dead_stock=False,
                 reorder_needed=False, within_budget=True, reorder_value=0.0),
            dict(common, product_id=cls.p_fallingB.id, urgency='ok', abc_class='A',
                 demand_trend='down', trend_pct=-40.0, is_dead_stock=False,
                 reorder_needed=False, within_budget=True, reorder_value=0.0),
        ])
        # A record in a DIFFERENT warehouse — must be excluded when filtering by cls.warehouse.
        cls.Suggestion.create(dict(
            common, warehouse_id=cls.warehouse2.id, product_id=cls.p_other_wh.id,
            urgency='critical', abc_class='A', demand_trend='down',
            is_dead_stock=False, reorder_needed=True, within_budget=True,
            reorder_value=999.0, qty_on_hand=-99.0,
        ))

    def test_counts_and_sums(self):
        data = self.Suggestion.get_dashboard_data(warehouse_id=self.warehouse.id)
        self.assertEqual(data['urgency'], {'critical': 1, 'urgent': 1, 'dead': 1, 'ok': 2})
        self.assertEqual(data['abc'], {'A': 3, 'B': 1, 'C': 1})
        self.assertEqual(data['trend'], {'down': 3, 'up': 1, 'stable': 1})
        self.assertEqual(data['dead_count'], 1)
        self.assertEqual(data['total_reorder_value'], 300.0)
        self.assertEqual(data['within_budget_value'], 100.0)
        self.assertEqual(data['budget_count'], 1)

    def test_top_lists_ordering_and_scope(self):
        data = self.Suggestion.get_dashboard_data(warehouse_id=self.warehouse.id)
        self.assertEqual(len(data['top']['critical']), 1)
        self.assertEqual(data['top']['critical'][0]['product_id'][0], self.p_critical.id)
        self.assertEqual(len(data['top']['urgent']), 1)
        self.assertEqual(data['top']['urgent'][0]['product_id'][0], self.p_urgent.id)
        self.assertEqual(len(data['top']['dead']), 1)
        self.assertEqual(data['top']['dead'][0]['product_id'][0], self.p_dead.id)
        self.assertEqual(len(data['top']['rising']), 1)
        self.assertEqual(data['top']['rising'][0]['product_id'][0], self.p_urgent.id)
        # falling: two candidates, most negative trend_pct sorted first
        falling_ids = [r['product_id'][0] for r in data['top']['falling']]
        self.assertEqual(falling_ids, [self.p_fallingB.id, self.p_fallingA.id])

    def test_warehouse_filter_excludes_other_warehouses(self):
        data = self.Suggestion.get_dashboard_data(warehouse_id=self.warehouse.id)
        all_top_product_ids = {
            r['product_id'][0] for lst in data['top'].values() for r in lst
        }
        self.assertNotIn(self.p_other_wh.id, all_top_product_ids)
        self.assertEqual(data['urgency']['critical'], 1, 'must not include the other warehouse record')

    def test_no_warehouse_filter_includes_all_warehouses(self):
        data = self.Suggestion.get_dashboard_data()
        self.assertEqual(data['urgency']['critical'], 2, 'unfiltered call must include both warehouses')


@tagged('post_install', '-at_install')
class TestDeltaTracking(TransactionCase):
    """T-25: prior_suggested_qty/delta_pct must track suggested_reorder_qty
    changes across successive generate_suggestions() runs, so buyers can spot
    a suggestion that swung dramatically since last week before approving a PO."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'safety_buffer_months': 0.0,
            'default_lead_time_months': 0.1,
        })
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.product = cls.env['product.product'].create({
            'name': 'Test Delta Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

    def _set_quant(self, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            self.product, self.warehouse.lot_stock_id, qty_delta
        )

    def _run(self):
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        return self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])

    def test_first_run_is_treated_as_100_percent_increase(self):
        self._set_quant(-5.0)
        suggestion = self._run()
        self.assertEqual(suggestion.suggested_reorder_qty, 5.0)
        self.assertEqual(suggestion.prior_suggested_qty, 0.0)
        self.assertEqual(suggestion.delta_pct, 100.0)

    def test_delta_tracks_increase_then_decrease_across_runs(self):
        # Run 1: -5 on hand → suggested qty 5.
        self._set_quant(-5.0)
        suggestion = self._run()
        self.assertEqual(suggestion.suggested_reorder_qty, 5.0)

        # Run 2: push further negative (-20 total) → suggested qty 20.
        self._set_quant(-15.0)
        suggestion = self._run()
        self.assertEqual(suggestion.suggested_reorder_qty, 20.0)
        self.assertEqual(suggestion.prior_suggested_qty, 5.0)
        self.assertEqual(suggestion.delta_pct, 300.0)

        # Run 3: still negative (-2), just less severe — stays in scope (no sales
        # history means it would otherwise be archived as an orphan the moment
        # stock turns non-negative) — suggested qty drops to 2.
        self._set_quant(18.0)
        suggestion = self._run()
        self.assertEqual(suggestion.suggested_reorder_qty, 2.0)
        self.assertEqual(suggestion.prior_suggested_qty, 20.0)
        self.assertEqual(suggestion.delta_pct, -90.0)


@tagged('post_install', '-at_install')
class TestNeedsReviewTriage(TransactionCase):
    """T-26: generate_suggestions() must auto-flag suggestions worth a human
    look — wild swings vs the prior run, negative-stock-only "demand", and
    dead-stock/positive-suggestion contradictions — using the configurable
    delta threshold from smart.reorder.config."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'safety_buffer_months': 0.0,
            'default_lead_time_months': 0.1,
            'needs_review_delta_threshold': 50.0,
        })
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.product = cls.env['product.product'].create({
            'name': 'Test Review Widget',
            'type': 'product',
            'standard_price': 1.0,
        })

    def _set_quant(self, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            self.product, self.warehouse.lot_stock_id, qty_delta
        )

    def _run(self):
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        return self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.warehouse.id),
            ('company_id', '=', self.company.id),
        ])

    def test_negative_stock_with_no_real_demand_is_flagged(self):
        # No sales history at all — the suggestion is driven purely by negative
        # stock, not real demand. Must be flagged regardless of the delta swing.
        self._set_quant(-5.0)
        suggestion = self._run()
        self.assertEqual(suggestion.avg_monthly_demand, 0.0)
        self.assertGreater(suggestion.suggested_reorder_qty, 0.0)
        self.assertTrue(suggestion.needs_review)
        self.assertIn('No real monthly demand', suggestion.needs_review_reason)

    def test_large_swing_is_flagged_on_second_run(self):
        self._set_quant(-2.0)
        first = self._run()
        self.assertEqual(first.suggested_reorder_qty, 2.0)

        # Jump to -20 → suggested qty 20, a +900% swing — well past the 50% threshold.
        self._set_quant(-18.0)
        second = self._run()
        self.assertEqual(second.suggested_reorder_qty, 20.0)
        self.assertEqual(second.delta_pct, 900.0)
        self.assertTrue(second.needs_review)
        self.assertIn('changed', second.needs_review_reason)

    def test_swing_threshold_is_configurable(self):
        self.config.write({'needs_review_delta_threshold': 0})
        self._set_quant(-2.0)
        self._run()  # establish a prior suggested qty to diff against
        self._set_quant(-18.0)
        second = self._run()
        self.assertEqual(second.delta_pct, 900.0)
        # The swing trigger is disabled, but the "no real demand" trigger still
        # fires (it always does for a pure negative-stock suggestion) — confirms
        # the threshold only switches off the swing check, not the whole feature.
        self.assertTrue(second.needs_review)
        self.assertNotIn('changed', second.needs_review_reason)
        self.assertIn('No real monthly demand', second.needs_review_reason)


@tagged('post_install', '-at_install')
class TestPdfReportTrendDetail(TransactionCase):
    """T-29: the per-product PDF report must show the numbers behind the trend
    label — current vs previous period monthly average, % change, and the
    same-period-last-year comparison — not just "↑ Rising (+25%)"."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.product = cls.env['product.product'].create({
            'name': 'Test Trend Report Widget',
            'type': 'product',
            'standard_price': 1.0,
        })
        cls.suggestion = cls.Suggestion.create({
            'company_id': cls.company.id,
            'warehouse_id': cls.warehouse.id,
            'product_id': cls.product.id,
            'analysis_months': 6,
            'avg_monthly_demand': 10.0,
            'prev_period_qty_sold': 18.0,
            'trend_comparison_months': 3,
            'demand_trend': 'up',
            'trend_pct': 25.0,
            'same_period_last_year_qty': 30.0,
            'seasonal_note': '↑ 20% vs same period last year (seasonal spike?)',
            'urgency': 'normal',
            'reorder_needed': False,
        })

    def test_trend_detail_table_renders_with_correct_numbers(self):
        # Render to HTML (not PDF) so the test doesn't need wkhtmltopdf installed.
        html, _ = self.env['ir.actions.report']._render_qweb_html(
            'smart_reorder_advisor.action_report_reorder_suggestion',
            self.suggestion.ids,
        )
        html = html.decode('utf-8')
        self.assertIn('Demand Trend Detail', html)
        self.assertIn('10.00 units/mo', html, 'current-period avg must show')
        self.assertIn('6.00 units/mo', html, 'prev-period avg = 18 / 3 months')
        self.assertIn('+25.0%', html)
        self.assertIn('30.00 units total', html, 'same-period-last-year total')
        self.assertIn('5.00 units/mo)', html, 'same-period-last-year avg = 30 / 6 months')
        self.assertIn('seasonal spike', html)

    def test_trend_detail_handles_no_prior_data_gracefully(self):
        suggestion = self.Suggestion.create({
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'product_id': self.env['product.product'].create({
                'name': 'Test Trend Report Widget — No History',
                'type': 'product',
                'standard_price': 1.0,
            }).id,
            'analysis_months': 6,
            'avg_monthly_demand': 0.0,
            'prev_period_qty_sold': 0.0,
            'trend_comparison_months': 0,
            'demand_trend': 'new',
            'trend_pct': 0.0,
            'same_period_last_year_qty': 0.0,
            'seasonal_note': 'No data for same period last year',
            'urgency': 'ok',
            'reorder_needed': False,
        })
        html, _ = self.env['ir.actions.report']._render_qweb_html(
            'smart_reorder_advisor.action_report_reorder_suggestion',
            suggestion.ids,
        )
        html = html.decode('utf-8')
        self.assertIn('No data for same period last year', html)
        self.assertIn('★ New / No History', html)


@tagged('post_install', '-at_install')
class TestExportSuggestionsWizard(TransactionCase):
    """T-30: the toolbar Excel export must respect whatever scope the list
    view's header button passed in (current filter via active_domain, or an
    explicit row selection via active_ids), and produce a real .xlsx with
    the documented columns."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Wizard = cls.env['smart.reorder.export.wizard']
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.vendor = cls.env['res.partner'].create({'name': 'Test Export Vendor'})
        cls.product_a = cls.env['product.product'].create({
            'name': 'Export Widget A', 'type': 'product', 'standard_price': 2.0,
            'default_code': 'EXP-A',
        })
        cls.product_b = cls.env['product.product'].create({
            'name': 'Export Widget B', 'type': 'product', 'standard_price': 3.0,
            'default_code': 'EXP-B',
        })
        common = {'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id}
        cls.sugg_a = cls.Suggestion.create(dict(
            common, product_id=cls.product_a.id, urgency='critical', abc_class='A',
            demand_trend='up', qty_on_hand=-5.0, avg_monthly_demand=4.0,
            lead_time_months=1.0, suggested_reorder_qty=10.0, reorder_value=20.0,
            budget_rank=1, vendor_id=cls.vendor.id,
            last_sale_date=fields.Date.today(),
        ))
        cls.sugg_b = cls.Suggestion.create(dict(
            common, product_id=cls.product_b.id, urgency='ok', abc_class='C',
            demand_trend='stable', qty_on_hand=20.0, avg_monthly_demand=1.0,
            lead_time_months=1.0, suggested_reorder_qty=0.0, reorder_value=0.0,
            budget_rank=0,
        ))

    def test_export_domain_prefers_active_domain(self):
        wizard = self.Wizard.with_context(
            active_domain=[('id', '=', self.sugg_a.id)],
            active_ids=[self.sugg_a.id, self.sugg_b.id],
        ).create({})
        self.assertEqual(wizard._get_export_domain(), [('id', '=', self.sugg_a.id)])

    def test_export_domain_falls_back_to_active_ids(self):
        wizard = self.Wizard.with_context(active_ids=[self.sugg_b.id]).create({})
        self.assertEqual(wizard._get_export_domain(), [('id', 'in', [self.sugg_b.id])])

    def test_export_domain_defaults_to_everything(self):
        wizard = self.Wizard.create({})
        self.assertEqual(wizard._get_export_domain(), [])

    def test_record_count_reflects_domain(self):
        wizard = self.Wizard.with_context(
            active_domain=[('id', '=', self.sugg_a.id)]
        ).create({})
        self.assertEqual(wizard.record_count, 1)

    def test_export_defaults_to_essential_format(self):
        wizard = self.Wizard.create({})
        self.assertEqual(wizard.export_format, 'essential', 'Task 9: Essential must be the default')

    def test_export_essential_produces_expected_columns_and_rows(self):
        wizard = self.Wizard.with_context(
            active_domain=[('id', 'in', [self.sugg_a.id, self.sugg_b.id])]
        ).create({})
        action = wizard.action_export()
        self.assertEqual(action['type'], 'ir.actions.act_url')
        self.assertIn('download=true', action['url'])

        attachment_id = int(action['url'].split('/web/content/')[1].split('?')[0])
        attachment = self.env['ir.attachment'].browse(attachment_id)
        self.assertTrue(attachment.exists())
        self.assertEqual(
            attachment.mimetype,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('Essential', attachment.name)

        wb = load_workbook(io.BytesIO(base64.b64decode(attachment.datas)))
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header_row, [
            'Product Name/Description', 'Warehouse', 'On Hand Qty',
            'Suggested Reorder Qty', 'Vendor', 'Unit Cost', 'Reorder Value',
            'Dead Stock?', 'Last Sale Date',
        ])
        self.assertEqual(ws.max_row, 3, 'header + 2 data rows')

        data_rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
        names = {row[0] for row in data_rows}
        self.assertEqual(names, {'Export Widget A', 'Export Widget B'}, 'no vendor part codes in the identity column')

        row_a = next(row for row in data_rows if row[0] == 'Export Widget A')
        self.assertEqual(row_a[1], self.warehouse.name)
        self.assertEqual(row_a[2], -5.0)
        self.assertEqual(row_a[3], 10.0)
        self.assertEqual(row_a[4], 'Test Export Vendor')
        self.assertEqual(row_a[6], 20.0)
        self.assertEqual(row_a[7], 'No')
        self.assertEqual(row_a[8], self.sugg_a.last_sale_date)

    def test_export_full_produces_all_columns(self):
        wizard = self.Wizard.with_context(
            active_domain=[('id', 'in', [self.sugg_a.id, self.sugg_b.id])]
        ).create({'export_format': 'full'})
        action = wizard.action_export()
        attachment_id = int(action['url'].split('/web/content/')[1].split('?')[0])
        attachment = self.env['ir.attachment'].browse(attachment_id)
        self.assertIn('Full', attachment.name)

        wb = load_workbook(io.BytesIO(base64.b64decode(attachment.datas)))
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(len(header_row), 46, 'Full format must keep every existing column plus the 4 new cost columns')
        self.assertEqual(header_row[0], 'Budget Rank')
        self.assertEqual(header_row[3], 'Urgency')
        self.assertEqual(header_row[10], 'Part Number')
        self.assertEqual(header_row[37], 'Vendor')
        self.assertEqual(header_row[42], 'Last Purchase Cost')
        self.assertEqual(header_row[43], 'Effective Unit Cost')
        self.assertEqual(header_row[44], 'Price Discrepancy?')
        self.assertEqual(header_row[45], 'Price Discrepancy (%)')
        self.assertEqual(ws.max_row, 3, 'header + 2 data rows')

        data_rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
        row_a = next(row for row in data_rows if row[10] == 'EXP-A')
        self.assertEqual(row_a[3], 'Critical — Negative Stock')
        self.assertEqual(row_a[37], 'Test Export Vendor')

    def test_export_full_includes_cost_columns(self):
        self.sugg_a.write({
            'last_purchase_cost': 8.5,
            'effective_unit_cost': 8.5,
            'has_price_discrepancy': True,
            'price_discrepancy_pct': 25.0,
        })
        wizard = self.Wizard.with_context(
            active_domain=[('id', '=', self.sugg_a.id)]
        ).create({'export_format': 'full'})
        action = wizard.action_export()
        attachment_id = int(action['url'].split('/web/content/')[1].split('?')[0])
        attachment = self.env['ir.attachment'].browse(attachment_id)

        wb = load_workbook(io.BytesIO(base64.b64decode(attachment.datas)))
        ws = wb.active
        data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(data_row[42], 8.5)
        self.assertEqual(data_row[43], 8.5)
        self.assertEqual(data_row[44], 'Yes')
        self.assertEqual(data_row[45], 25.0)

    def test_export_with_no_matching_records_raises(self):
        wizard = self.Wizard.with_context(active_domain=[('id', '=', -1)]).create({})
        with self.assertRaises(UserError):
            wizard.action_export()


@tagged('post_install', '-at_install')
class TestTransferLane(TransactionCase):
    """T-31: inter-warehouse transfer recommendations must use a per-pair lane's
    lead time when one is configured, and fall back to the company default
    (in days) when it isn't — instead of one global lead time for every pair."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Lane = cls.env['smart.reorder.transfer.lane']
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'default_lead_time_months': 2.0,   # → 60-day fallback
        })
        cls.wh_a = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.wh_a:
            cls.wh_a = cls.env['stock.warehouse'].create({
                'name': 'Lane Test WH A', 'code': 'LWHA', 'company_id': cls.company.id,
            })
        cls.wh_b = cls.env['stock.warehouse'].create({
            'name': 'Lane Test WH B', 'code': 'LWHB', 'company_id': cls.company.id,
        })
        cls.product = cls.env['product.product'].create({
            'name': 'Test Transfer Lane Widget', 'type': 'product', 'standard_price': 1.0,
        })

    def _set_quant(self, warehouse, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            self.product, warehouse.lot_stock_id, qty_delta
        )

    def _run(self):
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.wh_a.id, self.wh_b.id]
        )
        return self.Suggestion.search([
            ('product_id', '=', self.product.id),
            ('warehouse_id', '=', self.wh_a.id),
            ('company_id', '=', self.company.id),
        ])

    # ── Model-level constraints ──────────────────────────────────────────

    def test_lane_rejects_identical_source_and_destination(self):
        with self.assertRaises(ValidationError):
            self.Lane.create({
                'source_warehouse_id': self.wh_a.id,
                'dest_warehouse_id': self.wh_a.id,
                'lead_time_days': 3,
            })

    def test_lane_rejects_non_positive_lead_time(self):
        with self.assertRaises(ValidationError):
            self.Lane.create({
                'source_warehouse_id': self.wh_a.id,
                'dest_warehouse_id': self.wh_b.id,
                'lead_time_days': 0,
            })

    def test_lane_rejects_duplicate_pair(self):
        self.Lane.create({
            'source_warehouse_id': self.wh_a.id,
            'dest_warehouse_id': self.wh_b.id,
            'lead_time_days': 3,
        })
        with self.assertRaises(ValidationError):
            self.Lane.create({
                'source_warehouse_id': self.wh_a.id,
                'dest_warehouse_id': self.wh_b.id,
                'lead_time_days': 5,
            })

    # ── generate_suggestions() integration ───────────────────────────────

    def test_transfer_uses_lane_lead_time_when_configured(self):
        self.Lane.create({
            'source_warehouse_id': self.wh_b.id,
            'dest_warehouse_id': self.wh_a.id,
            'lead_time_days': 3,
        })
        self._set_quant(self.wh_a, -5.0)   # WH A: negative stock → critical
        self._set_quant(self.wh_b, 50.0)   # WH B: surplus, no sales → mos = 999

        suggestion = self._run()
        self.assertEqual(suggestion.urgency, 'critical')
        self.assertEqual(suggestion.transfer_source_warehouse_id, self.wh_b)
        self.assertEqual(suggestion.transfer_lead_time_days, 3, 'must use the lane, not the company default')
        self.assertIn('lane: Lane Test WH B', suggestion.notes)

    def test_transfer_falls_back_to_company_default_without_a_lane(self):
        # No lane created for this pair.
        self._set_quant(self.wh_a, -5.0)
        self._set_quant(self.wh_b, 50.0)

        suggestion = self._run()
        self.assertEqual(suggestion.transfer_source_warehouse_id, self.wh_b)
        self.assertEqual(
            suggestion.transfer_lead_time_days, 3,
            'no lane configured — must fall back to default_internal_transfer_days (3)'
        )
        self.assertIn('company default — no specific lane configured', suggestion.notes)

    def test_transfer_custom_internal_lead_time_fallback(self):
        # Set custom fallback
        self.config.write({
            'default_internal_transfer_days': 8
        })
        self._set_quant(self.wh_a, -5.0)
        self._set_quant(self.wh_b, 50.0)

        suggestion = self._run()
        self.assertEqual(suggestion.transfer_lead_time_days, 8)

    def test_donor_surplus_cap(self):
        """
        Donor warehouse has 20 units on hand, sells 10/month with a 1-month
        lead time → protected need = 10; safe surplus = 10.
        Destination needs 25 units.
        The transfer suggestion must be capped at min(10, 25) = 10.
        """
        self.config.write({
            'analysis_period': '6',
            'default_lead_time_months': 1.0,
            'transfer_surplus_threshold': 1.0,
        })
        self._set_quant(self.wh_a, -25.0)   # destination needs 25
        self._set_quant(self.wh_b, 20.0)   # donor has 20

        # Create sales history: 60 units in the last month on WH B (so demand = 10/mo)
        import datetime as _dt
        last_month = fields.Date.today().replace(day=1) - relativedelta(months=1)
        order_date = _dt.datetime.combine(last_month, _dt.time(12, 0))
        so = self.env['sale.order'].create({
            'partner_id': self.env['res.partner'].create({'name': 'WH-B Customer'}).id,
            'company_id': self.env.company.id,
            'warehouse_id': self.wh_b.id,
            'date_order': order_date,
            'order_line': [(0, 0, {
                'product_id': self.product.id,
                'product_uom_qty': 60.0,
                'price_unit': 1.0,
            })],
        })
        so.action_confirm()
        so.order_line.qty_delivered = 60.0

        suggestion = self._run()
        self.assertEqual(
            suggestion.transfer_source_warehouse_id, self.wh_b,
            'WH-B should be chosen as the donor'
        )
        self.assertEqual(
            suggestion.transfer_suggested_qty, 10.0,
            'Transfer suggested qty should be capped at donor surplus of 10.0'
        )
        self.assertIn('══ INTERNAL TRANSFER ══', suggestion.notes)
        self.assertIn('Transfer recommended: 10.00 units from donor warehouse "Lane Test WH B"', suggestion.notes)


@tagged('post_install', '-at_install')
class TestCronFrequencySync(TransactionCase):
    """T-32: saving smart.reorder.config.cron_frequency must update the shared
    'Smart Reorder: Weekly Analysis' scheduled action's interval directly, so a
    manager never has to open Settings → Technical → Scheduled Actions."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Config = cls.env['smart.reorder.config']
        cls.company = cls.env.company
        cls.cron = cls.env.ref('smart_reorder_advisor.cron_smart_reorder_weekly')

    def test_create_with_default_frequency_syncs_weekly(self):
        # Detune the cron first so the assertion proves the sync actually ran,
        # not that it just happened to already match.
        self.cron.write({'interval_number': 99, 'interval_type': 'months'})
        config = self.Config.create({'company_id': self.company.id})
        self.assertEqual(config.cron_frequency, 'weekly')
        self.assertEqual(self.cron.interval_number, 1)
        self.assertEqual(self.cron.interval_type, 'weeks')

    def test_write_biweekly_updates_cron_interval(self):
        config = self.Config.create({'company_id': self.company.id})
        config.write({'cron_frequency': 'biweekly'})
        self.assertEqual(self.cron.interval_number, 2)
        self.assertEqual(self.cron.interval_type, 'weeks')

    def test_write_monthly_updates_cron_interval(self):
        config = self.Config.create({'company_id': self.company.id})
        config.write({'cron_frequency': 'monthly'})
        self.assertEqual(self.cron.interval_number, 1)
        self.assertEqual(self.cron.interval_type, 'months')

    def test_unrelated_field_write_does_not_touch_cron(self):
        config = self.Config.create({'company_id': self.company.id})
        config.write({'cron_frequency': 'monthly'})
        self.cron.write({'interval_number': 1, 'interval_type': 'months'})  # known baseline

        config.write({'safety_buffer_months': 2.0})  # unrelated field

        self.assertEqual(self.cron.interval_number, 1)
        self.assertEqual(self.cron.interval_type, 'months')


@tagged('post_install', '-at_install')
class TestCronFrequencyOnchangeWarning(TransactionCase):
    """Task 8 / Finding 8: changing cron_frequency must pop a real onchange
    warning (not just static help text) when other companies' configs would
    also be affected by the save — a genuine UI interrupt, not passive text."""

    def test_onchange_warns_when_other_company_differs(self):
        company_a = self.env.company
        company_b = self.env['res.company'].create({'name': 'Cron Warning Test Co. B'})
        config_a = self.env['smart.reorder.config'].search(
            [('company_id', '=', company_a.id)], limit=1
        ) or self.env['smart.reorder.config'].create({'company_id': company_a.id})
        config_a.write({'cron_frequency': 'weekly'})
        self.env['smart.reorder.config'].create({
            'company_id': company_b.id, 'cron_frequency': 'weekly',
        })

        new_config = config_a.new({'company_id': company_a.id, 'cron_frequency': 'monthly'})
        result = new_config._onchange_cron_frequency_warn_shared()
        self.assertTrue(result and result.get('warning'), 'must pop a real warning dict, not just be silent')
        self.assertIn('Cron Warning Test Co. B', result['warning']['message'])

    def test_onchange_no_warning_when_value_unchanged_across_companies(self):
        company_a = self.env.company
        config_a = self.env['smart.reorder.config'].search(
            [('company_id', '=', company_a.id)], limit=1
        ) or self.env['smart.reorder.config'].create({'company_id': company_a.id})
        config_a.write({'cron_frequency': 'weekly'})

        new_config = config_a.new({'company_id': company_a.id, 'cron_frequency': 'weekly'})
        result = new_config._onchange_cron_frequency_warn_shared()
        self.assertFalse(result, 'no other company differs, so nothing to warn about')


@tagged('post_install', '-at_install')
class TestBulkSnooze(TransactionCase):
    """T-33: bulk snooze (triggered from the list view's Action menu on a
    checkbox selection) must snooze every selected record in one write, the
    same way the single-record buttons do for one record at a time."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        common = {'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id}
        cls.suggestions = cls.Suggestion.create([
            dict(common, product_id=cls.env['product.product'].create({
                'name': f'Bulk Snooze Widget {i}', 'type': 'product', 'standard_price': 1.0,
            }).id, urgency='critical', qty_on_hand=-5.0, suggested_reorder_qty=5.0,
                 reorder_needed=True)
            for i in range(3)
        ])

    def test_bulk_snooze_7_sets_all_selected_records(self):
        self.suggestions.action_bulk_snooze_7()
        expected_until = fields.Date.today() + timedelta(days=7)
        for rec in self.suggestions:
            self.assertEqual(rec.snoozed_until, expected_until)
            self.assertFalse(rec.reorder_needed)
            self.assertIn('Bulk-snoozed 7 days', rec.snoozed_note)
            self.assertIn('3 suggestions', rec.snoozed_note)
        self.assertTrue(all(self.suggestions.mapped('is_snoozed')))

    def test_bulk_snooze_30_sets_all_selected_records(self):
        self.suggestions.action_bulk_snooze_30()
        expected_until = fields.Date.today() + timedelta(days=30)
        for rec in self.suggestions:
            self.assertEqual(rec.snoozed_until, expected_until)
            self.assertIn('Bulk-snoozed 30 days', rec.snoozed_note)

    def test_bulk_snooze_on_partial_selection_leaves_others_untouched(self):
        selected = self.suggestions[:2]
        untouched = self.suggestions[2]
        selected.action_bulk_snooze_7()
        self.assertTrue(all(selected.mapped('is_snoozed')))
        self.assertFalse(untouched.is_snoozed)
        self.assertTrue(untouched.reorder_needed)

    def test_bulk_snooze_on_empty_recordset_does_not_raise(self):
        empty = self.Suggestion.browse()
        empty.action_bulk_snooze_7()  # must be a silent no-op


@tagged('post_install', '-at_install')
class TestMarkAsOrdered(TransactionCase):
    """Task 5: one-click 'Mark as Ordered' — no PO, no wizard. Must suppress
    reorder_needed immediately, then survive exactly ONE generate_suggestions()
    run before auto-clearing, so a genuinely delayed/lost order resurfaces
    instead of being silently forgotten forever."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Mark Ordered Test WH', 'code': 'MOWH', 'company_id': cls.company.id,
            })
        if not cls.env['smart.reorder.config'].search([('company_id', '=', cls.company.id)], limit=1):
            cls.env['smart.reorder.config'].create({'company_id': cls.company.id})

    def _set_quant(self, product, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            product, self.warehouse.lot_stock_id, qty_delta
        )

    def test_mark_ordered_suppresses_immediately(self):
        product = self.env['product.product'].create({
            'name': 'Mark Ordered Widget', 'type': 'product', 'standard_price': 1.0,
        })
        suggestion = self.Suggestion.create({
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': product.id, 'urgency': 'urgent',
            'suggested_reorder_qty': 5.0, 'reorder_needed': True,
        })
        suggestion.action_mark_ordered()
        self.assertTrue(suggestion.is_marked_ordered)
        self.assertFalse(suggestion.reorder_needed)
        self.assertTrue(suggestion.marked_ordered_at)
        self.assertEqual(suggestion.marked_ordered_by_id, self.env.user)

    def test_unmark_ordered_restores_reorder_needed(self):
        product = self.env['product.product'].create({
            'name': 'Unmark Ordered Widget', 'type': 'product', 'standard_price': 1.0,
        })
        suggestion = self.Suggestion.create({
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': product.id, 'urgency': 'critical', 'qty_on_hand': -3.0,
            'suggested_reorder_qty': 5.0, 'reorder_needed': True,
        })
        suggestion.action_mark_ordered()
        self.assertFalse(suggestion.reorder_needed)
        suggestion.action_unmark_ordered()
        self.assertFalse(suggestion.is_marked_ordered)
        self.assertFalse(suggestion.marked_ordered_at)
        self.assertTrue(suggestion.reorder_needed, 'stock is still negative, so unmarking must re-flag it')

    def test_bulk_mark_ordered_sets_all_selected(self):
        products = [
            self.env['product.product'].create({
                'name': f'Bulk Mark Ordered Widget {i}', 'type': 'product', 'standard_price': 1.0,
            }) for i in range(3)
        ]
        suggestions = self.Suggestion.create([{
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': p.id, 'urgency': 'urgent',
            'suggested_reorder_qty': 5.0, 'reorder_needed': True,
        } for p in products])
        suggestions.action_bulk_mark_ordered()
        self.assertTrue(all(suggestions.mapped('is_marked_ordered')))
        self.assertFalse(any(suggestions.mapped('reorder_needed')))

    def test_bulk_mark_ordered_on_empty_recordset_does_not_raise(self):
        empty = self.Suggestion.browse()
        empty.action_bulk_mark_ordered()  # must be a silent no-op

    def test_mark_ordered_requires_reorder_advisor_group(self):
        outsider = self.env['res.users'].create({
            'name': 'No Reorder Group User',
            'login': 'no_reorder_group_user_mark_ordered_test',
            'groups_id': [(6, 0, [self.env.ref('base.group_user').id])],
        })
        product = self.env['product.product'].create({
            'name': 'Access Guard Widget', 'type': 'product', 'standard_price': 1.0,
        })
        suggestion = self.Suggestion.create({
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': product.id, 'urgency': 'urgent',
            'suggested_reorder_qty': 5.0, 'reorder_needed': True,
        })
        with self.assertRaises(AccessError):
            suggestion.with_user(outsider).action_mark_ordered()
        self.assertFalse(suggestion.is_marked_ordered, 'the guard must fire before any field is touched')

    def test_mark_ordered_suppresses_exactly_one_run_then_resurfaces(self):
        product = self.env['product.product'].create({
            'name': 'Mark Ordered Lifecycle Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, -4.0)  # negative stock -> always critical/reorder_needed

        # Run 1: creates the suggestion, negative stock -> reorder_needed True.
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertTrue(suggestion.reorder_needed)

        # Buyer marks it ordered before the next scheduled run.
        suggestion.action_mark_ordered()
        self.assertFalse(suggestion.reorder_needed)

        # Run 2 (stock still hasn't arrived — unchanged): must stay suppressed
        # AND consume the flag.
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertFalse(suggestion.reorder_needed, 'must still be suppressed for the run right after marking')
        self.assertFalse(suggestion.is_marked_ordered, 'the one-time suppression must be consumed after this run')

        # Run 3 (still no delivery logged): the order appears genuinely lost —
        # must resurface instead of staying silently hidden forever.
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion.invalidate_recordset()
        self.assertTrue(suggestion.reorder_needed, 'a still-unresolved shortage must resurface after the one suppressed cycle')

    def test_mark_ordered_creates_pending_log_entry(self):
        product = self.env['product.product'].create({
            'name': 'Mark Ordered Log Widget', 'type': 'product', 'standard_price': 1.0,
        })
        suggestion = self.Suggestion.create({
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': product.id, 'urgency': 'urgent',
            'suggested_reorder_qty': 7.0, 'reorder_needed': True,
        })
        suggestion.action_mark_ordered()
        log = self.env['smart.reorder.mark.ordered.log'].search([
            ('suggestion_id', '=', suggestion.id),
        ])
        self.assertEqual(len(log), 1)
        self.assertEqual(log.outcome, 'pending')
        self.assertEqual(log.suggested_qty_at_mark, 7.0)
        self.assertEqual(log.marked_by_id, self.env.user)
        self.assertEqual(log.product_id, product)
        self.assertFalse(log.resolved_at)

    def test_bulk_mark_ordered_creates_one_log_entry_per_record(self):
        products = [
            self.env['product.product'].create({
                'name': f'Bulk Mark Ordered Log Widget {i}', 'type': 'product', 'standard_price': 1.0,
            }) for i in range(2)
        ]
        suggestions = self.Suggestion.create([{
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': p.id, 'urgency': 'urgent',
            'suggested_reorder_qty': 3.0, 'reorder_needed': True,
        } for p in products])
        suggestions.action_bulk_mark_ordered()
        logs = self.env['smart.reorder.mark.ordered.log'].search([
            ('suggestion_id', 'in', suggestions.ids),
        ])
        self.assertEqual(len(logs), 2)
        self.assertTrue(all(l.outcome == 'pending' for l in logs))

    def test_reconciliation_log_resolves_as_resurfaced_when_still_needed(self):
        product = self.env['product.product'].create({
            'name': 'Mark Ordered Log Resurface Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, -4.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        suggestion.action_mark_ordered()

        # Stock still hasn't arrived — the underlying shortage is unchanged.
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        log = self.env['smart.reorder.mark.ordered.log'].search([
            ('suggestion_id', '=', suggestion.id),
        ])
        self.assertEqual(log.outcome, 'resurfaced')
        self.assertTrue(log.resolved_at)

    def test_reconciliation_log_resolves_as_cleared_when_resolved(self):
        product = self.env['product.product'].create({
            'name': 'Mark Ordered Log Cleared Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, -4.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        suggestion.action_mark_ordered()

        # The stock actually arrived and got logged before the next run.
        self._set_quant(product, 4.0)
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        log = self.env['smart.reorder.mark.ordered.log'].search([
            ('suggestion_id', '=', suggestion.id),
        ])
        self.assertEqual(log.outcome, 'cleared')
        self.assertTrue(log.resolved_at)


@tagged('post_install', '-at_install')
class TestLastRunBanner(TransactionCase):
    """T-34: the 'Last Analysis' banner above the suggestions list (banner_route)
    must prefer the cron log (Feature 1.2) for an accurate run timestamp/status,
    and fall back gracefully when there's no run history yet."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from ..controllers.main import SmartReorderBannerController
        cls.build_html = SmartReorderBannerController._build_banner_html
        cls.CronLog = cls.env['smart.reorder.cron.log']
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )

    def test_no_history_at_all(self):
        html = self.build_html(self.env)
        self.assertIn('No analysis has been run yet', html)

    def test_falls_back_to_suggestion_analysis_date_without_cron_log(self):
        product = self.env['product.product'].create({
            'name': 'Banner Test Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self.Suggestion.create({
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'product_id': product.id,
            'analysis_date': fields.Date.today(),
        })
        html = self.build_html(self.env)
        self.assertIn('no run history found yet', html)
        self.assertIn(str(fields.Date.today()), html)

    def test_completed_run_shows_info_banner(self):
        now = fields.Datetime.now()
        self.CronLog.create({
            'company_id': self.company.id,
            'started_at': now - timedelta(minutes=5),
            'finished_at': now,
            'trigger_type': 'cron',
            'status': 'completed',
        })
        html = self.build_html(self.env)
        self.assertIn('alert-info', html)
        self.assertIn('Last Analysis', html)
        self.assertNotIn('warnings', html)

    def test_completed_with_errors_shows_warning_banner(self):
        now = fields.Datetime.now()
        self.CronLog.create({
            'company_id': self.company.id,
            'started_at': now - timedelta(minutes=5),
            'finished_at': now,
            'trigger_type': 'cron',
            'status': 'completed_with_errors',
        })
        html = self.build_html(self.env)
        self.assertIn('alert-warning', html)
        self.assertIn('completed with some warnings', html)

    def test_most_recent_completed_run_wins(self):
        now = fields.Datetime.now()
        self.CronLog.create({
            'company_id': self.company.id,
            'started_at': now - timedelta(days=8),
            'finished_at': now - timedelta(days=8),
            'trigger_type': 'cron', 'status': 'completed',
        })
        recent = self.CronLog.create({
            'company_id': self.company.id,
            'started_at': now - timedelta(hours=1),
            'finished_at': now,
            'trigger_type': 'manual', 'status': 'completed',
        })
        html = self.build_html(self.env)
        local_dt = fields.Datetime.context_timestamp(self.env.user, recent.finished_at)
        self.assertIn(local_dt.strftime('%d %b %Y, %H:%M'), html)

    def test_running_status_is_ignored_in_favor_of_last_completed(self):
        now = fields.Datetime.now()
        completed = self.CronLog.create({
            'company_id': self.company.id,
            'started_at': now - timedelta(hours=2),
            'finished_at': now - timedelta(hours=1),
            'trigger_type': 'cron', 'status': 'completed',
        })
        self.CronLog.create({
            'company_id': self.company.id,
            'started_at': now,
            'trigger_type': 'cron', 'status': 'running',
        })
        html = self.build_html(self.env)
        local_dt = fields.Datetime.context_timestamp(self.env.user, completed.finished_at)
        self.assertIn(local_dt.strftime('%d %b %Y, %H:%M'), html)
        self.assertNotIn('No analysis has been run yet', html)


class TestProductSupersession(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'analysis_period': '3',
            'safety_buffer_months': 0.0,
            'default_lead_time_months': 1.0,
        })
        cls.warehouse = cls.env['stock.warehouse'].search([('company_id', '=', cls.company.id)], limit=1)

        # Create predecessor and successor templates/products
        cls.pred_tmpl = cls.env['product.template'].create({
            'name': 'Old Part A',
            'type': 'product',
        })
        cls.succ_tmpl = cls.env['product.template'].create({
            'name': 'New Part B',
            'type': 'product',
        })
        cls.pred_tmpl.superseded_by_id = cls.succ_tmpl.id

        cls.pred_prod = cls.pred_tmpl.product_variant_id
        cls.succ_prod = cls.succ_tmpl.product_variant_id

        # Add sales to old part in previous completed month (so it enters the analysis window)
        cls.partner = cls.env['res.partner'].create({'name': 'Supersession Customer'})

        # Last month start
        last_month_start = date.today().replace(day=1) - relativedelta(months=1)
        order_date = datetime.combine(last_month_start + timedelta(days=5), datetime.min.time())

        cls.order = cls.env['sale.order'].create({
            'partner_id': cls.partner.id,
            'company_id': cls.company.id,
            'warehouse_id': cls.warehouse.id,
            'date_order': order_date,
            'order_line': [(0, 0, {
                'product_id': cls.pred_prod.id,
                'product_uom_qty': 10.0,
                'price_unit': 10.0,
            })],
        })
        cls.order.action_confirm()
        cls.order.order_line.qty_delivered = 10.0

    def test_superseded_product_rollup_and_replenishment_exclusion(self):
        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )

        pred_suggestion = self.Suggestion.search([
            ('product_id', '=', self.pred_prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])
        succ_suggestion = self.Suggestion.search([
            ('product_id', '=', self.succ_prod.id),
            ('warehouse_id', '=', self.warehouse.id),
        ])

        self.assertTrue(pred_suggestion, "Predecessor suggestion should exist")
        self.assertTrue(succ_suggestion, "Successor suggestion should exist")

        # Predecessor assertions
        self.assertEqual(pred_suggestion.suggested_reorder_qty, 0.0)
        self.assertFalse(pred_suggestion.reorder_needed)
        self.assertEqual(pred_suggestion.urgency, 'dead')
        self.assertTrue(pred_suggestion.is_dead_stock)
        self.assertIn('SUPERSEDED: This part has been superseded', pred_suggestion.notes)

        # Successor assertions
        self.assertGreater(succ_suggestion.avg_monthly_demand, 0.0)
        self.assertIn('Demand includes rolled-up sales history from superseded parts', succ_suggestion.notes)
        self.assertIn('Old Part A', succ_suggestion.notes)

    def test_supersession_cycle_guard(self):
        # Direct self-reference
        with self.assertRaises(ValidationError):
            self.pred_tmpl.superseded_by_id = self.pred_tmpl.id

        # Loop: A -> B -> A
        with self.assertRaises(ValidationError):
            self.succ_tmpl.superseded_by_id = self.pred_tmpl.id


@tagged('post_install', '-at_install')
class TestSupersededNegativeStockUrgency(TransactionCase):
    """Fix 11: negative on-hand is promised/shipped stock that doesn't exist —
    that must never be hidden under a "Dead Stock" label, even for a
    superseded part. Each test method creates its own predecessor/successor
    product pair (rather than sharing one via setUpClass), so one test's
    quant can't leak into and contaminate the other's on-hand assertion."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.config = cls.env['smart.reorder.config'].create({
            'company_id': cls.company.id,
            'analysis_period': '3',
            'safety_buffer_months': 0.0,
            'default_lead_time_months': 1.0,
        })
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Test WH', 'code': 'SNSWH', 'company_id': cls.company.id,
            })
        cls.partner = cls.env['res.partner'].create({'name': 'Negative Stock Supersession Customer'})

    def _make_superseded_pair(self, name_suffix):
        pred_tmpl = self.env['product.template'].create({
            'name': f'Superseded Part {name_suffix}', 'type': 'product',
        })
        succ_tmpl = self.env['product.template'].create({
            'name': f'Successor Part {name_suffix}', 'type': 'product',
        })
        pred_tmpl.superseded_by_id = succ_tmpl.id
        pred_prod = pred_tmpl.product_variant_id

        last_month_start = date.today().replace(day=1) - relativedelta(months=1)
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(last_month_start + timedelta(days=5), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': pred_prod.id, 'product_uom_qty': 10.0, 'price_unit': 10.0,
            })],
        })
        order.action_confirm()
        order.order_line.qty_delivered = 10.0
        return pred_prod, succ_tmpl

    def test_superseded_with_negative_stock_is_critical_not_dead(self):
        pred_prod, succ_tmpl = self._make_superseded_pair('Negative')
        self.env['stock.quant'].create({
            'product_id': pred_prod.id,
            'location_id': self.warehouse.lot_stock_id.id,
            'quantity': -4.0,
        })

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', pred_prod.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertEqual(suggestion.qty_on_hand, -4.0)

        # No-replenishment behavior is unchanged even though urgency flips.
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertFalse(suggestion.reorder_needed)

        # Negative stock must never be hidden under a Dead Stock label.
        self.assertEqual(suggestion.urgency, 'critical')
        self.assertFalse(suggestion.is_dead_stock)
        self.assertIn('SUPERSEDED WITH NEGATIVE STOCK', suggestion.notes)
        self.assertIn(succ_tmpl.name, suggestion.notes)

    def test_superseded_with_positive_stock_stays_dead(self):
        pred_prod, succ_tmpl = self._make_superseded_pair('Positive')
        self.env['stock.quant'].create({
            'product_id': pred_prod.id,
            'location_id': self.warehouse.lot_stock_id.id,
            'quantity': 6.0,
        })

        self.Suggestion.generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id]
        )
        suggestion = self.Suggestion.search([
            ('product_id', '=', pred_prod.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertEqual(suggestion.qty_on_hand, 6.0)
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertFalse(suggestion.reorder_needed)
        self.assertEqual(suggestion.urgency, 'dead')
        self.assertTrue(suggestion.is_dead_stock)
        self.assertIn('SUPERSEDED: This part has been superseded', suggestion.notes)


class TestReorderConfigOrderCycle(TransactionCase):

    def test_order_cycle_derivation_and_manual_authoritativeness(self):
        config = self.env['smart.reorder.config'].create({
            'company_id': self.env.company.id,
            'cron_frequency': 'weekly',
        })
        # Default derived value should be 0.25 (weekly)
        self.assertEqual(config.order_cycle_months, 0.25)

        # Change frequency to biweekly -> updates to 0.5
        config.cron_frequency = 'biweekly'
        self.assertEqual(config.order_cycle_months, 0.5)

        # Change frequency to monthly -> updates to 1.0
        config.cron_frequency = 'monthly'
        self.assertEqual(config.order_cycle_months, 1.0)

        # Set a custom manual value -> manual value survives
        config.order_cycle_months = 2.5
        self.assertEqual(config.order_cycle_months, 2.5)

        # Changing frequency again with custom manual value -> manual value still survives
        config.cron_frequency = 'weekly'
        self.assertEqual(config.order_cycle_months, 2.5)


class TestReorderOverstockGuard(TransactionCase):

    def test_overstock_guard_trigger_and_flagging(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })
            
        product = self.env['product.product'].create({
            'name': 'Overstock Test Part',
            'type': 'product',
        })
        
        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'overstock_ceiling_months': 10.0,
            })
        else:
            config.write({'overstock_ceiling_months': 10.0})

        supplier = self.env['res.partner'].create({'name': 'Giant MOQ Vendor'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': supplier.id,
            'min_qty': 100.0,
            'price': 10.0,
        })

        sale_date = date(2026, 6, 15)
        so = self.env['sale.order'].create({
            'partner_id': supplier.id,
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'date_order': sale_date,
            'order_line': [(0, 0, {
                'product_id': product.id,
                'product_uom_qty': 6.0,
                'price_unit': 20.0,
            })],
        })
        so.action_confirm()
        so.order_line.qty_delivered = 6.0

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )

        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])

        self.assertTrue(suggestion, "Suggestion should be generated")
        self.assertTrue(suggestion.is_overstocked)
        self.assertTrue(suggestion.needs_review)
        self.assertIn("Vendor MOQ forces", suggestion.needs_review_reason)
        self.assertIn("months of cover", suggestion.needs_review_reason)


class TestReorderBudgetCapVendorPrice(TransactionCase):

    def test_budget_cap_valued_at_vendor_price(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })

        product = self.env['product.product'].create({
            'name': 'Budget Test Part',
            'type': 'product',
            'standard_price': 5.0,
        })

        supplier = self.env['res.partner'].create({'name': 'Vendor Price supplier'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': supplier.id,
            'price': 10.0,
            'min_qty': 1.0,
        })

        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'budget_cap': 20.0,
            })
        else:
            config.write({'budget_cap': 20.0})

        sale_date = date(2026, 6, 15)
        so = self.env['sale.order'].create({
            'partner_id': supplier.id,
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'date_order': sale_date,
            'order_line': [(0, 0, {
                'product_id': product.id,
                'product_uom_qty': 6.0,
                'price_unit': 20.0,
            })],
        })
        so.action_confirm()
        so.order_line.qty_delivered = 6.0

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )

        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])
        self.assertTrue(suggestion, "Suggestion should be generated")
        self.assertEqual(suggestion.reorder_value, 15.0)
        self.assertEqual(suggestion.estimated_purchase_value, 30.0)
        self.assertFalse(suggestion.within_budget, "Should exceed budget when valued at vendor price (30.0 > 20.0)")


@tagged('post_install', '-at_install')
class TestLastPurchaseCost(TransactionCase):
    """Task 1: last_purchase_cost / last_purchase_date / last_purchase_vendor_id
    must come from the most recent confirmed PO line (state purchase/done,
    highest date_order), and effective_unit_cost must fall back in order:
    Last Purchase Cost -> Vendor Price -> Standard Cost (Finding 6)."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.warehouse = self.env['stock.warehouse'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.warehouse:
            self.warehouse = self.env['stock.warehouse'].create({
                'name': 'LPC Test WH', 'code': 'LPCWH', 'company_id': self.company.id,
            })
        config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not config:
            self.env['smart.reorder.config'].create({'company_id': self.company.id})

    def _set_quant(self, product, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            product, self.warehouse.lot_stock_id, qty_delta
        )

    def _make_po(self, product, vendor, price_unit, date_order, qty=10.0):
        po = self.env['purchase.order'].create({
            'partner_id': vendor.id,
            'company_id': self.company.id,
            'date_order': date_order,
            'order_line': [(0, 0, {
                'product_id': product.id,
                'name': product.name,
                'product_qty': qty,
                'price_unit': price_unit,
            })],
        })
        po.write({'state': 'purchase'})
        return po

    def test_last_purchase_cost_uses_most_recent_po(self):
        product = self.env['product.product'].create({
            'name': 'LPC Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, 5.0)
        vendor_old = self.env['res.partner'].create({'name': 'Old Vendor'})
        vendor_new = self.env['res.partner'].create({'name': 'New Vendor'})
        self._make_po(product, vendor_old, 8.0, date(2026, 1, 5))
        self._make_po(product, vendor_new, 12.0, date(2026, 6, 5))

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id],
            include_zero_demand=True, trigger_type='manual',
        )
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion, 'zero-demand product with on-hand stock must still be analysed')
        self.assertEqual(suggestion.last_purchase_cost, 12.0, 'must pick the LATER PO (June), not the earlier one (January)')
        self.assertEqual(suggestion.last_purchase_date, date(2026, 6, 5))
        self.assertEqual(suggestion.last_purchase_vendor_id, vendor_new)
        self.assertEqual(suggestion.effective_unit_cost, 12.0)

    def test_effective_unit_cost_falls_back_to_vendor_price(self):
        product = self.env['product.product'].create({
            'name': 'LPC No History Widget', 'type': 'product', 'standard_price': 3.0,
        })
        self._set_quant(product, 5.0)
        vendor = self.env['res.partner'].create({'name': 'Pricelist Vendor'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': vendor.id,
            'price': 7.0,
            'min_qty': 1.0,
        })
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id],
            include_zero_demand=True, trigger_type='manual',
        )
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertEqual(suggestion.last_purchase_cost, 0.0, 'no PO history yet')
        self.assertEqual(suggestion.effective_unit_cost, 7.0, 'no purchase history: fall back to vendor price')

    def test_effective_unit_cost_falls_back_to_standard_cost(self):
        product = self.env['product.product'].create({
            'name': 'LPC No Vendor Widget', 'type': 'product', 'standard_price': 4.5,
        })
        self._set_quant(product, 5.0)
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id],
            include_zero_demand=True, trigger_type='manual',
        )
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertEqual(suggestion.last_purchase_cost, 0.0)
        self.assertEqual(suggestion.effective_unit_cost, 4.5, 'no purchase or vendor price: fall back to standard cost')


@tagged('post_install', '-at_install')
class TestPriceDiscrepancyFlag(TransactionCase):
    """Task 2: has_price_discrepancy/price_discrepancy_pct must only fire when
    BOTH a vendor pricelist price and a last-purchase cost are on record, and
    only past the fixed 15% threshold (a constant, not a config field, by
    explicit product decision)."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.warehouse = self.env['stock.warehouse'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.warehouse:
            self.warehouse = self.env['stock.warehouse'].create({
                'name': 'PDF Test WH', 'code': 'PDFWH', 'company_id': self.company.id,
            })
        if not self.env['smart.reorder.config'].search([('company_id', '=', self.company.id)], limit=1):
            self.env['smart.reorder.config'].create({'company_id': self.company.id})

    def _set_quant(self, product, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            product, self.warehouse.lot_stock_id, qty_delta
        )

    def _make_po(self, product, vendor, price_unit, date_order):
        po = self.env['purchase.order'].create({
            'partner_id': vendor.id,
            'company_id': self.company.id,
            'date_order': date_order,
            'order_line': [(0, 0, {
                'product_id': product.id,
                'name': product.name,
                'product_qty': 10.0,
                'price_unit': price_unit,
            })],
        })
        po.write({'state': 'purchase'})
        return po

    def _generate(self):
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id],
            include_zero_demand=True, trigger_type='manual',
        )

    def test_flagged_when_divergence_exceeds_threshold(self):
        product = self.env['product.product'].create({
            'name': 'PDF Widget Big Gap', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, 5.0)
        vendor = self.env['res.partner'].create({'name': 'PDF Vendor Big Gap'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': vendor.id,
            'price': 20.0,
            'min_qty': 1.0,
        })
        self._make_po(product, vendor, 10.0, date(2026, 6, 1))  # actually paid half the pricelist quote

        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertTrue(suggestion.has_price_discrepancy)
        self.assertAlmostEqual(suggestion.price_discrepancy_pct, 100.0)

    def test_not_flagged_within_threshold(self):
        product = self.env['product.product'].create({
            'name': 'PDF Widget Small Gap', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, 5.0)
        vendor = self.env['res.partner'].create({'name': 'PDF Vendor Small Gap'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': vendor.id,
            'price': 10.5,
            'min_qty': 1.0,
        })
        self._make_po(product, vendor, 10.0, date(2026, 6, 1))  # 5% above what was paid

        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertFalse(suggestion.has_price_discrepancy, '5% divergence must stay under the 15% threshold')
        self.assertAlmostEqual(suggestion.price_discrepancy_pct, 5.0)

    def test_not_flagged_when_no_purchase_history(self):
        product = self.env['product.product'].create({
            'name': 'PDF Widget No History', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, 5.0)
        vendor = self.env['res.partner'].create({'name': 'PDF Vendor No History'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': vendor.id,
            'price': 50.0,
            'min_qty': 1.0,
        })
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertFalse(
            suggestion.has_price_discrepancy,
            'must never flag with no last-purchase data to compare the pricelist against'
        )
        self.assertEqual(suggestion.price_discrepancy_pct, 0.0)


@tagged('post_install', '-at_install')
class TestBossWeeklyOrderReport(TransactionCase):
    """Task 3: the boss's report must be vendor-grouped, use Product Name only
    (no vendor part codes / internal reference), price lines using the
    fallback-chain effective_unit_cost, and route no-vendor / placeholder-vendor
    / dead-stock-with-negative-balance items to "Needs Attention" instead of
    presenting them as a normal, urgent line item."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.real_vendor = cls.env['res.partner'].create({'name': 'Real Spare Parts Co.'})
        cls.temp_vendor = cls.env['res.partner'].create({'name': 'Temporary Supplier'})

        config = cls.env['smart.reorder.config'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not config:
            config = cls.env['smart.reorder.config'].create({'company_id': cls.company.id})
        config.write({'temp_vendor_ids': [(6, 0, [cls.temp_vendor.id])]})

        cls.product_ok = cls.env['product.product'].create({
            'name': 'Boss Report Normal Widget',
            'default_code': 'VENDOR-SKU-SHOULD-NOT-APPEAR',
            'type': 'product', 'standard_price': 2.0,
        })
        cls.product_temp_vendor = cls.env['product.product'].create({
            'name': 'Boss Report Temp Vendor Widget', 'type': 'product', 'standard_price': 2.0,
        })
        cls.product_no_vendor = cls.env['product.product'].create({
            'name': 'Boss Report No Vendor Widget', 'type': 'product', 'standard_price': 2.0,
        })
        cls.product_dead_critical = cls.env['product.product'].create({
            'name': 'Boss Report Dead Critical Widget', 'type': 'product', 'standard_price': 2.0,
        })

        cls.sug_ok = cls.Suggestion.create({
            'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id,
            'product_id': cls.product_ok.id, 'vendor_id': cls.real_vendor.id,
            'suggested_reorder_qty': 10.0, 'effective_unit_cost': 5.0,
            'urgency': 'urgent', 'reorder_needed': True,
        })
        cls.sug_temp_vendor = cls.Suggestion.create({
            'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id,
            'product_id': cls.product_temp_vendor.id, 'vendor_id': cls.temp_vendor.id,
            'suggested_reorder_qty': 4.0, 'effective_unit_cost': 3.0,
            'urgency': 'normal', 'reorder_needed': True,
        })
        cls.sug_no_vendor = cls.Suggestion.create({
            'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id,
            'product_id': cls.product_no_vendor.id,
            'suggested_reorder_qty': 6.0, 'effective_unit_cost': 1.5,
            'urgency': 'urgent', 'reorder_needed': True,
        })
        cls.sug_dead_critical = cls.Suggestion.create({
            'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id,
            'product_id': cls.product_dead_critical.id, 'vendor_id': cls.real_vendor.id,
            'suggested_reorder_qty': 8.0, 'effective_unit_cost': 4.0,
            'urgency': 'critical', 'is_dead_stock': True, 'qty_on_hand': -8.0,
            'reorder_needed': True,
        })

    def _render(self):
        docs = (self.sug_ok | self.sug_temp_vendor | self.sug_no_vendor | self.sug_dead_critical)
        html, _ = self.env['ir.actions.report']._render_qweb_html(
            'smart_reorder_advisor.action_report_boss_weekly_order', docs.ids,
        )
        return html.decode('utf-8')

    def test_normal_item_appears_under_its_vendor_with_subtotal(self):
        html = self._render()
        self.assertIn('Real Spare Parts Co.', html)
        self.assertIn('Boss Report Normal Widget', html)
        self.assertIn('subtotal', html)
        # 10 units * 5.0 effective unit cost = 50.00
        self.assertIn('50.00', html)

    def test_no_vendor_part_codes_ever_shown(self):
        html = self._render()
        self.assertNotIn('VENDOR-SKU-SHOULD-NOT-APPEAR', html, 'no vendor part codes must ever appear in the boss report')

    def test_temp_vendor_item_routed_to_needs_attention(self):
        html = self._render()
        self.assertIn('Needs Attention', html)
        self.assertIn('Boss Report Temp Vendor Widget', html)
        self.assertIn('Temporary/placeholder vendor', html)

    def test_no_vendor_item_routed_to_needs_attention(self):
        html = self._render()
        self.assertIn('Boss Report No Vendor Widget', html)
        self.assertIn('No vendor assigned', html)

    def test_dead_stock_critical_item_never_shown_as_urgent(self):
        html = self._render()
        self.assertIn('Boss Report Dead Critical Widget', html)
        self.assertIn('Dead stock with negative balance', html)
        # The item must not appear inside a vendor block row (which would
        # show it as "Critical" alongside genuinely orderable items) — it
        # must only appear in the Needs Attention table.
        self.assertNotIn('>Critical<', html, 'a dead-stock/negative-balance item must never display as Critical to the boss')

    def test_grand_total_excludes_needs_attention_items(self):
        html = self._render()
        # Grand total = only sug_ok (50.00) since the other three are routed
        # to Needs Attention, not summed into any vendor subtotal or the total.
        self.assertIn('GRAND TOTAL', html)
        self.assertIn('50.00', html)


@tagged('post_install', '-at_install')
class TestDataCleanupTriageFlags(TransactionCase):
    """Recommendation: bulk triage tool — needs_vendor_assignment and
    dead_stock_critical_review must correctly identify the two backlog
    categories the boss report already routes away from the normal list,
    and the vendor-assign wizard must durably fix the vendor (via
    product.supplierinfo), not just cosmetically patch the suggestion."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.warehouse = self.env['stock.warehouse'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.warehouse:
            self.warehouse = self.env['stock.warehouse'].create({
                'name': 'Triage Test WH', 'code': 'TRWH', 'company_id': self.company.id,
            })
        self.config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.config:
            self.config = self.env['smart.reorder.config'].create({'company_id': self.company.id})

    def _set_quant(self, product, qty_delta):
        self.env['stock.quant']._update_available_quantity(
            product, self.warehouse.lot_stock_id, qty_delta
        )

    def _generate(self):
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company.id], warehouse_ids=[self.warehouse.id],
            include_zero_demand=True, trigger_type='manual',
        )

    def test_needs_vendor_assignment_true_when_no_vendor(self):
        product = self.env['product.product'].create({
            'name': 'Triage No Vendor Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, -3.0)
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertTrue(suggestion.needs_vendor_assignment)

    def test_needs_vendor_assignment_true_when_temp_vendor(self):
        temp_vendor = self.env['res.partner'].create({'name': 'Triage Temp Vendor'})
        self.config.write({'temp_vendor_ids': [(6, 0, [temp_vendor.id])]})
        product = self.env['product.product'].create({
            'name': 'Triage Temp Vendor Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': temp_vendor.id,
        })
        self._set_quant(product, -3.0)
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertTrue(suggestion.needs_vendor_assignment)

    def test_needs_vendor_assignment_false_with_real_vendor(self):
        real_vendor = self.env['res.partner'].create({'name': 'Triage Real Vendor'})
        product = self.env['product.product'].create({
            'name': 'Triage Real Vendor Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': real_vendor.id,
        })
        self._set_quant(product, -3.0)
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertFalse(suggestion.needs_vendor_assignment)

    def test_dead_stock_critical_review_true(self):
        product = self.env['product.product'].create({
            'name': 'Triage Dead Critical Widget', 'type': 'product', 'standard_price': 1.0,
        })
        partner = self.env['res.partner'].create({'name': 'Triage Dead Critical Customer'})
        old_date = fields.Date.today() - relativedelta(months=8)
        order = self.env['sale.order'].create({
            'partner_id': partner.id,
            'company_id': self.company.id,
            'warehouse_id': self.warehouse.id,
            'date_order': datetime.combine(old_date, datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': product.id, 'product_uom_qty': 5.0, 'price_unit': 10.0,
            })],
        })
        order.action_confirm()
        order.order_line.qty_delivered = 5.0

        self._set_quant(product, -3.0)  # negative now, no RECENT sales
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertEqual(suggestion.urgency, 'critical')
        self.assertTrue(suggestion.is_dead_stock)
        self.assertTrue(suggestion.dead_stock_critical_review)

    def test_dead_stock_critical_review_false_when_not_critical(self):
        product = self.env['product.product'].create({
            'name': 'Triage Not Critical Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, 5.0)  # positive, dead, but not critical
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion)
        self.assertNotEqual(suggestion.urgency, 'critical')
        self.assertFalse(suggestion.dead_stock_critical_review)

    def test_vendor_assign_wizard_creates_supplierinfo_and_updates_suggestion(self):
        product = self.env['product.product'].create({
            'name': 'Triage Wizard Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, -3.0)
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        self.assertTrue(suggestion.needs_vendor_assignment)

        new_vendor = self.env['res.partner'].create({'name': 'Newly Assigned Vendor'})
        wizard = self.env['smart.reorder.vendor.assign.wizard'].with_context(
            active_ids=[suggestion.id]
        ).create({'vendor_id': new_vendor.id})
        wizard.action_assign_vendor()

        supplierinfo = self.env['product.supplierinfo'].search([
            ('product_tmpl_id', '=', product.product_tmpl_id.id),
            ('partner_id', '=', new_vendor.id),
        ])
        self.assertTrue(supplierinfo, 'must create a real product.supplierinfo, not just patch the suggestion')
        suggestion.invalidate_recordset()
        self.assertEqual(suggestion.vendor_id, new_vendor)
        self.assertFalse(suggestion.needs_vendor_assignment)

    def test_vendor_assign_wizard_requires_manager_group(self):
        product = self.env['product.product'].create({
            'name': 'Triage Wizard Guard Widget', 'type': 'product', 'standard_price': 1.0,
        })
        self._set_quant(product, -3.0)
        self._generate()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id), ('warehouse_id', '=', self.warehouse.id),
        ])
        basic_user = self.env['res.users'].create({
            'name': 'Triage Basic User',
            'login': 'triage_basic_user_test',
            'groups_id': [(6, 0, [
                self.env.ref('base.group_user').id,
                self.env.ref('smart_reorder_advisor.group_smart_reorder_user').id,
            ])],
        })
        new_vendor = self.env['res.partner'].create({'name': 'Guard Test Vendor'})
        with self.assertRaises(AccessError):
            wizard = self.env['smart.reorder.vendor.assign.wizard'].with_user(basic_user).with_context(
                active_ids=[suggestion.id]
            ).create({'vendor_id': new_vendor.id})
            wizard.action_assign_vendor()


@tagged('post_install', '-at_install')
class TestActionCenter(TransactionCase):
    """Recommendation: attention_flag_count / attention_reasons must combine
    all six triage flags reactively — correct the instant any flag changes,
    not just after the next generate_suggestions() run."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.warehouse = self.env['stock.warehouse'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.warehouse:
            self.warehouse = self.env['stock.warehouse'].create({
                'name': 'Action Center Test WH', 'code': 'ACWH', 'company_id': self.company.id,
            })

    def _make_suggestion(self, **vals):
        product = self.env['product.product'].create({
            'name': f'Action Center Widget {vals.get("urgency", "x")} {len(vals)}',
            'type': 'product', 'standard_price': 1.0,
        })
        base = {
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': product.id, 'urgency': 'normal',
        }
        base.update(vals)
        return self.env['smart.reorder.suggestion'].create(base)

    def test_no_flags_gives_zero_count_and_no_reasons(self):
        suggestion = self._make_suggestion()
        self.assertEqual(suggestion.attention_flag_count, 0)
        self.assertFalse(suggestion.attention_reasons)

    def test_single_flag_counts_and_names_it(self):
        suggestion = self._make_suggestion(needs_review=True, needs_review_reason='swing')
        self.assertEqual(suggestion.attention_flag_count, 1)
        self.assertEqual(suggestion.attention_reasons, 'Needs Review')

    def test_multiple_flags_stack_and_all_are_named(self):
        suggestion = self._make_suggestion(
            needs_review=True, is_stale=True, has_price_discrepancy=True,
        )
        self.assertEqual(suggestion.attention_flag_count, 3)
        self.assertIn('Needs Review', suggestion.attention_reasons)
        self.assertIn('Stale Data', suggestion.attention_reasons)
        self.assertIn('Price Discrepancy', suggestion.attention_reasons)

    def test_recomputes_reactively_when_a_flag_is_cleared(self):
        suggestion = self._make_suggestion(needs_review=True, is_stale=True)
        self.assertEqual(suggestion.attention_flag_count, 2)
        suggestion.write({'is_stale': False})
        self.assertEqual(suggestion.attention_flag_count, 1)
        self.assertEqual(suggestion.attention_reasons, 'Needs Review')

    def test_action_center_domain_finds_only_flagged_suggestions(self):
        flagged = self._make_suggestion(needs_review=True)
        unflagged = self._make_suggestion()
        results = self.env['smart.reorder.suggestion'].search([('attention_flag_count', '>', 0)])
        self.assertIn(flagged, results)
        self.assertNotIn(unflagged, results)


class TestReorderPriceBreaksAltVendors(TransactionCase):

    def test_price_breaks_and_alt_vendors(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })

        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'alt_vendor_lead_margin_days': 5,
            })
        else:
            config.write({'alt_vendor_lead_margin_days': 5})

        product = self.env['product.product'].create({
            'name': 'Price Break Part',
            'type': 'product',
            'standard_price': 5.0,
        })
        # This test is about vendor/price-break selection, not demand
        # robustness — both sales below land in the same single month, which
        # would otherwise be excluded as a one-time-order spike (Fix 2) and
        # zero out the forecast the price-break tiers are supposed to react
        # to. bulk_regular bypasses that (plain average, no exclusion),
        # matching what this test's quantities/tiers were designed against.
        product.product_tmpl_id.reorder_behavior = 'bulk_regular'

        primary_vendor = self.env['res.partner'].create({'name': 'Primary Vendor'})
        alt_vendor = self.env['res.partner'].create({'name': 'Alternative Vendor'})

        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': primary_vendor.id,
            'price': 10.0,
            'min_qty': 1.0,
            'delay': 30,
            'sequence': 10,
        })
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': primary_vendor.id,
            'price': 8.0,
            'min_qty': 50.0,
            'delay': 25,
            'sequence': 10,
        })

        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': alt_vendor.id,
            'price': 12.0,
            'min_qty': 1.0,
            'delay': 15,
            'sequence': 20,
        })

        sale_date = date(2026, 6, 15)
        so = self.env['sale.order'].create({
            'partner_id': primary_vendor.id,
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'date_order': sale_date,
            'order_line': [(0, 0, {
                'product_id': product.id,
                'product_uom_qty': 20.0,
                'price_unit': 20.0,
            })],
        })
        so.action_confirm()
        so.order_line.qty_delivered = 20.0

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )

        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])
        self.assertTrue(suggestion, "Suggestion should be generated")
        self.assertEqual(suggestion.vendor_price, 10.0)
        self.assertEqual(suggestion.alt_vendor_id.id, alt_vendor.id)
        self.assertEqual(suggestion.alt_vendor_lead_days, 15)
        self.assertIn("Alternative vendor available: Alternative Vendor", suggestion.notes)
        self.assertIn("can deliver in 15 days", suggestion.notes)

        so_large = self.env['sale.order'].create({
            'partner_id': primary_vendor.id,
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'date_order': sale_date,
            'order_line': [(0, 0, {
                'product_id': product.id,
                'product_uom_qty': 120.0,
                'price_unit': 20.0,
            })],
        })
        so_large.action_confirm()
        so_large.order_line.qty_delivered = 120.0

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )
        suggestion.invalidate_recordset()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])
        self.assertEqual(suggestion.vendor_price, 8.0)
        self.assertEqual(suggestion.alt_vendor_id.id, alt_vendor.id)
        self.assertEqual(suggestion.alt_vendor_lead_days, 15)

        picking = self.env['stock.picking'].create({
            'picking_type_id': self.env['stock.picking.type'].search([('code', '=', 'incoming'), ('warehouse_id', '=', warehouse.id)], limit=1).id,
            'location_id': self.env.ref('stock.stock_location_suppliers').id,
            'location_dest_id': warehouse.lot_stock_id.id,
            'company_id': company.id,
        })
        move = self.env['stock.move'].create({
            'name': 'Test Move',
            'product_id': product.id,
            'product_uom_qty': 500.0,
            'product_uom': product.uom_id.id,
            'location_id': picking.location_id.id,
            'location_dest_id': picking.location_dest_id.id,
            'picking_id': picking.id,
        })
        picking.action_confirm()
        picking.action_assign()
        move.quantity = 500.0
        picking.button_validate()

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )
        suggestion.invalidate_recordset()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])
        
        self.assertEqual(suggestion.qty_on_hand, 500.0)
        self.assertEqual(suggestion.suggested_reorder_qty, 0.0)
        self.assertNotIn("Alternative vendor available", suggestion.notes or "")

        # 4. If alt_vendor_lead_margin_days is 0 (disabled), it should not scan/find alternative vendors
        config.write({'alt_vendor_lead_margin_days': 0})
        # Reset standard stock so it triggers a suggested quantity again
        self.env['stock.quant'].search([
            ('product_id', '=', product.id),
            ('location_id', '=', warehouse.lot_stock_id.id),
        ]).write({'quantity': 0.0})
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )
        suggestion.invalidate_recordset()
        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])
        self.assertFalse(suggestion.alt_vendor_id, "Alt vendor should be disabled when margin is 0")
        self.assertNotIn("Alternative vendor available", suggestion.notes or "")


class TestReorderProvisionalNegativeStock(TransactionCase):

    def test_provisional_negative_stock_flag(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })

        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'auto_flag_on_negative': True,
            })
        else:
            config.write({'auto_flag_on_negative': True})

        product = self.env['product.product'].create({
            'name': 'Provisional Test Part',
            'type': 'product',
            'standard_price': 5.0,
        })
        # Part 3 of this test drives a single 12-unit month through a full
        # generate_suggestions() run and expects a plain average (12/6=2.0).
        # A single month that size would otherwise be excluded as a one-time
        # spike (Fix 2), zeroing the forecast — bulk_regular bypasses that.
        # _flag_negative_stock_product() (parts 1-2) doesn't consult
        # reorder_behavior at all, so this has no effect on those assertions.
        product.product_tmpl_id.reorder_behavior = 'bulk_regular'

        supplier = self.env['res.partner'].create({'name': 'Provisional supplier'})
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': product.product_tmpl_id.id,
            'partner_id': supplier.id,
            'price': 10.0,
            'min_qty': 10.0,  # MOQ = 10
            'delay': 30,
        })

        # Set stock negative: let's mock it by editing standard stock quants or manually driving negative.
        # But we can also simulate the flag call directly:
        # Since _flag_negative_stock_product reads quants, we create a quant with quantity -3.0
        # In Odoo, negative quants represent negative stock.
        self.env['stock.quant'].create({
            'product_id': product.id,
            'location_id': warehouse.lot_stock_id.id,
            'quantity': -3.0,
        })

        # 1. No prior suggestion exists
        self.env['smart.reorder.suggestion']._flag_negative_stock_product(
            product.id, warehouse.id, company.id
        )

        suggestion = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', product.id),
            ('warehouse_id', '=', warehouse.id),
        ])
        self.assertTrue(suggestion, "Suggestion should be generated")
        self.assertTrue(suggestion.is_provisional, "Should be marked as provisional")
        self.assertEqual(suggestion.qty_on_hand, -3.0)
        # Suggested qty should be absolute negative qty (3.0) raised to vendor minimum MOQ (10) -> 10
        self.assertEqual(suggestion.suggested_reorder_qty, 10.0)

        # 2. Existing suggestion exists
        # Update suggestion to simulate a previous analysis run that calculated some demand
        suggestion.write({
            'avg_monthly_demand': 15.0,
            'lead_time_months': 1.0,
            'safety_buffer_months': 1.0,
            'order_cycle_months': 1.0,
            'moq': 10.0,
            'qty_incoming': 0.0,
            'qty_outgoing': 0.0,
            'is_provisional': False,
        })

        # Let's call the hook again with a different negative qty (e.g. -5.0)
        self.env['stock.quant'].search([
            ('product_id', '=', product.id),
            ('location_id', '=', warehouse.lot_stock_id.id),
        ]).write({'quantity': -5.0})

        self.env['smart.reorder.suggestion']._flag_negative_stock_product(
            product.id, warehouse.id, company.id
        )

        suggestion.invalidate_recordset()
        self.assertTrue(suggestion.is_provisional, "Should be flagged provisional again")
        self.assertEqual(suggestion.qty_on_hand, -5.0)
        # Calculations:
        # avg_monthly = 15.0, min_level = 15 * (1 + 1) = 30
        # max_level = 30 + 15 * 1 = 45
        # raw_qty = 45 - (-5) = 50. MOQ = 10, so raised to minimum of 10 is 50.
        # Max of 50 and abs(-5) raised to MOQ 10 is 50.
        self.assertEqual(suggestion.suggested_reorder_qty, 50.0)

        # 3. Next full analysis clears the provisional flag
        # We need to run full analysis. Let's create a sale order to simulate sales history first
        so = self.env['sale.order'].create({
            'partner_id': supplier.id,
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'date_order': date(2026, 6, 15),
            'order_line': [(0, 0, {
                'product_id': product.id,
                'product_uom_qty': 12.0,
                'price_unit': 20.0,
            })],
        })
        so.action_confirm()
        so.order_line.qty_delivered = 12.0

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )

        suggestion.invalidate_recordset()
        self.assertFalse(suggestion.is_provisional, "Provisional flag should be cleared by full analysis")
        # Numbers should be replaced with actual forecast numbers
        # avg_monthly = 2.0 (since 12 / 6 = 2)
        # Net available = -5.0. Min level = 2 * (1 + 1) = 4
        # max_level = 4 + 2 * 0.25 = 4.5
        # raw_qty = 4.5 - (-5.0) = 9.5 -> raised to vendor minimum of 10.
        self.assertEqual(suggestion.suggested_reorder_qty, 10.0)
        self.assertEqual(suggestion.avg_monthly_demand, 2.0)

    def test_combined_picking_notifications(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })

        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'auto_flag_on_negative': True,
            })
        else:
            config.write({'auto_flag_on_negative': True})

        product1 = self.env['product.product'].create({
            'name': 'Combined Alert Part 1',
            'type': 'product',
            'standard_price': 5.0,
        })
        product2 = self.env['product.product'].create({
            'name': 'Combined Alert Part 2',
            'type': 'product',
            'standard_price': 5.0,
        })

        self.env['stock.quant'].create({
            'product_id': product1.id,
            'location_id': warehouse.lot_stock_id.id,
            'quantity': -3.0,
        })
        self.env['stock.quant'].create({
            'product_id': product2.id,
            'location_id': warehouse.lot_stock_id.id,
            'quantity': -5.0,
        })

        picking = self.env['stock.picking'].create({
            'picking_type_id': self.env['stock.picking.type'].search([('code', '=', 'outgoing'), ('warehouse_id', '=', warehouse.id)], limit=1).id,
            'location_id': warehouse.lot_stock_id.id,
            'location_dest_id': self.env.ref('stock.stock_location_customers').id,
            'company_id': company.id,
        })
        self.env['stock.move'].create({
            'name': 'Test Out Move 1',
            'product_id': product1.id,
            'product_uom_qty': 3.0,
            'product_uom': product1.uom_id.id,
            'location_id': picking.location_id.id,
            'location_dest_id': picking.location_dest_id.id,
            'picking_id': picking.id,
        })
        self.env['stock.move'].create({
            'name': 'Test Out Move 2',
            'product_id': product2.id,
            'product_uom_qty': 5.0,
            'product_uom': product2.uom_id.id,
            'location_id': picking.location_id.id,
            'location_dest_id': picking.location_dest_id.id,
            'picking_id': picking.id,
        })

        notification_calls = []
        def mock_send_notifications(comp, conf, count, warehouse_id=None):
            notification_calls.append((comp.id, count, warehouse_id))
            return True

        with patch.object(type(self.Suggestion), '_send_notifications', side_effect=mock_send_notifications):
            picking._action_done()

        self.assertEqual(len(notification_calls), 1)
        self.assertEqual(notification_calls[0][1], 2)
        self.assertEqual(notification_calls[0][2], warehouse.id)


@tagged('post_install', '-at_install')
class TestMultiCompanyProductCost(TransactionCase):
    """standard_price is a company-dependent property field. Both the bulk
    cost fetch in _fetch_warehouse_data() (Q4) and the cost read in
    _flag_negative_stock_product() must resolve it within the context of the
    company being analyzed/flagged — not whatever company the ORM defaults to
    under sudo() — or a multi-company setup prices suggestions with the wrong
    company's average cost."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company_a = cls.env.company
        cls.company_b = cls.env['res.company'].create({'name': 'Cost Test Company B'})

        cls.warehouse_a = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company_a.id)], limit=1
        )
        if not cls.warehouse_a:
            cls.warehouse_a = cls.env['stock.warehouse'].create({
                'name': 'Cost Test WH A', 'code': 'CTWA', 'company_id': cls.company_a.id,
            })
        cls.warehouse_b = cls.env['stock.warehouse'].create({
            'name': 'Cost Test WH B', 'code': 'CTWB', 'company_id': cls.company_b.id,
        })

        for comp, wh in ((cls.company_a, cls.warehouse_a), (cls.company_b, cls.warehouse_b)):
            if not cls.env['smart.reorder.config'].search([('company_id', '=', comp.id)], limit=1):
                cls.env['smart.reorder.config'].create({'company_id': comp.id})

        # A single product shared across both companies, with a different
        # company-dependent standard_price per company.
        cls.product = cls.env['product.product'].create({
            'name': 'Multi-Company Cost Widget',
            'type': 'product',
        })
        cls.product.with_company(cls.company_a).standard_price = 10.0
        cls.product.with_company(cls.company_b).standard_price = 40.0

        cls.partner = cls.env['res.partner'].create({'name': 'Cost Test Customer'})

    def test_generate_suggestions_uses_own_company_cost(self):
        for comp, wh in ((self.company_a, self.warehouse_a), (self.company_b, self.warehouse_b)):
            so = self.env['sale.order'].create({
                'partner_id': self.partner.id,
                'company_id': comp.id,
                'warehouse_id': wh.id,
                'date_order': datetime.combine(date.today(), datetime.min.time()),
                'order_line': [(0, 0, {
                    'product_id': self.product.id,
                    'product_uom_qty': 5.0,
                    'price_unit': 1.0,
                })],
            })
            so.action_confirm()
            so.order_line.qty_delivered = 5.0

        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company_a.id], warehouse_ids=[self.warehouse_a.id]
        )
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[self.company_b.id], warehouse_ids=[self.warehouse_b.id]
        )

        suggestion_a = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', self.product.id), ('warehouse_id', '=', self.warehouse_a.id),
        ])
        suggestion_b = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', self.product.id), ('warehouse_id', '=', self.warehouse_b.id),
        ])
        self.assertEqual(len(suggestion_a), 1)
        self.assertEqual(len(suggestion_b), 1)
        self.assertEqual(suggestion_a.product_cost, 10.0, "Company A's suggestion must use Company A's cost")
        self.assertEqual(suggestion_b.product_cost, 40.0, "Company B's suggestion must use Company B's cost")
        self.assertEqual(
            suggestion_a.reorder_value,
            suggestion_a.suggested_reorder_qty * 10.0,
        )
        self.assertEqual(
            suggestion_b.reorder_value,
            suggestion_b.suggested_reorder_qty * 40.0,
        )

    def test_flag_negative_stock_uses_own_company_cost(self):
        self.env['stock.quant'].create({
            'product_id': self.product.id,
            'location_id': self.warehouse_a.lot_stock_id.id,
            'quantity': -2.0,
        })
        self.env['stock.quant'].create({
            'product_id': self.product.id,
            'location_id': self.warehouse_b.lot_stock_id.id,
            'quantity': -2.0,
        })

        self.env['smart.reorder.suggestion']._flag_negative_stock_product(
            self.product.id, self.warehouse_a.id, self.company_a.id
        )
        self.env['smart.reorder.suggestion']._flag_negative_stock_product(
            self.product.id, self.warehouse_b.id, self.company_b.id
        )

        suggestion_a = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', self.product.id), ('warehouse_id', '=', self.warehouse_a.id),
        ])
        suggestion_b = self.env['smart.reorder.suggestion'].search([
            ('product_id', '=', self.product.id), ('warehouse_id', '=', self.warehouse_b.id),
        ])
        self.assertEqual(suggestion_a.product_cost, 10.0)
        self.assertEqual(suggestion_b.product_cost, 40.0)


class TestReorderForecastBacktesting(TransactionCase):

    def test_forecast_backtesting_flow(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })

        # Set retention configuration
        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'snapshot_retention_months': 3,
            })
        else:
            config.write({'snapshot_retention_months': 3})

        product = self.env['product.product'].create({
            'name': 'Backtest Test Product',
            'type': 'product',
            'standard_price': 10.0,
        })

        # 1. Create snapshots manually to simulate old runs (older than 1 lead time)
        # Snapshot date: 35 days ago, lead time: 30 days
        snap_date_old = date.today() - timedelta(days=35)
        old_snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': snap_date_old,
            'forecast_demand': 10.0,  # 10 units/month forecast
            'confidence': 85.0,
            'lead_time_days': 30,
            'abc_class': 'A',
            'evaluated': False,
        })

        # Snapshot date: 15 days ago, lead time: 30 days (too new to evaluate)
        snap_date_new = date.today() - timedelta(days=15)
        new_snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': snap_date_new,
            'forecast_demand': 5.0,
            'confidence': 90.0,
            'lead_time_days': 30,
            'abc_class': 'B',
            'evaluated': False,
        })

        # Snapshot date: 120 days ago (older than 3 months retention)
        snap_date_expired = date.today() - timedelta(days=120)
        expired_snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': snap_date_expired,
            'forecast_demand': 20.0,
            'confidence': 70.0,
            'lead_time_days': 30,
            'abc_class': 'C',
            'evaluated': True,
        })

        # 2. Record sales during the lead time period for the old snapshot
        # Period: [today-35, today-5]
        # Let's say we sold 8 units
        so = self.env['sale.order'].create({
            'partner_id': self.env['res.partner'].create({'name': 'Test Cust'}).id,
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'date_order': date.today() - timedelta(days=20),
            'order_line': [(0, 0, {
                'product_id': product.id,
                'product_uom_qty': 8.0,
                'price_unit': 15.0,
            })],
        })
        so.action_confirm()
        # Deliver the units
        for line in so.order_line:
            line.qty_delivered = 8.0

        # 3. Score Snapshots
        self.env['smart.reorder.forecast.snapshot']._score_snapshots()

        old_snap.invalidate_recordset()
        new_snap.invalidate_recordset()

        self.assertTrue(old_snap.evaluated, "Old snapshot should be evaluated")
        self.assertFalse(new_snap.evaluated, "New snapshot should not be evaluated yet")

        # Calculations:
        # Forecast demand = 10.0 monthly. Period = 30 days -> forecasted qty = 10 * (30/30) = 10.0 units.
        # Actual sales = 8.0 units.
        # APE = abs(10.0 - 8.0) / 10.0 = 0.2 -> 20.0%
        self.assertEqual(old_snap.actual_sales, 8.0)
        self.assertEqual(old_snap.absolute_error_pct, 20.0)

        # 4. Verify retention settings pruning on generate suggestions
        # Create suggestion to run generation
        self.env['smart.reorder.suggestion'].generate_suggestions(
            company_ids=[company.id], warehouse_ids=[warehouse.id]
        )

        # The expired snapshot (120 days ago) should be deleted by the retention purge
        still_exists = self.env['smart.reorder.forecast.snapshot'].search([('id', '=', expired_snap.id)])
        self.assertFalse(still_exists, "Expired snapshot should have been deleted")

        # 5. Dashboard aggregation check
        dashboard_data = self.env['smart.reorder.suggestion'].get_dashboard_data(warehouse_id=warehouse.id)
        self.assertIn('backtest', dashboard_data)
        backtest_res = dashboard_data['backtest']
        self.assertEqual(backtest_res['overall_mape'], 20.0)
        self.assertEqual(backtest_res['mape_by_abc']['A'], 20.0)


@tagged('post_install', '-at_install')
class TestSnapshotEvaluatedNotEditable(TransactionCase):
    """A manager clicking a toggle in the snapshot list must not be able to flip
    'Evaluated' without running the scorer — that would inject a zero-error row
    into the MAPE averages on the dashboard. Only _score_snapshots() / the
    Evaluate button may set this field."""

    def test_evaluated_field_is_readonly_at_model_level(self):
        field_info = self.env['smart.reorder.forecast.snapshot'].fields_get(['evaluated'])
        self.assertTrue(
            field_info['evaluated']['readonly'],
            "'evaluated' must be readonly at the model level so no view can "
            "accidentally expose it as writable"
        )

    def test_tree_view_does_not_use_toggle_widget(self):
        tree_view = self.env.ref('smart_reorder_advisor.view_smart_reorder_forecast_snapshot_tree')
        self.assertNotIn(
            'boolean_toggle', tree_view.arch_db,
            "the snapshot list must render 'evaluated' as a plain read-only "
            "boolean, not a clickable toggle"
        )

    def test_scoring_still_sets_evaluated_correctly(self):
        # The readonly field must not block the scorer's own ORM write.
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH Eval', 'code': 'TWHE', 'company_id': company.id,
            })
        product = self.env['product.product'].create({
            'name': 'Evaluated Field Test Product', 'type': 'product',
        })
        snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': date.today() - timedelta(days=45),
            'forecast_demand': 10.0,
            'lead_time_days': 30,
            'evaluated': False,
        })
        self.assertFalse(snap.evaluated)
        self.env['smart.reorder.forecast.snapshot']._score_snapshots()
        snap.invalidate_recordset()
        self.assertTrue(snap.evaluated)


@tagged('post_install', '-at_install')
class TestForecastSnapshotCompositeIndex(TransactionCase):
    """Fix 12: the scorer's pending-snapshot search and the dashboard's
    back-test aggregation both filter on evaluated + company (+ warehouse/
    product) — a composite index matching that access pattern must exist,
    created in ForecastSnapshot.init() at module initialization."""

    def test_composite_eval_scope_index_exists_with_expected_columns(self):
        self.env.cr.execute("""
            SELECT indexdef FROM pg_indexes
            WHERE tablename = 'smart_reorder_forecast_snapshot'
              AND indexname = 'smart_reorder_forecast_snapshot_eval_scope_idx'
        """)
        row = self.env.cr.fetchone()
        self.assertIsNotNone(
            row,
            'composite index on (evaluated, company_id, warehouse_id, product_id) '
            'must exist on smart_reorder_forecast_snapshot'
        )
        indexdef = row[0]
        for col in ('evaluated', 'company_id', 'warehouse_id', 'product_id'):
            self.assertIn(col, indexdef)

        # Column order matters: evaluated must lead so both the scorer's
        # evaluated-only search and the dashboard's evaluated+company(+...)
        # filter can use a leading-column index scan.
        positions = [indexdef.index(col) for col in ('evaluated', 'company_id', 'warehouse_id', 'product_id')]
        self.assertEqual(positions, sorted(positions), 'evaluated, company_id, warehouse_id, product_id must be in that order')


@tagged('post_install', '-at_install')
class TestReorderPureFunction(TransactionCase):
    """
    Direct unit tests for the pure per-product calculation function `_calculate_product_suggestion`.
    These pass plain Python datatypes to ensure correct formula application without DB calls.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.config_data = {
            'default_lead_time_months': 1.0,
            'track_vendor_performance': False,
            'dead_stock_months': 6,
            'flag_dead_stock': True,
            'safety_buffer_months': 1.0,
            'order_cycle_months': 2.0,
            'alt_vendor_lead_margin_days': 5,
            'abc_a_threshold': 5.0,
            'abc_b_threshold': 1.0,
            'min_spike_size': 10.0,
            'overstock_ceiling_months': 12.0,
            'transfer_surplus_threshold': 6.0,
            'default_internal_transfer_days': 3,
        }
        cls.dates = {
            'date_from': date.today() - timedelta(days=90),
            'date_to': date.today(),
            'analysis_months': 3.0,
            'comparison_months': 3.0,
            'month_starts': [
                date.today() - timedelta(days=30),
                date.today() - timedelta(days=60),
                date.today() - timedelta(days=90),
            ],
        }

    def test_pure_function_normal_part(self):
        # Normal part: monthly_series = [10.0, 10.0, 10.0] -> avg = 10.0.
        # On hand = 5.0. Incoming = 0. Outgoing = 0.
        # Lead time = 1 month (30 days stated). Safety buffer = 1 month. Order cycle = 2 months.
        # Min level = 10.0 * (1 + 1) = 20.0.
        # Max level = 20.0 + 10.0 * 2 = 40.0.
        # Triggered because qty_available (5.0) < min_level (20.0).
        # Raw reorder qty = 40.0 - 5.0 = 35.0.
        # MOQ = 10.0 -> Suggested qty = 35.0 (raised to vendor minimum of 10.0).
        res = self.Suggestion._calculate_product_suggestion(
            product_id=1,
            product_code='NORM',
            product_name='Normal Part',
            tmpl_id=10,
            is_superseded=False,
            successor_display_name=False,
            company_id=1,
            warehouse_id=1,
            dates=self.dates,
            config_data=self.config_data,
            warehouses_list=[{'id': 1, 'name': 'WH 1'}],
            monthly_series=[10.0, 10.0, 10.0],
            qty_on_hand=5.0,
            qty_incoming=0.0,
            qty_outgoing=0.0,
            cost=1.5,
            last_sale=date.today(),
            current_month_sales=0.0,
            prev_qty=30.0,
            ly_qty=30.0,
            tmpl_suppliers=[],
            primary_vendor_info=(101, 30.0, 10.0, 5.0, False), # partner_id, delay, min_qty, price, currency_id
            actual_avg_days=0.0,
            overdue_lines=[],
            predecessors=set(),
            predecessor_names=[],
            global_stocks={1: 5.0},
            global_sales={1: 30.0},
            lane_lead_times={},
            partner_names_map={101: 'Primary Vendor'},
            currency_convert_fn=lambda price, currency_id: price
        )
        self.assertEqual(res['suggested_reorder_qty'], 35.0)
        self.assertEqual(res['min_stock_level'], 20.0)
        self.assertEqual(res['max_stock_level'], 40.0)
        self.assertEqual(res['urgency'], 'urgent')
        self.assertEqual(res['abc_class'], 'A')

    def test_pure_function_negative_stock_part(self):
        # Negative-stock part: on hand = -5.0.
        # Even with zero demand and zero forecast, negative stock triggers reorder.
        res = self.Suggestion._calculate_product_suggestion(
            product_id=2,
            product_code='NEG',
            product_name='Negative Stock Part',
            tmpl_id=11,
            is_superseded=False,
            successor_display_name=False,
            company_id=1,
            warehouse_id=1,
            dates=self.dates,
            config_data=self.config_data,
            warehouses_list=[{'id': 1, 'name': 'WH 1'}],
            monthly_series=[0.0, 0.0, 0.0],
            qty_on_hand=-5.0,
            qty_incoming=0.0,
            qty_outgoing=0.0,
            cost=1.5,
            last_sale=date.today() - timedelta(days=200),
            current_month_sales=0.0,
            prev_qty=0.0,
            ly_qty=0.0,
            tmpl_suppliers=[],
            primary_vendor_info=False,
            actual_avg_days=0.0,
            overdue_lines=[],
            predecessors=set(),
            predecessor_names=[],
            global_stocks={1: -5.0},
            global_sales={1: 0.0},
            lane_lead_times={},
            partner_names_map={},
            currency_convert_fn=lambda price, currency_id: price
        )
        self.assertEqual(res['urgency'], 'critical')
        self.assertGreater(res['suggested_reorder_qty'], 0.0)

    def test_pure_function_dead_part(self):
        # Dead part: no sales in dates or last_sale is very old.
        # Config dead_stock_months is 6, last_sale is 300 days ago.
        res = self.Suggestion._calculate_product_suggestion(
            product_id=3,
            product_code='DEAD',
            product_name='Dead Part',
            tmpl_id=12,
            is_superseded=False,
            successor_display_name=False,
            company_id=1,
            warehouse_id=1,
            dates=self.dates,
            config_data=self.config_data,
            warehouses_list=[{'id': 1, 'name': 'WH 1'}],
            monthly_series=[0.0, 0.0, 0.0],
            qty_on_hand=10.0,
            qty_incoming=0.0,
            qty_outgoing=0.0,
            cost=1.5,
            last_sale=date.today() - timedelta(days=300),
            current_month_sales=0.0,
            prev_qty=0.0,
            ly_qty=0.0,
            tmpl_suppliers=[],
            primary_vendor_info=False,
            actual_avg_days=0.0,
            overdue_lines=[],
            predecessors=set(),
            predecessor_names=[],
            global_stocks={1: 10.0},
            global_sales={1: 0.0},
            lane_lead_times={},
            partner_names_map={},
            currency_convert_fn=lambda price, currency_id: price
        )
        self.assertTrue(res['is_dead_stock'])
        self.assertEqual(res['urgency'], 'dead')

    def test_pure_function_transfer_eligible_part(self):
        # Transfer eligible part: urgent/critical in WH 1, but WH 2 has surplus stock (mos_other > 6.0).
        # WH 2 stock: 100.0, demand in WH 2: 1.0 (sold_other = 3.0 / 3 months = 1.0 units/month).
        # mos_other = 100.0 / 1.0 = 100.0 (> 6.0).
        # Expect transfer source warehouse to be WH 2.
        res = self.Suggestion._calculate_product_suggestion(
            product_id=4,
            product_code='TRANS',
            product_name='Transfer Eligible Part',
            tmpl_id=13,
            is_superseded=False,
            successor_display_name=False,
            company_id=1,
            warehouse_id=1,
            dates=self.dates,
            config_data=self.config_data,
            warehouses_list=[{'id': 1, 'name': 'WH 1'}, {'id': 2, 'name': 'WH 2'}],
            monthly_series=[10.0, 10.0, 10.0],
            qty_on_hand=5.0,
            qty_incoming=0.0,
            qty_outgoing=0.0,
            cost=1.5,
            last_sale=date.today(),
            current_month_sales=0.0,
            prev_qty=30.0,
            ly_qty=30.0,
            tmpl_suppliers=[],
            primary_vendor_info=False,
            actual_avg_days=0.0,
            overdue_lines=[],
            predecessors=set(),
            predecessor_names=[],
            global_stocks={1: 5.0, 2: 100.0},
            global_sales={1: 30.0, 2: 3.0}, # WH 2 sold 3 units total -> 1.0 avg
            lane_lead_times={2: 5},
            partner_names_map={},
            currency_convert_fn=lambda price, currency_id: price
        )
        self.assertEqual(res['transfer_source_warehouse_id'], 2)
        self.assertEqual(res['transfer_lead_time_days'], 5)

    def test_pure_function_overdue_po_lines(self):
        """
        Regression test for the datetime-vs-date comparison bug fixed in
        _calculate_product_suggestion:  oldest_dt = min(...).date()

        Previously oldest_dt held a datetime.datetime object and
        `date.today() - oldest_dt` raised TypeError.  This test passes an
        overdue PO line (quantity overdue, delivery 10 days in the past) and
        asserts:
          1. The function returns without raising.
          2. The notes string contains the OVERDUE SUPPLY marker.
        """
        overdue_dt = datetime.now() - timedelta(days=10)
        res = self.Suggestion._calculate_product_suggestion(
            product_id=5,
            product_code='OVER',
            product_name='Overdue PO Part',
            tmpl_id=14,
            is_superseded=False,
            successor_display_name=False,
            company_id=1,
            warehouse_id=1,
            dates=self.dates,
            config_data=self.config_data,
            warehouses_list=[{'id': 1, 'name': 'WH 1'}],
            monthly_series=[5.0, 5.0, 5.0],
            qty_on_hand=2.0,
            qty_incoming=8.0,   # an open PO line — overdue
            qty_outgoing=0.0,
            cost=2.0,
            last_sale=date.today(),
            current_month_sales=0.0,
            prev_qty=15.0,
            ly_qty=15.0,
            tmpl_suppliers=[],
            primary_vendor_info=(102, 30.0, 5.0, 4.0, False),
            actual_avg_days=0.0,
            overdue_lines=[(8.0, overdue_dt)],   # <-- triggers the fixed code path
            predecessors=set(),
            predecessor_names=[],
            global_stocks={1: 2.0},
            global_sales={1: 15.0},
            lane_lead_times={},
            partner_names_map={102: 'Overdue Vendor'},
            currency_convert_fn=lambda price, currency_id: price
        )
        # Must not raise; basic sanity checks
        self.assertIsNotNone(res)
        # The overdue block must have been entered and written into notes
        self.assertIn('OVERDUE SUPPLY', res.get('notes', ''))


@tagged('post_install', '-at_install')
class TestAutomatedWeeklyEmailReport(TransactionCase):
    """Task 4: weekly email delivery — gated on Critical/Urgent existing
    (reusing the existing 'Notify Only for Critical / Urgent Items' toggle),
    and now attaches the new boss-friendly report instead of the old
    all-columns technical summary."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Suggestion = cls.env['smart.reorder.suggestion']
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        cls.user = cls.env['res.users'].create({
            'name': 'Weekly Report Recipient',
            'login': 'weekly_report_recipient_test',
            'email': 'weekly_report_recipient_test@example.com',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id])],
        })
        cls.config = cls.env['smart.reorder.config'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.config:
            cls.config = cls.env['smart.reorder.config'].create({'company_id': cls.company.id})
        cls.config.write({
            'notify_user_ids': [(6, 0, [cls.user.id])],
            'critical_notify_only': True,
        })

    def _make_suggestion(self, urgency, reorder_needed=True):
        product = self.env['product.product'].create({
            'name': f'Weekly Report Widget {urgency} {id(self)}', 'type': 'product', 'standard_price': 1.0,
        })
        return self.Suggestion.create({
            'company_id': self.company.id, 'warehouse_id': self.warehouse.id,
            'product_id': product.id, 'urgency': urgency,
            'suggested_reorder_qty': 5.0, 'reorder_needed': reorder_needed,
        })

    def test_no_email_when_only_normal_items_and_critical_only_gate_on(self):
        self._make_suggestion('normal')
        mail_count_before = self.env['mail.mail'].search_count([])
        result = self.Suggestion._send_email_report(self.company, self.config)
        self.assertIsNone(result, 'must be a no-op, not an error, when nothing critical/urgent exists')
        self.assertEqual(self.env['mail.mail'].search_count([]), mail_count_before)

    def test_email_sent_and_uses_boss_report_when_critical_item_exists(self):
        self._make_suggestion('critical')
        with patch.object(
            type(self.env['ir.actions.report']), '_render_qweb_pdf',
            return_value=(b'%PDF-FAKE%', 'pdf'),
        ) as mock_render:
            result = self.Suggestion._send_email_report(self.company, self.config)
        self.assertTrue(result)
        mock_render.assert_called_once()
        report_ref = mock_render.call_args[0][0]
        self.assertEqual(
            report_ref, 'smart_reorder_advisor.action_report_boss_weekly_order',
            'must attach the new boss-friendly report, not the old technical summary'
        )

    def test_email_sent_when_only_urgent_item_exists(self):
        self._make_suggestion('urgent')
        with patch.object(
            type(self.env['ir.actions.report']), '_render_qweb_pdf',
            return_value=(b'%PDF-FAKE%', 'pdf'),
        ):
            result = self.Suggestion._send_email_report(self.company, self.config)
        self.assertTrue(result, 'Urgent (not just Critical) must also pass the gate')

    def test_critical_only_gate_disabled_sends_regardless(self):
        self.config.write({'critical_notify_only': False})
        self._make_suggestion('normal')
        with patch.object(
            type(self.env['ir.actions.report']), '_render_qweb_pdf',
            return_value=(b'%PDF-FAKE%', 'pdf'),
        ):
            result = self.Suggestion._send_email_report(self.company, self.config)
        self.assertTrue(result)

    def test_weekly_cron_enabled_by_default(self):
        cron = self.env.ref('smart_reorder_advisor.cron_smart_reorder_weekly')
        self.assertTrue(cron.active, 'Task 4: the weekly analysis cron must be active by default')

    def test_new_config_defaults_email_report_on(self):
        other_company = self.env['res.company'].create({'name': 'Task 4 Default Test Co.'})
        config = self.env['smart.reorder.config'].create({'company_id': other_company.id})
        self.assertTrue(config.send_email_report, 'Task 4: new configs must default to email-on')


@tagged('post_install', '-at_install')
class TestReorderObservability(TransactionCase):

    def test_observability_dashboard_and_hook_errors(self):
        company = self.env.company
        config = self.env['smart.reorder.config'].create({
            'company_id': company.id,
        })
        self.assertEqual(config.picking_hook_error_count, 0)

        # Increment picking hook error count
        config.write({'picking_hook_error_count': 5})

        # Create dummy cron log records
        CronLog = self.env['smart.reorder.cron.log']
        
        # Log 1: Completed
        CronLog.create({
            'company_id': company.id,
            'started_at': fields.Datetime.now() - timedelta(hours=2),
            'finished_at': fields.Datetime.now() - timedelta(hours=1, minutes=58),
            'status': 'completed',
        })
        # Log 2: Completed with Errors
        CronLog.create({
            'company_id': company.id,
            'started_at': fields.Datetime.now() - timedelta(hours=1),
            'finished_at': fields.Datetime.now() - timedelta(minutes=58),
            'status': 'completed_with_errors',
        })

        # Compute dashboard stats
        dashboard = self.env['smart.reorder.observability.dashboard'].create({
            'company_id': company.id,
        })
        self.assertEqual(dashboard.total_runs, 2)
        self.assertEqual(dashboard.failed_runs, 1)
        # failure_rate_30_days is 0.5 (1/2) for percentage widget format
        self.assertEqual(dashboard.failure_rate_30_days, 0.5)
        self.assertEqual(dashboard.picking_hook_errors, 5)

    def test_timeout_cap_observability(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })

        # We pass a pre-timed-out timestamp using _cron_start
        with patch('time.time', return_value=fields.Datetime.now().timestamp()):
            # Pass a _cron_start that is 46 minutes in the past
            self.Suggestion.generate_suggestions(
                company_ids=[company.id],
                warehouse_ids=[warehouse.id],
                _cron_start=fields.Datetime.now().timestamp() - 46 * 60
            )

        # Check that the log status is 'completed_with_errors' and contains 'TIME CAP'
        log = self.env['smart.reorder.cron.log'].search([('company_id', '=', company.id)], limit=1, order='started_at desc')
        self.assertEqual(log.status, 'completed_with_errors')
        self.assertIn('TIME CAP', log.error_notes)


@tagged('post_install', '-at_install')
class TestDataHealthDashboard(TransactionCase):
    """Recommendation: data-quality/setup health check — surfaces product
    master-data gaps (no cost, no vendor, no part number, placeholder-only
    vendor) that silently weaken forecasts without ever erroring."""

    def setUp(self):
        super().setUp()
        self.company = self.env.company
        self.config = self.env['smart.reorder.config'].search(
            [('company_id', '=', self.company.id)], limit=1
        )
        if not self.config:
            self.config = self.env['smart.reorder.config'].create({'company_id': self.company.id})

    def test_counts_missing_cost_and_part_number(self):
        self.env['product.product'].create({
            'name': 'Health Missing Cost Widget', 'type': 'product',
            'standard_price': 0.0, 'default_code': 'HLT-001',
        })
        self.env['product.product'].create({
            'name': 'Health Missing Part Number Widget', 'type': 'product',
            'standard_price': 5.0, 'default_code': False,
        })
        dashboard = self.env['smart.reorder.data.health.dashboard'].create({
            'company_id': self.company.id,
        })
        self.assertGreaterEqual(dashboard.missing_cost_count, 1)
        self.assertGreaterEqual(dashboard.missing_part_number_count, 1)

    def test_counts_no_vendor_and_placeholder_only_vendor(self):
        temp_vendor = self.env['res.partner'].create({'name': 'Health Temp Vendor'})
        self.config.write({'temp_vendor_ids': [(6, 0, [temp_vendor.id])]})
        real_vendor = self.env['res.partner'].create({'name': 'Health Real Vendor'})

        no_vendor_product = self.env['product.product'].create({
            'name': 'Health No Vendor Widget', 'type': 'product', 'standard_price': 5.0,
        })
        temp_vendor_product = self.env['product.product'].create({
            'name': 'Health Temp Vendor Product', 'type': 'product', 'standard_price': 5.0,
        })
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': temp_vendor_product.product_tmpl_id.id,
            'partner_id': temp_vendor.id,
        })
        real_vendor_product = self.env['product.product'].create({
            'name': 'Health Real Vendor Product', 'type': 'product', 'standard_price': 5.0,
        })
        self.env['product.supplierinfo'].create({
            'product_tmpl_id': real_vendor_product.product_tmpl_id.id,
            'partner_id': real_vendor.id,
        })

        dashboard = self.env['smart.reorder.data.health.dashboard'].create({
            'company_id': self.company.id,
        })
        self.assertGreaterEqual(dashboard.missing_vendor_count, 1)
        self.assertGreaterEqual(dashboard.temp_vendor_only_count, 1)

        action = dashboard.action_view_missing_vendor()
        self.assertIn(no_vendor_product.product_tmpl_id.id, action['domain'][0][2])
        self.assertNotIn(real_vendor_product.product_tmpl_id.id, action['domain'][0][2])

    def test_excluded_products_not_counted(self):
        excluded_product = self.env['product.product'].create({
            'name': 'Health Excluded Widget', 'type': 'product',
            'standard_price': 0.0, 'exclude_from_reorder_advisor': True,
        })
        dashboard = self.env['smart.reorder.data.health.dashboard'].create({
            'company_id': self.company.id,
        })
        action = dashboard.action_view_missing_cost()
        matches = self.env['product.template'].search(action['domain'])
        self.assertNotIn(
            excluded_product.product_tmpl_id, matches,
            'a product explicitly excluded from the advisor must not appear as a data-quality gap'
        )

    def test_action_open_dashboard_creates_and_returns_record(self):
        action = self.env['smart.reorder.data.health.dashboard'].action_open_dashboard()
        self.assertEqual(action['res_model'], 'smart.reorder.data.health.dashboard')
        self.assertTrue(action['res_id'])


@tagged('post_install', '-at_install')
class TestCostFieldVisibility(TransactionCase):
    """Recommendation: cost/pricing fields must be gated behind a Cost Viewer
    group (or Manager, which always implies it) — enforced at the ORM level
    via fields_get(), not just hidden in a view, and the export must respect
    the same boundary explicitly."""

    COST_FIELDS = [
        'product_cost', 'vendor_price', 'last_purchase_cost', 'effective_unit_cost',
        'has_price_discrepancy', 'price_discrepancy_pct', 'reorder_value',
        'estimated_purchase_value',
    ]

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search(
            [('company_id', '=', cls.company.id)], limit=1
        )
        if not cls.warehouse:
            cls.warehouse = cls.env['stock.warehouse'].create({
                'name': 'Cost Visibility Test WH', 'code': 'CVWH', 'company_id': cls.company.id,
            })
        product = cls.env['product.product'].create({
            'name': 'Cost Visibility Widget', 'type': 'product', 'standard_price': 5.0,
        })
        cls.suggestion = cls.env['smart.reorder.suggestion'].create({
            'company_id': cls.company.id, 'warehouse_id': cls.warehouse.id,
            'product_id': product.id, 'urgency': 'normal',
            'product_cost': 5.0, 'vendor_price': 6.0,
            'last_purchase_cost': 4.5, 'effective_unit_cost': 4.5,
            'suggested_reorder_qty': 3.0,
        })

        user_group = cls.env.ref('smart_reorder_advisor.group_smart_reorder_user')
        cost_viewer_group = cls.env.ref('smart_reorder_advisor.group_smart_reorder_cost_viewer')
        manager_group = cls.env.ref('smart_reorder_advisor.group_smart_reorder_manager')

        cls.plain_user = cls.env['res.users'].create({
            'name': 'Cost Visibility Plain User',
            'login': 'cost_visibility_plain_user_test',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id, user_group.id])],
        })
        cls.cost_viewer_user = cls.env['res.users'].create({
            'name': 'Cost Visibility Cost Viewer',
            'login': 'cost_visibility_cost_viewer_test',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id, cost_viewer_group.id])],
        })
        cls.manager_user = cls.env['res.users'].create({
            'name': 'Cost Visibility Manager',
            'login': 'cost_visibility_manager_test',
            'groups_id': [(6, 0, [cls.env.ref('base.group_user').id, manager_group.id])],
        })

    def test_plain_user_cannot_see_cost_fields(self):
        fields_info = self.env['smart.reorder.suggestion'].with_user(
            self.plain_user
        ).fields_get(self.COST_FIELDS)
        self.assertEqual(fields_info, {}, 'a plain Reorder User must not see any cost field metadata')

    def test_cost_viewer_sees_cost_fields_without_being_a_manager(self):
        fields_info = self.env['smart.reorder.suggestion'].with_user(
            self.cost_viewer_user
        ).fields_get(self.COST_FIELDS)
        self.assertEqual(set(fields_info.keys()), set(self.COST_FIELDS))
        # Confirm this user still isn't a manager (Cost Viewer is a separate axis).
        self.assertFalse(self.cost_viewer_user.has_group('smart_reorder_advisor.group_smart_reorder_manager'))

    def test_manager_sees_cost_fields_via_implied_group(self):
        fields_info = self.env['smart.reorder.suggestion'].with_user(
            self.manager_user
        ).fields_get(self.COST_FIELDS)
        self.assertEqual(set(fields_info.keys()), set(self.COST_FIELDS))
        self.assertTrue(self.manager_user.has_group('smart_reorder_advisor.group_smart_reorder_cost_viewer'))

    def test_export_redacts_cost_for_plain_user(self):
        wizard = self.env['smart.reorder.export.wizard'].with_user(self.plain_user).with_context(
            active_ids=[self.suggestion.id]
        ).create({'export_format': 'full'})
        xlsx_bytes = wizard._build_xlsx(self.suggestion)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(data_row[31], '— (no access)')  # Unit Cost (col 32, 0-indexed 31)
        self.assertEqual(data_row[42], '— (no access)')  # Last Purchase Cost

    def test_export_shows_real_cost_for_cost_viewer(self):
        wizard = self.env['smart.reorder.export.wizard'].with_user(self.cost_viewer_user).with_context(
            active_ids=[self.suggestion.id]
        ).create({'export_format': 'full'})
        xlsx_bytes = wizard._build_xlsx(self.suggestion)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(data_row[31], 5.0)
        self.assertEqual(data_row[42], 4.5)


@tagged('post_install', '-at_install')
class TestReorderConfigConstraints(TransactionCase):

    def test_negative_constraints(self):
        company = self.env.company
        # Create a new config bypass constraint checks on creation or write
        config = self.env['smart.reorder.config'].create({
            'company_id': company.id,
        })
        
        # Test negative transfer_surplus_threshold raises ValidationError
        with self.assertRaises(ValidationError):
            config.write({'transfer_surplus_threshold': -1.0})
            
        # Test negative default_internal_transfer_days raises ValidationError
        with self.assertRaises(ValidationError):
            config.write({'default_internal_transfer_days': -5})

    def test_spike_dominance_pct_must_be_between_0_and_100(self):
        company = self.env.company
        config = self.env['smart.reorder.config'].create({'company_id': company.id})
        with self.assertRaises(ValidationError):
            config.write({'spike_dominance_pct': 0.0})
        with self.assertRaises(ValidationError):
            config.write({'spike_dominance_pct': 150.0})
        with self.assertRaises(ValidationError):
            config.write({'spike_dominance_pct': -10.0})
        config.write({'spike_dominance_pct': 60.0})  # must not raise
        self.assertEqual(config.spike_dominance_pct, 60.0)

    def test_spike_multiplier_must_be_at_least_one(self):
        company = self.env.company
        config = self.env['smart.reorder.config'].create({'company_id': company.id})
        with self.assertRaises(ValidationError):
            config.write({'spike_multiplier': 0.5})
        config.write({'spike_multiplier': 3.0})  # must not raise
        self.assertEqual(config.spike_multiplier, 3.0)


@tagged('post_install', '-at_install')
class TestReorderRound2Regressions(TransactionCase):

    def test_warehouse_failure_is_isolated(self):
        # 1. Company with two warehouses
        company = self.env.company
        wh_a = self.env['stock.warehouse'].create({
            'name': 'Warehouse A',
            'code': 'WHA',
            'company_id': company.id,
        })
        wh_b = self.env['stock.warehouse'].create({
            'name': 'Warehouse B',
            'code': 'WHB',
            'company_id': company.id,
        })
        
        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
            })

        product = self.env['product.product'].create({
            'name': 'Test Regression Product',
            'type': 'product',
        })

        # Pre-create a suggestion for WH_B so we can test that it gets marked as is_stale=True
        self.env['smart.reorder.suggestion'].create({
            'company_id': company.id,
            'warehouse_id': wh_b.id,
            'product_id': product.id,
            'active': True,
            'is_stale': False,
        })

        # 2. Patch _fetch_warehouse_data to raise error for WH_B only
        orig_fetch = self.env['smart.reorder.suggestion']._fetch_warehouse_data
        
        def side_effect(company, config, warehouse, *args, **kwargs):
            if warehouse.id == wh_b.id:
                raise RuntimeError("Simulated failure")
            return orig_fetch(company, config, warehouse, *args, **kwargs)

        with patch.object(type(self.env['smart.reorder.suggestion']), '_fetch_warehouse_data', side_effect):
            # 3. Run suggestions generation
            self.env['smart.reorder.suggestion'].generate_suggestions(
                company_ids=[company.id],
                warehouse_ids=[wh_a.id, wh_b.id],
                trigger_type='manual'
            )

        # 4. Assert isolation
        sugg_b = self.env['smart.reorder.suggestion'].search([
            ('company_id', '=', company.id),
            ('warehouse_id', '=', wh_b.id),
            ('product_id', '=', product.id),
        ], limit=1)
        self.assertTrue(sugg_b)
        self.assertTrue(sugg_b.is_stale)
        self.assertIn("Simulated failure", sugg_b.stale_reason)

        # Cron log check
        log = self.env['smart.reorder.cron.log'].search([
            ('company_id', '=', company.id),
        ], order='started_at desc', limit=1)
        self.assertTrue(log)
        self.assertEqual(log.status, 'completed_with_errors')
        self.assertEqual(log.error_count, 1)
        self.assertIn("Warehouse B", log.error_notes)
        self.assertIn("Simulated failure", log.error_notes)

    def test_snapshot_retention_purge(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company.id)], limit=1)
        if not warehouse:
            warehouse = self.env['stock.warehouse'].create({
                'name': 'Test WH',
                'code': 'TWH',
                'company_id': company.id,
            })
        product = self.env['product.product'].create({
            'name': 'Test Snapshot Product',
            'type': 'product',
        })

        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'snapshot_retention_months': 2,
            })
        else:
            config.write({'snapshot_retention_months': 2})

        old_date = fields.Date.today() - relativedelta(months=3)
        recent_date = fields.Date.today()

        old_snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': old_date,
            'lead_time_days': 30,
        })
        recent_snap = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': recent_date,
            'lead_time_days': 30,
        })

        self.env['smart.reorder.forecast.snapshot']._score_snapshots()

        self.assertFalse(old_snap.exists())
        self.assertTrue(recent_snap.exists())

        config.write({'snapshot_retention_months': 0})
        old_snap_2 = self.env['smart.reorder.forecast.snapshot'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'snapshot_date': old_date,
            'lead_time_days': 30,
        })
        self.env['smart.reorder.forecast.snapshot']._score_snapshots()
        self.assertTrue(old_snap_2.exists())

    def test_receipt_brings_stock_above_min_below_max(self):
        company = self.env.company
        warehouse = self.env['stock.warehouse'].create({
            'name': 'Test Receipt WH',
            'code': 'TRW',
            'company_id': company.id,
        })
        product = self.env['product.product'].create({
            'name': 'Test Receipt Product',
            'type': 'product',
        })
        
        config = self.env['smart.reorder.config'].search([('company_id', '=', company.id)], limit=1)
        if not config:
            config = self.env['smart.reorder.config'].create({
                'company_id': company.id,
                'default_lead_time_months': 1.0,
                'safety_buffer_months': 1.0,
                'order_cycle_months': 2.0,
            })
        else:
            config.write({
                'default_lead_time_months': 1.0,
                'safety_buffer_months': 1.0,
                'order_cycle_months': 2.0,
            })

        sug = self.env['smart.reorder.suggestion'].create({
            'company_id': company.id,
            'warehouse_id': warehouse.id,
            'product_id': product.id,
            'avg_monthly_demand': 10.0,
            'lead_time_months': 1.0,
            'safety_buffer_months': 1.0,
            'order_cycle_months': 2.0,
            'moq': 1.0,
            'qty_on_hand': 10.0,
            'qty_incoming': 0.0,
            'qty_outgoing': 0.0,
            'reorder_needed': True,
        })

        picking_type = self.env['stock.picking.type'].search([
            ('warehouse_id', '=', warehouse.id),
            ('code', '=', 'incoming'),
        ], limit=1)
        if not picking_type:
            picking_type = self.env['stock.picking.type'].create({
                'name': 'Receipts',
                'code': 'incoming',
                'sequence_code': 'IN',
                'warehouse_id': warehouse.id,
                'default_location_dest_id': warehouse.lot_stock_id.id,
                'company_id': company.id,
            })
        
        self.env['stock.quant']._update_available_quantity(
            product, warehouse.lot_stock_id, 10.0
        )
        
        picking = self.env['stock.picking'].create({
            'picking_type_id': picking_type.id,
            'location_id': self.env.ref('stock.stock_location_suppliers').id,
            'location_dest_id': warehouse.lot_stock_id.id,
            'company_id': company.id,
        })
        self.env['stock.move'].create({
            'name': 'Test receipt move',
            'product_id': product.id,
            'product_uom_qty': 15.0,
            'product_uom': product.uom_id.id,
            'location_id': self.env.ref('stock.stock_location_suppliers').id,
            'location_dest_id': warehouse.lot_stock_id.id,
            'picking_id': picking.id,
        })
        
        picking.action_confirm()
        picking.action_assign()
        
        for move in picking.move_ids:
            move.quantity = 15.0
        
        picking.button_validate()

        sug.invalidate_recordset(['qty_on_hand', 'qty_available', 'reorder_needed'])
        self.assertEqual(sug.qty_on_hand, 25.0)
        self.assertEqual(sug.qty_available, 25.0)
        self.assertTrue(sug.reorder_needed)

        picking2 = self.env['stock.picking'].create({
            'picking_type_id': picking_type.id,
            'location_id': self.env.ref('stock.stock_location_suppliers').id,
            'location_dest_id': warehouse.lot_stock_id.id,
            'company_id': company.id,
        })
        self.env['stock.move'].create({
            'name': 'Test receipt move 2',
            'product_id': product.id,
            'product_uom_qty': 20.0,
            'product_uom': product.uom_id.id,
            'location_id': self.env.ref('stock.stock_location_suppliers').id,
            'location_dest_id': warehouse.lot_stock_id.id,
            'picking_id': picking2.id,
        })
        
        picking2.action_confirm()
        picking2.action_assign()
        for move in picking2.move_ids:
            move.quantity = 20.0
        picking2.button_validate()

        sug.invalidate_recordset(['qty_on_hand', 'qty_available', 'reorder_needed'])
        self.assertEqual(sug.qty_on_hand, 45.0)
        self.assertEqual(sug.qty_available, 45.0)
        self.assertFalse(sug.reorder_needed)




