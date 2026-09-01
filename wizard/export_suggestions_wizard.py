import base64
import io

from odoo import models, fields, api, _
from odoo.exceptions import UserError

URGENCY_LABELS = {
    'critical': 'Critical — Negative Stock',
    'urgent':   'Urgent — Below Lead Time',
    'normal':   'Normal — Reorder Recommended',
    'dead':     'Dead Stock — No Movement',
    'ok':       'OK — Sufficient Stock',
}
TREND_LABELS = {
    'up':     'Rising',
    'stable': 'Stable',
    'down':   'Falling',
    'new':    'New / No History',
}
PATTERN_LABELS = {
    'regular': 'Sells Regularly',
    'sometimes': 'Sells Sometimes',
    'big_order_mixed': 'Big Order Mixed In',
    'one_time_big_order': 'One-Time Big Order Only',
    'new': 'New — No Sales History',
}
RESOLUTION_LABELS = {
    'reorder': 'Genuine Shortfall — Reorder',
    'stock_correction': 'Likely Data Error — Consider Adjustment',
}

EXPORT_HEADERS = [
    'Budget Rank',
    'Stale?',
    'Review?',
    'Urgency',
    'ABC Class',
    'Demand Trend',
    'Sales Pattern',
    'Company',
    'Warehouse',
    'Analysis Date',
    'Part Number',
    'Product Name',
    'Superseded By',
    'Category',
    'On Hand Qty',
    'Incoming Qty',
    'Outgoing Qty',
    'Net Available Qty',
    'Months of Stock',
    'Months of Stock After Order',
    'Overstocked?',
    'Avg Monthly Demand',
    'Lead Time (Months)',
    'MOQ',
    'Min Stock Level',
    'Max Stock Level',
    'Raw Reorder Qty',
    'Suggested Reorder Qty',
    'Prior Suggested Qty',
    'Change % vs Last Run',
    'Confidence (%)',
    'Unit Cost',
    'Reorder Value',
    'Dead Stock?',
    'Months Since Last Sale',
    'Within Budget?',
    'Provisional?',
    'Vendor',
    'Transfer Source Warehouse',
    'Transfer Qty',
    'Transfer Lead Time (Days)',
    'Draft PO Ref',
    'Last Purchase Cost',
    'Effective Unit Cost',
    'Price Discrepancy?',
    'Price Discrepancy (%)',
]

# Task 9: "Essential (Working List)" — the small column set a buyer actually
# works from day to day, as opposed to the full technical breakdown above.
ESSENTIAL_HEADERS = [
    'Part Number',
    'Product Name/Description',
    'Warehouse',
    'On Hand Qty',
    'Suggested Reorder Qty',
    'Vendor',
    'Unit Cost',
    'Reorder Value',
    'Dead Stock?',
    'Last Sale Date',
    'Suggested Resolution',
]


class ExportSuggestionsWizard(models.TransientModel):
    """
    Exports the currently filtered/selected smart.reorder.suggestion list to
    an Excel file — triggered from a button in the list view's header, which
    Odoo passes either active_domain (current search filter, no row selected)
    or active_ids (rows explicitly selected) in context.
    """
    _name = 'smart.reorder.export.wizard'
    _description = 'Export Reorder Suggestions to Excel'

    record_count = fields.Integer(string='Suggestions to Export', readonly=True)

    export_format = fields.Selection([
        ('essential', 'Essential (Working List)'),
        ('full', 'Full (All Columns)'),
    ], string='Export Format', default='essential', required=True,
       help='Essential: the short working-list column set (product, warehouse, '
            'quantities, vendor, cost, dead-stock, last sale). '
            'Full: every column on the suggestion, including analysis internals.')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if 'record_count' in fields_list:
            res['record_count'] = self.env['smart.reorder.suggestion'].search_count(
                self._get_export_domain()
            )
        return res

    def _get_export_domain(self):
        """Mirrors how the list view's header button was triggered: a domain
        (search filter applied, nothing selected), explicit row selection, or
        — if invoked outside that context entirely — everything."""
        domain = self.env.context.get('active_domain')
        if domain is not None:
            return domain
        active_ids = self.env.context.get('active_ids')
        if active_ids:
            return [('id', 'in', active_ids)]
        return []

    def action_export(self):
        self.ensure_one()
        suggestions = self.env['smart.reorder.suggestion'].search(self._get_export_domain())
        if not suggestions:
            raise UserError(_('No suggestions match the current filter — nothing to export.'))

        xlsx_data = self._build_xlsx(suggestions)
        name_suffix = 'Essential' if self.export_format == 'essential' else 'Full'
        attachment = self.env['ir.attachment'].create({
            'name': f'Reorder_Suggestions_{name_suffix}_{fields.Date.today()}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _build_xlsx(self, suggestions):
        """Pure data → bytes — no DB writes, easy to unit test directly.
        Dispatches on export_format; defaults to 'essential' when called
        without a wizard record bound (e.g. directly in a test)."""
        export_format = self.export_format if self and self.export_format else 'essential'
        if export_format == 'full':
            return self._build_full_xlsx(suggestions)
        return self._build_essential_xlsx(suggestions)

    @api.model
    def _new_workbook(self, sheet_title, headers):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise UserError(_(
                'The Python package "openpyxl" is required for Excel export '
                'but is not installed on this server.'
            ))

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        return wb, ws

    @api.model
    def _finalize_sheet(self, wb, ws, headers):
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = max(14, len(headers[col - 1]) + 2)
        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _user_can_see_cost(self):
        """Recommendation: field-level `groups=` on the model already blocks
        cost data in list/form views; this makes the export respect the same
        boundary explicitly and predictably, rather than relying on ORM
        edge-case behavior for programmatic field access."""
        return (
            self.env.user.has_group('smart_reorder_advisor.group_smart_reorder_cost_viewer')
            or self.env.user.has_group('smart_reorder_advisor.group_smart_reorder_manager')
        )

    @api.model
    def _build_essential_xlsx(self, suggestions):
        wb, ws = self._new_workbook('Reorder Suggestions (Essential)', ESSENTIAL_HEADERS)
        can_see_cost = self._user_can_see_cost()
        redacted = '— (no access)'

        for row, rec in enumerate(suggestions, start=2):
            ws.cell(row=row, column=1, value=rec.product_id.default_code or '')
            ws.cell(row=row, column=2, value=rec.product_id.name or '')
            ws.cell(row=row, column=3, value=rec.warehouse_id.name or '')
            ws.cell(row=row, column=4, value=rec.qty_on_hand)
            ws.cell(row=row, column=5, value=rec.suggested_reorder_qty)
            ws.cell(row=row, column=6, value=rec.vendor_id.name or '')
            ws.cell(row=row, column=7, value=rec.product_cost if can_see_cost else redacted)
            ws.cell(row=row, column=8, value=rec.reorder_value if can_see_cost else redacted)
            ws.cell(row=row, column=9, value='Yes' if rec.is_dead_stock else 'No')
            ws.cell(row=row, column=10, value=rec.last_sale_date or None)
            ws.cell(row=row, column=11, value=RESOLUTION_LABELS.get(rec.suggested_resolution, ''))

        return self._finalize_sheet(wb, ws, ESSENTIAL_HEADERS)

    @api.model
    def _build_full_xlsx(self, suggestions):
        wb, ws = self._new_workbook('Reorder Suggestions', EXPORT_HEADERS)
        can_see_cost = self._user_can_see_cost()
        redacted = '— (no access)'

        for row, rec in enumerate(suggestions, start=2):
            ws.cell(row=row, column=1,  value=rec.budget_rank)
            ws.cell(row=row, column=2,  value='Yes' if rec.is_stale else 'No')
            ws.cell(row=row, column=3,  value='Yes' if rec.needs_review else 'No')
            ws.cell(row=row, column=4,  value=URGENCY_LABELS.get(rec.urgency, rec.urgency or ''))
            ws.cell(row=row, column=5,  value=rec.abc_class or '')
            ws.cell(row=row, column=6,  value=TREND_LABELS.get(rec.demand_trend, rec.demand_trend or ''))
            ws.cell(row=row, column=7,  value=PATTERN_LABELS.get(rec.sales_pattern, rec.sales_pattern or ''))
            ws.cell(row=row, column=8,  value=rec.company_id.name or '')
            ws.cell(row=row, column=9,  value=rec.warehouse_id.name or '')
            ws.cell(row=row, column=10, value=rec.analysis_date or None)
            ws.cell(row=row, column=11, value=rec.default_code or '')
            ws.cell(row=row, column=12, value=rec.product_id.display_name or '')
            ws.cell(row=row, column=13, value=rec.superseded_by_id.display_name or '')
            ws.cell(row=row, column=14, value=rec.product_categ_id.name or '')
            ws.cell(row=row, column=15, value=rec.qty_on_hand)
            ws.cell(row=row, column=16, value=rec.qty_incoming)
            ws.cell(row=row, column=17, value=rec.qty_outgoing)
            ws.cell(row=row, column=18, value=rec.qty_available)
            ws.cell(row=row, column=19, value=rec.months_of_stock)
            ws.cell(row=row, column=20, value=rec.months_of_stock_after_order)
            ws.cell(row=row, column=21, value='Yes' if rec.is_overstocked else 'No')
            ws.cell(row=row, column=22, value=rec.avg_monthly_demand)
            ws.cell(row=row, column=23, value=rec.lead_time_months)
            ws.cell(row=row, column=24, value=rec.moq)
            ws.cell(row=row, column=25, value=rec.min_stock_level)
            ws.cell(row=row, column=26, value=rec.max_stock_level)
            ws.cell(row=row, column=27, value=rec.raw_reorder_qty)
            ws.cell(row=row, column=28, value=rec.suggested_reorder_qty)
            ws.cell(row=row, column=29, value=rec.prior_suggested_qty)
            ws.cell(row=row, column=30, value=rec.delta_pct)
            ws.cell(row=row, column=31, value=rec.confidence)
            ws.cell(row=row, column=32, value=rec.product_cost if can_see_cost else redacted)
            ws.cell(row=row, column=33, value=rec.reorder_value if can_see_cost else redacted)
            ws.cell(row=row, column=34, value='Yes' if rec.is_dead_stock else 'No')
            ws.cell(row=row, column=35, value=rec.months_since_last_sale)
            ws.cell(row=row, column=36, value='Yes' if rec.within_budget else 'No')
            ws.cell(row=row, column=37, value='Yes' if rec.is_provisional else 'No')
            ws.cell(row=row, column=38, value=rec.vendor_id.name or '')
            ws.cell(row=row, column=39, value=rec.transfer_source_warehouse_id.name or '')
            ws.cell(row=row, column=40, value=rec.transfer_suggested_qty)
            ws.cell(row=row, column=41, value=rec.transfer_lead_time_days)
            ws.cell(row=row, column=42, value=rec.draft_po_ref or '')
            ws.cell(row=row, column=43, value=rec.last_purchase_cost if can_see_cost else redacted)
            ws.cell(row=row, column=44, value=rec.effective_unit_cost if can_see_cost else redacted)
            ws.cell(row=row, column=45, value=('Yes' if rec.has_price_discrepancy else 'No') if can_see_cost else redacted)
            ws.cell(row=row, column=46, value=rec.price_discrepancy_pct if can_see_cost else redacted)

        return self._finalize_sheet(wb, ws, EXPORT_HEADERS)
